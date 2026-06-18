from __future__ import annotations

from collections.abc import Iterable
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from leadgen.utils.locale_text import normalize_lang

if TYPE_CHECKING:
    from leadgen.db.models import Lead


# Column header per language + column width. Localised by the
# exporting user's ``language_code`` (ru by default).
COLUMNS: list[tuple[dict[str, str], int]] = [
    ({"ru": "Название", "uk": "Назва", "en": "Name"}, 36),
    ({"ru": "AI-скор", "uk": "AI-скор", "en": "AI score"}, 10),
    ({"ru": "Теги", "uk": "Теги", "en": "Tags"}, 18),
    ({"ru": "Резюме", "uk": "Резюме", "en": "Summary"}, 40),
    (
        {
            "ru": "Совет: как зайти",
            "uk": "Порада: як зайти",
            "en": "Advice: how to approach",
        },
        50,
    ),
    (
        {"ru": "Сильные стороны", "uk": "Сильні сторони", "en": "Strengths"},
        35,
    ),
    (
        {
            "ru": "Точки роста / слабые",
            "uk": "Точки зростання / слабкі",
            "en": "Growth points / weaknesses",
        },
        35,
    ),
    ({"ru": "Риски", "uk": "Ризики", "en": "Risks"}, 30),
    ({"ru": "Категория", "uk": "Категорія", "en": "Category"}, 22),
    ({"ru": "Телефон", "uk": "Телефон", "en": "Phone"}, 18),
    ({"ru": "Сайт", "uk": "Сайт", "en": "Website"}, 32),
    ({"ru": "Соцсети", "uk": "Соцмережі", "en": "Social media"}, 24),
    ({"ru": "Адрес", "uk": "Адреса", "en": "Address"}, 45),
    (
        {
            "ru": "Рейтинг Google",
            "uk": "Рейтинг Google",
            "en": "Google rating",
        },
        12,
    ),
    ({"ru": "Отзывов", "uk": "Відгуків", "en": "Reviews"}, 10),
    (
        {
            "ru": "Отзывы (кратко)",
            "uk": "Відгуки (стисло)",
            "en": "Reviews (brief)",
        },
        45,
    ),
    ({"ru": "Широта", "uk": "Широта", "en": "Latitude"}, 12),
    ({"ru": "Долгота", "uk": "Довгота", "en": "Longitude"}, 12),
    ({"ru": "Источник", "uk": "Джерело", "en": "Source"}, 14),
]

HEADER_FILL = PatternFill(start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")


def build_excel(leads: Iterable[Lead], lang: str | None = None) -> bytes:
    """Render a list of leads into an XLSX file and return its bytes.

    ``lang`` is the exporting user's ``language_code`` and controls
    the header row language (ru / uk / en, ru default).
    """
    lang = normalize_lang(lang)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Leads"

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col_idx, (titles, width) in enumerate(COLUMNS, start=1):
        title = titles.get(lang, titles["ru"])
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    body_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, lead in enumerate(leads, start=2):
        ws.cell(row=row_idx, column=1, value=lead.name)
        ws.cell(
            row=row_idx,
            column=2,
            value=int(lead.score_ai) if lead.score_ai is not None else None,
        )
        ws.cell(row=row_idx, column=3, value=", ".join(lead.tags or []))
        ws.cell(row=row_idx, column=4, value=lead.summary)
        ws.cell(row=row_idx, column=5, value=lead.advice)
        ws.cell(row=row_idx, column=6, value="\n".join(lead.strengths or []))
        ws.cell(row=row_idx, column=7, value="\n".join(lead.weaknesses or []))
        ws.cell(row=row_idx, column=8, value="\n".join(lead.red_flags or []))
        ws.cell(row=row_idx, column=9, value=lead.category)
        ws.cell(row=row_idx, column=10, value=lead.phone)
        ws.cell(row=row_idx, column=11, value=lead.website)
        ws.cell(
            row=row_idx,
            column=12,
            value=", ".join((lead.social_links or {}).keys()),
        )
        ws.cell(row=row_idx, column=13, value=lead.address)
        ws.cell(row=row_idx, column=14, value=lead.rating)
        ws.cell(row=row_idx, column=15, value=lead.reviews_count)
        ws.cell(row=row_idx, column=16, value=lead.reviews_summary)
        ws.cell(row=row_idx, column=17, value=lead.latitude)
        ws.cell(row=row_idx, column=18, value=lead.longitude)
        ws.cell(row=row_idx, column=19, value=lead.source)

        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = body_align

    ws.freeze_panes = "C2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
