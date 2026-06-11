"""FastAPI application factory for the web frontend.

Listens on ``PORT`` (default 8080), serves ``/health``, ``/metrics``
and ``/api/v1/*``. This is the only delivery surface the product has
since the Telegram bot was removed.

Auth note: real email + password auth is in place. ``WEB_API_KEY``
still gates the SSE progress stream (since that's the only endpoint
where the client can't retry).
"""

from __future__ import annotations

import asyncio
import io
import json
import logging
import os
import secrets
import uuid
from collections.abc import AsyncIterator
from contextlib import asynccontextmanager, suppress
from datetime import datetime, timedelta, timezone
from typing import Any

import sqlalchemy as sa
from argon2 import PasswordHasher
from argon2.exceptions import VerifyMismatchError
from fastapi import (
    BackgroundTasks,
    Depends,
    FastAPI,
    HTTPException,
    Query,
    Request,
    UploadFile,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import (
    PlainTextResponse,
    RedirectResponse,
    Response,
    StreamingResponse,
)
from prometheus_client import CONTENT_TYPE_LATEST, REGISTRY, generate_latest
from sqlalchemy import func, select, update

from leadgen.adapters.web_api.auth import (
    get_current_user,
    request_ip,
)
from leadgen.adapters.web_api.csrf import CsrfMiddleware
from leadgen.adapters.web_api.schemas import (
    AffiliateCodeCreateRequest,
    AffiliateCodeSchema,
    AffiliateCodeUpdate,
    AffiliateOverview,
    BulkDraftEmailItem,
    BulkDraftEmailRequest,
    BulkDraftEmailResponse,
    BulkSendRequest,
    CityEntryResponse,
    CityListResponse,
    DashboardStats,
    GmailAuthorizeResponse,
    GmailIntegrationStatus,
    GmailSendRequest,
    GmailSendResponse,
    HealthResponse,
    HubspotAuthorizeResponse,
    HubspotExportItem,
    HubspotExportRequest,
    HubspotExportResponse,
    HubspotIntegrationStatus,
    LeadActivityListResponse,
    LeadBulkUpdateRequest,
    LeadBulkUpdateResponse,
    LeadCustomFieldsResponse,
    LeadCustomFieldUpsert,
    LeadEmailDraftRequest,
    LeadEmailDraftResponse,
    LeadListResponse,
    LeadMarkRequest,
    LeadResponse,
    LeadStatusCreate,
    LeadStatusListResponse,
    LeadStatusReorderRequest,
    LeadStatusSchema,
    LeadStatusUpdate,
    LeadTagSchema,
    LeadTaskCreate,
    LeadTaskListResponse,
    LeadTaskUpdate,
    LeadUpdate,
    NicheTaxonomyEntry,
    NicheTaxonomyResponse,
    NotionAuthorizeResponse,
    NotionConnectRequest,
    NotionDatabase,
    NotionDatabaseList,
    NotionExportItem,
    NotionExportRequest,
    NotionExportResponse,
    NotionIntegrationStatus,
    NotionSetDatabaseRequest,
    OutlookAuthorizeResponse,
    OutlookIntegrationStatus,
    PendingAction,
    PipedriveAuthorizeResponse,
    PipedriveConfigUpdate,
    PipedriveExportItem,
    PipedriveExportRequest,
    PipedriveExportResponse,
    PipedriveIntegrationStatus,
    PipedrivePipelinesResponse,
    PipedrivePipelineView,
    PipedriveStageView,
    PriorTeamSearch,
    SavedSearchCreate,
    SavedSearchListResponse,
    SavedSearchSchema,
    SavedSearchUpdate,
    SearchCreateResponse,
    SearchSummary,
    TeamAnalytics,
    TeamAnalyticsMemberBucket,
    TeamAnalyticsNicheBucket,
    TeamAnalyticsSourceBucket,
    TeamAnalyticsStatusBucket,
    TeamAnalyticsTimepoint,
    TeamDetailResponse,
    TeamMemberResponse,
    TeamMemberSummary,
    UserProfile,
)
from leadgen.adapters.web_api.schemas import (
    LeadActivity as LeadActivitySchema,
)
from leadgen.adapters.web_api.schemas import (
    LeadCustomField as LeadCustomFieldSchema,
)
from leadgen.adapters.web_api.schemas import (
    LeadTask as LeadTaskSchema,
)
from leadgen.adapters.web_api.sinks import WebDeliverySink
from leadgen.analysis.ai_analyzer import AIAnalyzer
from leadgen.config import assert_production_secrets, get_settings
from leadgen.core.services import (
    default_broker,
    mask_email,
    render_verification_email,
    send_email,
)
from leadgen.core.services.assistant_memory import (
    prune_old,
    record_memory,
)
from leadgen.core.services.progress_broker import BrokerProgressSink
from leadgen.core.services.team_permissions import (
    ROLE_OWNER as _ROLE_OWNER,
)
from leadgen.core.services.team_permissions import (
    can_manage_members as _can_manage_members,
)
from leadgen.core.services.team_permissions import (
    normalize_role as _normalize_role,
)
from leadgen.core.services.webhooks import (
    emit_event_sync as emit_webhook_event_sync,
)
from leadgen.core.services.webhooks import (
    serialize_lead as serialize_lead_for_webhook,
)
from leadgen.db.models import (
    AffiliateCode,
    EmailVerificationToken,
    Lead,
    LeadActivity,
    LeadCustomField,
    LeadMark,
    LeadStatus,
    LeadTag,
    LeadTagAssignment,
    LeadTask,
    OAuthCredential,
    Referral,
    SavedSearch,
    SearchQuery,
    Team,
    TeamInvite,
    TeamMembership,
    TeamSeenLead,
    User,
    UserAuditLog,
    UserIntegrationCredential,
    UserSeenLead,
)
from leadgen.db.session import session_factory
from leadgen.integrations.slack import send_slack_notification
from leadgen.pipeline.search import run_search_with_timeout
from leadgen.queue import enqueue_search, is_queue_enabled
from leadgen.utils import spawn
from leadgen.utils.locale_text import normalize_lang
from leadgen.utils.locale_text import pick as locale_pick

logger = logging.getLogger(__name__)


# Demo avatars for team page until seat management is wired up.
_DEMO_TEAM_COLORS = [
    "#3D5AFE",
    "#F59E0B",
    "#16A34A",
    "#EC4899",
    "#8B5CF6",
    "#06B6D4",
]


# Legacy hard-coded lead-status keys. Personal-mode searches still
# use these directly; team-mode searches resolve against the team's
# ``lead_statuses`` palette (which is seeded with the same five keys
# at team creation, so existing rows remain valid).
LEGACY_LEAD_STATUS_KEYS: frozenset[str] = frozenset(
    {"new", "contacted", "replied", "won", "archived"}
)


# Default lead-status palette lives in ``routes/_helpers.py`` (the
# single source of truth, localised per creating user). The aliases
# below keep the legacy ``app.py`` import path working for callers
# and tests that still reference the old private names.
from leadgen.adapters.web_api.routes._helpers import (  # noqa: E402, F401, I001
    _DEFAULT_LEAD_STATUSES,
    seed_default_lead_statuses as _seed_default_lead_statuses,
)


async def _bootstrap_admins() -> None:
    # Promote every email listed in BOOTSTRAP_ADMIN_EMAILS (comma-separated)
    # to platform admin on startup. Missing users are skipped silently — the
    # next boot after they register will pick them up. Lets non-technical
    # operators flip the flag from Railway's Variables UI without SQL.
    raw = os.environ.get("BOOTSTRAP_ADMIN_EMAILS", "").strip()
    if not raw:
        return
    emails = [e.strip().lower() for e in raw.split(",") if e.strip()]
    if not emails:
        return
    try:
        from leadgen.db.models import User
        from leadgen.db.session import get_session

        async with get_session() as session:
            result = await session.execute(
                update(User)
                .where(User.email.in_(emails), User.is_admin.is_(False))
                .values(is_admin=True)
                .returning(User.email)
            )
            promoted = [row[0] for row in result.all()]
            await session.commit()
        if promoted:
            logging.getLogger(__name__).info(
                "bootstrap_admin: promoted %s", ", ".join(promoted)
            )
    except Exception:
        logging.getLogger(__name__).exception("bootstrap_admin failed")


@asynccontextmanager
async def _lifespan(app: FastAPI) -> AsyncIterator[None]:
    await _bootstrap_admins()
    # In-process saved-search scheduler. Runs only when Redis is
    # absent — production deploys with arq run a separate worker that
    # does the same scan. Safe in dev because each scan completes in
    # a few ms when no rows are due. Toggle with SAVED_SEARCH_SCHEDULER=0.
    scheduler_task: asyncio.Task[None] | None = None
    if (
        not get_settings().redis_url
        and os.environ.get("SAVED_SEARCH_SCHEDULER", "1") == "1"
    ):
        scheduler_task = asyncio.create_task(
            _saved_search_scheduler_loop(),
            name="convioo-saved-search-scheduler",
        )
    try:
        yield
    finally:
        if scheduler_task is not None:
            scheduler_task.cancel()
            with suppress(asyncio.CancelledError, Exception):
                await scheduler_task


def create_app() -> FastAPI:
    # Fail fast on Railway when the security-critical secrets are
    # missing — better a crashed deploy than sessions signed with an
    # empty secret or OAuth tokens that reset on every restart.
    # No-ops outside Railway (local dev, pytest).
    assert_production_secrets()

    app = FastAPI(
        title="Convioo API",
        version="0.3.0",
        docs_url="/docs",
        redoc_url=None,
        lifespan=_lifespan,
    )

    cors = get_settings().web_cors_origins
    if cors:
        origins = [o.strip() for o in cors.split(",") if o.strip()]
        if any(o == "*" for o in origins):
            raise RuntimeError(
                "WEB_CORS_ORIGINS contains '*' which is incompatible "
                "with allow_credentials=True. Set an explicit allowlist."
            )
        # CsrfMiddleware sits inside CORSMiddleware so it runs *after*
        # the preflight OPTIONS handshake — preflights stay 200.
        app.add_middleware(CsrfMiddleware, allowed_origins=origins)
        app.add_middleware(
            CORSMiddleware,
            allow_origins=origins,
            allow_credentials=True,
            allow_methods=["*"],
            allow_headers=["*"],
        )

    @app.middleware("http")
    async def _security_headers(request: Request, call_next):
        response = await call_next(request)
        # The API only serves JSON / SSE / file downloads, no HTML. A
        # strict default-src 'none' policy means even an accidental
        # HTML response can't pull in third-party scripts / frames.
        response.headers.setdefault(
            "Content-Security-Policy",
            "default-src 'none'; frame-ancestors 'none'",
        )
        response.headers.setdefault("X-Content-Type-Options", "nosniff")
        response.headers.setdefault("Referrer-Policy", "no-referrer")
        response.headers.setdefault(
            "Strict-Transport-Security",
            "max-age=31536000; includeSubDomains",
        )
        return response

    @app.get("/", response_class=PlainTextResponse, include_in_schema=False)
    async def root() -> str:
        return "convioo alive. /health, /metrics and /api/v1/* available.\n"

    @app.get("/health", response_model=HealthResponse)
    async def health() -> HealthResponse:
        from leadgen.core.services.health_probes import probes_for_health

        probes = await probes_for_health()
        db_ok = bool(probes["db"])
        # Redis is optional. If REDIS_URL is unset the bot/web path
        # works without arq, so don't fail the overall status on it.
        # If REDIS_URL is set and the ping failed, surface unhealthy.
        redis_state = probes["redis"]
        redis_down = redis_state is False
        return HealthResponse(
            status="healthy" if db_ok and not redis_down else "unhealthy",
            db=db_ok,
            redis=redis_state,
            queue_depth=probes["queue_depth"],
            commit=(os.environ.get("RAILWAY_GIT_COMMIT_SHA", "unknown"))[:12],
        )

    @app.get("/metrics", include_in_schema=False)
    async def metrics() -> Response:
        payload = generate_latest(REGISTRY)
        return Response(
            content=payload,
            media_type=CONTENT_TYPE_LATEST.split(";")[0],
        )

    # /api/v1/auth/* and /api/v1/api-keys/* moved to routes/auth.py
    # /api/v1/webhooks moved to routes/webhooks.py


    # /api/v1/users/me/* moved to routes/users.py

    @app.post("/api/v1/users/me/icp-profile")
    async def upload_icp_profile(
        file: UploadFile,
        current_user: User = Depends(get_current_user),
    ) -> dict:
        """Upload a CSV of best clients → Claude extracts ICP → stored on user profile."""
        from leadgen.core.services.icp_analyzer import analyze_client_csv

        if not file.filename or not file.filename.endswith(".csv"):
            raise HTTPException(status_code=400, detail="Only CSV files are accepted")

        content = await file.read()
        try:
            csv_text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            csv_text = content.decode("latin-1")

        try:
            icp = await analyze_client_csv(csv_text)
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc

        async with session_factory() as session:
            user = await session.get(User, current_user.id)
            if user is None:
                raise HTTPException(status_code=404, detail="User not found")
            user.icp_profile = icp
            await session.commit()

        return {"icp_profile": icp}

    @app.get("/api/v1/users/me/icp-profile")
    async def get_icp_profile(
        current_user: User = Depends(get_current_user),
    ) -> dict:
        return {"icp_profile": current_user.icp_profile}

    # /api/v1/users/me PATCH moved to routes/users.py

    # /api/v1/teams/* moved to routes/teams.py

    # /api/v1/search/consult, /api/v1/assistant/*, decision-makers, import-csv, suggest-niches moved to routes/assistant.py

    # /api/v1/searches/* moved to routes/search.py


    @app.get("/api/v1/leads", response_model=LeadListResponse)
    async def list_all_leads(
        team_id: uuid.UUID | None = None,
        member_user_id: int | None = None,
        lead_status: str | None = None,
        temp: str | None = None,
        created_after: datetime | None = None,
        untouched_days: int | None = None,
        tag_id: uuid.UUID | None = None,
        archived: bool = False,
        limit: int = 200,
        current_user: User = Depends(get_current_user),
    ) -> LeadListResponse:
        """Cross-session CRM listing.

        Personal mode → caller's own leads. Team mode → caller's own
        leads inside that team by default. Team owners can pass
        ``member_user_id`` to inspect a specific teammate's CRM.

        Filter knobs the frontend's smart-filter chips lean on:
        - ``temp`` ∈ {"hot","warm","cold"} → filters by score buckets
          (hot ≥ 75, warm 50-74, cold < 50).
        - ``created_after`` → ISO timestamp; "новые сегодня" / "за неделю".
        - ``untouched_days`` → leads whose ``last_touched_at`` is older
          than N days (or never touched at all). "Без касания 14+ дней".

        ``mark_color`` on each row is always the *caller's* private
        mark (never the viewed-as user's), so an owner browsing a
        teammate's CRM still sees their own colour codes.
        """
        user_id = current_user.id
        limit = max(1, min(limit, 500))
        # ``archived`` flag splits the list into two zones — active
        # CRM (default) vs the Archive tab. Both still hide soft-deleted
        # rows, only ``archived_at`` flips between IS NULL / IS NOT NULL.
        archived_predicate = (
            Lead.archived_at.is_not(None) if archived else Lead.archived_at.is_(None)
        )
        async with session_factory() as session:
            stmt = (
                select(Lead, SearchQuery.niche, SearchQuery.region)
                .join(SearchQuery, SearchQuery.id == Lead.query_id)
                .where(SearchQuery.source == "web")
                .where(Lead.deleted_at.is_(None))
                .where(archived_predicate)
                .order_by(Lead.score_ai.desc().nullslast(), Lead.created_at.desc())
                .limit(limit)
            )
            total_stmt = (
                select(func.count(Lead.id))
                .join(SearchQuery, SearchQuery.id == Lead.query_id)
                .where(SearchQuery.source == "web")
                .where(Lead.deleted_at.is_(None))
                .where(archived_predicate)
            )
            if team_id is not None:
                target_user = await _resolve_team_view(
                    session, team_id, user_id, member_user_id
                )
                stmt = stmt.where(SearchQuery.team_id == team_id).where(
                    SearchQuery.user_id == target_user
                )
                total_stmt = total_stmt.where(
                    SearchQuery.team_id == team_id
                ).where(SearchQuery.user_id == target_user)
            else:
                stmt = stmt.where(SearchQuery.user_id == user_id).where(
                    SearchQuery.team_id.is_(None)
                )
                total_stmt = total_stmt.where(
                    SearchQuery.user_id == user_id
                ).where(SearchQuery.team_id.is_(None))
            if lead_status:
                stmt = stmt.where(Lead.lead_status == lead_status)
                total_stmt = total_stmt.where(Lead.lead_status == lead_status)
            if temp == "hot":
                stmt = stmt.where(Lead.score_ai >= 75)
                total_stmt = total_stmt.where(Lead.score_ai >= 75)
            elif temp == "warm":
                stmt = stmt.where(Lead.score_ai >= 50).where(Lead.score_ai < 75)
                total_stmt = total_stmt.where(Lead.score_ai >= 50).where(
                    Lead.score_ai < 75
                )
            elif temp == "cold":
                stmt = stmt.where(
                    (Lead.score_ai < 50) | (Lead.score_ai.is_(None))
                )
                total_stmt = total_stmt.where(
                    (Lead.score_ai < 50) | (Lead.score_ai.is_(None))
                )
            if created_after is not None:
                stmt = stmt.where(Lead.created_at >= created_after)
                total_stmt = total_stmt.where(Lead.created_at >= created_after)
            if untouched_days and untouched_days > 0:
                cutoff = datetime.now(timezone.utc) - timedelta(
                    days=untouched_days
                )
                stmt = stmt.where(
                    (Lead.last_touched_at < cutoff)
                    | (Lead.last_touched_at.is_(None))
                )
                total_stmt = total_stmt.where(
                    (Lead.last_touched_at < cutoff)
                    | (Lead.last_touched_at.is_(None))
                )
            if tag_id is not None:
                tagged_subq = (
                    select(LeadTagAssignment.lead_id)
                    .where(LeadTagAssignment.tag_id == tag_id)
                    .subquery()
                )
                stmt = stmt.where(Lead.id.in_(select(tagged_subq.c.lead_id)))
                total_stmt = total_stmt.where(
                    Lead.id.in_(select(tagged_subq.c.lead_id))
                )
            rows = (await session.execute(stmt)).all()

            lead_ids = [lead.id for lead, _n, _r in rows]
            total = int((await session.execute(total_stmt)).scalar() or 0)
            marks = await _marks_for_user(session, user_id, lead_ids)
            tags_by_lead = await _tags_by_lead(session, lead_ids)

        leads: list[LeadResponse] = []
        sessions_by_id: dict[str, dict[str, Any]] = {}
        for lead, niche, region in rows:
            leads.append(
                _to_lead_response(
                    lead, marks.get(lead.id), tags_by_lead.get(lead.id)
                )
            )
            sessions_by_id[str(lead.query_id)] = {"niche": niche, "region": region}
        return LeadListResponse(leads=leads, total=total, sessions_by_id=sessions_by_id)

    @app.get("/api/v1/leads/export.csv", include_in_schema=False)
    async def export_leads_csv(
        team_id: uuid.UUID | None = None,
        member_user_id: int | None = None,
        current_user: User = Depends(get_current_user),
    ) -> StreamingResponse:
        """Export the caller's CRM rows as a CSV file.

        Streamed in 500-row chunks so the response starts before the
        whole result set is in memory. Mirrors the same scoping as the
        JSON list endpoint (personal / team / view-as) but ignores the
        smart-filter knobs — export is always "everything in this
        scope" so the file is the complete copy.
        """
        user_id = current_user.id
        import csv as _csv
        import io as _io

        # Hand-rolled CSV — keeps the deps tight (no openpyxl/pandas in
        # the request path). Columns are intentionally narrow: the
        # things you'd actually paste into another CRM.
        header = [
            "name",
            "niche",
            "region",
            "score",
            "lead_status",
            "rating",
            "reviews_count",
            "phone",
            "website",
            "address",
            "category",
            "notes",
            "last_touched_at",
            "created_at",
        ]

        def _row_bytes(values: list[Any]) -> bytes:
            buf = _io.StringIO()
            _csv.writer(buf, quoting=_csv.QUOTE_MINIMAL).writerow(values)
            return buf.getvalue().encode("utf-8")

        async def generate() -> AsyncIterator[bytes]:
            # UTF-8 BOM so Excel on Windows opens Cyrillic columns cleanly.
            yield b"\xef\xbb\xbf"
            yield _row_bytes(header)
            async with session_factory() as session:
                stmt = (
                    select(Lead, SearchQuery.niche, SearchQuery.region)
                    .join(SearchQuery, SearchQuery.id == Lead.query_id)
                    .where(SearchQuery.source == "web")
                    .where(Lead.deleted_at.is_(None))
                    .order_by(
                        Lead.score_ai.desc().nullslast(),
                        Lead.created_at.desc(),
                    )
                    .limit(50_000)
                )
                if team_id is not None:
                    target_user = await _resolve_team_view(
                        session, team_id, user_id, member_user_id
                    )
                    stmt = stmt.where(
                        SearchQuery.team_id == team_id
                    ).where(SearchQuery.user_id == target_user)
                else:
                    stmt = stmt.where(SearchQuery.user_id == user_id).where(
                        SearchQuery.team_id.is_(None)
                    )
                result = await session.stream(stmt.execution_options(yield_per=500))
                async for lead, niche, region in result:
                    yield _row_bytes(
                        [
                            lead.name or "",
                            niche or "",
                            region or "",
                            ""
                            if lead.score_ai is None
                            else int(round(lead.score_ai)),
                            lead.lead_status or "",
                            "" if lead.rating is None else lead.rating,
                            ""
                            if lead.reviews_count is None
                            else lead.reviews_count,
                            lead.phone or "",
                            lead.website or "",
                            lead.address or "",
                            lead.category or "",
                            (lead.notes or "").replace("\n", " "),
                            lead.last_touched_at.isoformat()
                            if lead.last_touched_at
                            else "",
                            lead.created_at.isoformat()
                            if lead.created_at
                            else "",
                        ]
                    )

        filename = f"convioo-leads-{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"
        return StreamingResponse(
            generate(),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )

    @app.get(
        "/api/v1/searches/{query_id}/export.xlsx", include_in_schema=False
    )
    async def export_session_xlsx(
        query_id: uuid.UUID,
        current_user: User = Depends(get_current_user),
    ) -> Response:
        """Export one search session as a styled Excel workbook.

        One sheet, header row formatted bold, frozen first row, columns
        auto-fit-ish. The deliberately narrow column set matches the CSV
        export so the user gets the same shape they're used to plus the
        extra polish (cell types, no BOM hack) that Excel users expect.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        async with session_factory() as session:
            query = await session.get(SearchQuery, query_id)
            if query is None:
                raise HTTPException(status_code=404, detail="search not found")
            # Cross-user access answers 404 (not 403) so the export URL
            # can't be used to probe which session ids exist.
            allowed = query.user_id == current_user.id
            if not allowed and query.team_id is not None:
                allowed = (
                    await _membership(session, query.team_id, current_user.id)
                ) is not None
            if not allowed:
                raise HTTPException(status_code=404, detail="search not found")
            rows = list(
                (
                    await session.execute(
                        select(Lead)
                        .where(Lead.query_id == query_id)
                        .where(Lead.deleted_at.is_(None))
                        .order_by(
                            Lead.score_ai.desc().nullslast(),
                            Lead.created_at.desc(),
                        )
                    )
                )
                .scalars()
                .all()
            )

        _xlsx_lang = normalize_lang(current_user.language_code)
        _xlsx_headers = {
            "ru": [
                "Название",
                "Скор",
                "Статус",
                "Рейтинг",
                "Отзывов",
                "Телефон",
                "Сайт",
                "Адрес",
                "Категория",
                "Заметки",
                "Последнее касание",
                "Создан",
            ],
            "uk": [
                "Назва",
                "Скор",
                "Статус",
                "Рейтинг",
                "Відгуків",
                "Телефон",
                "Сайт",
                "Адреса",
                "Категорія",
                "Нотатки",
                "Останній контакт",
                "Створено",
            ],
            "en": [
                "Name",
                "Score",
                "Status",
                "Rating",
                "Reviews",
                "Phone",
                "Website",
                "Address",
                "Category",
                "Notes",
                "Last touched",
                "Created",
            ],
        }
        headers = _xlsx_headers[_xlsx_lang]

        # openpyxl is pure-Python and CPU-bound; running it inline blocks
        # the event loop for the entire workbook build + zip. Offload to
        # the default thread pool so other requests keep flowing while
        # the export builds.
        def _build_workbook() -> bytes:
            wb = Workbook()
            ws = wb.active
            ws.title = (query.niche or "leads")[:30]
            ws.append(headers)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="3D5AFE",
                end_color="3D5AFE",
                fill_type="solid",
            )
            header_align = Alignment(vertical="center")
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            for lead in rows:
                ws.append(
                    [
                        lead.name or "",
                        ""
                        if lead.score_ai is None
                        else int(round(lead.score_ai)),
                        lead.lead_status or "",
                        "" if lead.rating is None else lead.rating,
                        ""
                        if lead.reviews_count is None
                        else lead.reviews_count,
                        lead.phone or "",
                        lead.website or "",
                        lead.address or "",
                        lead.category or "",
                        (lead.notes or "").replace("\n", " "),
                        lead.last_touched_at.isoformat()
                        if lead.last_touched_at
                        else "",
                        lead.created_at.isoformat()
                        if lead.created_at
                        else "",
                    ]
                )

            widths = [32, 8, 12, 8, 10, 18, 36, 36, 22, 40, 22, 22]
            for i, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 22

            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

        body = await asyncio.to_thread(_build_workbook)
        slug = (query.niche or "session").replace(" ", "-").lower()[:40]
        date = datetime.now(timezone.utc).strftime("%Y%m%d")
        filename = f"convioo-{slug}-{date}.xlsx"
        return Response(
            content=body,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )

    # NOTE: registered BEFORE the /api/v1/leads/{lead_id} routes —
    # Starlette matches in registration order and the literal
    # "bulk" segment would otherwise be captured by {lead_id}.
    @app.patch(
        "/api/v1/leads/bulk", response_model=LeadBulkUpdateResponse
    )
    async def bulk_update_leads(
        body: LeadBulkUpdateRequest,
        current_user: User = Depends(get_current_user),
    ) -> LeadBulkUpdateResponse:
        """Apply ``lead_status`` and/or the caller's mark to many leads
        in one round-trip. The CRM bulk-toolbar uses this so the user
        can sweep dozens of rows in one click.

        Only leads the caller owns (or shares a team with via the
        parent search) are touched — foreign ids in the payload are
        silently dropped from the update set.
        """
        if not body.lead_status and not body.set_mark_color:
            raise HTTPException(
                status_code=400, detail="nothing to update"
            )

        async with session_factory() as session:
            # Authorise per-lead: keep only ids whose parent search the
            # caller owns or can reach through a team membership.
            owner_rows = (
                await session.execute(
                    select(Lead.id, SearchQuery.user_id, SearchQuery.team_id)
                    .join(SearchQuery, SearchQuery.id == Lead.query_id)
                    .where(Lead.id.in_(body.lead_ids))
                )
            ).all()
            team_member_cache: dict[uuid.UUID, bool] = {}
            allowed_ids: list[uuid.UUID] = []
            for row_lead_id, owner_id, owner_team_id in owner_rows:
                if owner_id == current_user.id:
                    allowed_ids.append(row_lead_id)
                    continue
                if owner_team_id is None:
                    continue
                is_member = team_member_cache.get(owner_team_id)
                if is_member is None:
                    is_member = (
                        await _membership(
                            session, owner_team_id, current_user.id
                        )
                    ) is not None
                    team_member_cache[owner_team_id] = is_member
                if is_member:
                    allowed_ids.append(row_lead_id)
            # Permissive validation: accept if matches a legacy key OR
            # any team's custom palette. Bulk operations span teams so
            # a strict per-team check would block mixed selections.
            if (
                body.lead_status
                and body.lead_status not in LEGACY_LEAD_STATUS_KEYS
            ):
                custom = (
                    await session.execute(
                        select(LeadStatus.key).where(
                            LeadStatus.key == body.lead_status
                        ).limit(1)
                    )
                ).scalar_one_or_none()
                if custom is None:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "lead_status is not a valid key in any "
                            "team palette or the default set"
                        ),
                    )
            if body.lead_status and allowed_ids:
                await session.execute(
                    update(Lead)
                    .where(Lead.id.in_(allowed_ids))
                    .values(
                        lead_status=body.lead_status,
                        last_touched_at=datetime.now(timezone.utc),
                    )
                )

            if body.set_mark_color and allowed_ids:
                color = (body.mark_color or "").strip() or None
                if color is None:
                    await session.execute(
                        sa.delete(LeadMark)
                        .where(LeadMark.user_id == current_user.id)
                        .where(LeadMark.lead_id.in_(allowed_ids))
                    )
                else:
                    # Per-row upsert. Postgres ON CONFLICT keeps it cheap;
                    # SQLite (test harness) iterates Python-side.
                    from sqlalchemy.dialects.postgresql import (
                        insert as pg_insert,
                    )

                    rows = [
                        {
                            "user_id": current_user.id,
                            "lead_id": lid,
                            "color": color,
                            "updated_at": datetime.now(timezone.utc),
                        }
                        for lid in allowed_ids
                    ]
                    if session.bind.dialect.name == "postgresql":
                        stmt = pg_insert(LeadMark).values(rows)
                        stmt = stmt.on_conflict_do_update(
                            index_elements=["user_id", "lead_id"],
                            set_={
                                "color": color,
                                "updated_at": datetime.now(timezone.utc),
                            },
                        )
                        await session.execute(stmt)
                    else:
                        for r in rows:
                            existing = (
                                await session.execute(
                                    select(LeadMark)
                                    .where(LeadMark.user_id == r["user_id"])
                                    .where(LeadMark.lead_id == r["lead_id"])
                                )
                            ).scalar_one_or_none()
                            if existing:
                                existing.color = color
                                existing.updated_at = r["updated_at"]
                            else:
                                session.add(LeadMark(**r))

            await session.commit()

            # Touched rows = requested ids that exist AND passed the
            # ownership filter.
            return LeadBulkUpdateResponse(updated=len(allowed_ids))

    @app.patch("/api/v1/leads/{lead_id}", response_model=LeadResponse)
    async def update_lead(
        lead_id: uuid.UUID,
        body: LeadUpdate,
        background_tasks: BackgroundTasks,
        current_user: User = Depends(get_current_user),
    ) -> LeadResponse:
        """Partial update: status, owner, notes. Touches last_touched_at.

        Writes ``lead_activities`` rows per changed field so the
        timeline + team feed have something to render. The actor is
        always the authenticated user — used to point at a query-param
        default which broke the lead_activities FK.
        """
        actor_user_id = current_user.id
        async with session_factory() as session:
            lead = await session.get(Lead, lead_id)
            if lead is None:
                raise HTTPException(status_code=404, detail="lead not found")

            # Ownership: the caller must own the parent search or be a
            # member of the team it belongs to. Cross-user access gets
            # a 404 (not 403) so lead ids can't be probed for existence.
            search = await session.get(SearchQuery, lead.query_id)
            allowed = search is not None and search.user_id == actor_user_id
            if not allowed and search is not None and search.team_id is not None:
                allowed = (
                    await _membership(session, search.team_id, actor_user_id)
                ) is not None
            if not allowed:
                raise HTTPException(status_code=404, detail="lead not found")

            # Lead-status validation: team-mode searches use the
            # team's custom palette; personal-mode searches keep the
            # legacy hard-coded keys. Either way an unknown key fails.
            if body.lead_status is not None:
                search_for_status = search
                valid_keys: set[str] | frozenset[str]
                if search_for_status and search_for_status.team_id is not None:
                    valid_keys = {
                        k for (k,) in (
                            await session.execute(
                                select(LeadStatus.key).where(
                                    LeadStatus.team_id == search_for_status.team_id
                                )
                            )
                        ).all()
                    }
                    # Defensive fallback — if the team's palette wasn't
                    # seeded for some reason, accept the legacy keys
                    # rather than rejecting every drag-and-drop.
                    if not valid_keys:
                        valid_keys = set(LEGACY_LEAD_STATUS_KEYS)
                else:
                    valid_keys = LEGACY_LEAD_STATUS_KEYS
                if body.lead_status not in valid_keys:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "lead_status must be one of "
                            + ", ".join(sorted(valid_keys))
                        ),
                    )

            # Capture before/after so we can write meaningful activity
            # rows. The fields list mirrors what LeadUpdate exposes —
            # if a new field gets added there, add it here too.
            activities: list[dict[str, Any]] = []
            now = datetime.now(timezone.utc)

            if body.lead_status is not None and body.lead_status != lead.lead_status:
                activities.append(
                    {
                        "kind": "status",
                        "payload": {
                            "from": lead.lead_status,
                            "to": body.lead_status,
                        },
                    }
                )
                lead.lead_status = body.lead_status
            if "owner_user_id" in body.model_fields_set:
                if body.owner_user_id != lead.owner_user_id:
                    activities.append(
                        {
                            "kind": "assigned",
                            "payload": {
                                "from": lead.owner_user_id,
                                "to": body.owner_user_id,
                            },
                        }
                    )
                lead.owner_user_id = body.owner_user_id
            if "deal_value" in body.model_fields_set:
                lead.deal_value = body.deal_value

            if body.notes is not None and body.notes != (lead.notes or ""):
                activities.append(
                    {
                        "kind": "notes",
                        "payload": {"len": len(body.notes)},
                    }
                )
                lead.notes = body.notes

            if not activities and (
                body.lead_status is None
                and body.notes is None
                and "owner_user_id" not in body.model_fields_set
                and "deal_value" not in body.model_fields_set
            ):
                raise HTTPException(status_code=400, detail="no fields to update")

            lead.last_touched_at = now

            # Pull team_id off the parent search query so the activity
            # row can land in the team feed when the lead is shared.
            team_id_for_activity = search.team_id if search else None

            for act in activities:
                session.add(
                    LeadActivity(
                        lead_id=lead.id,
                        user_id=actor_user_id,
                        team_id=team_id_for_activity,
                        kind=act["kind"],
                        payload=act["payload"],
                    )
                )
            await session.commit()
            await session.refresh(lead)

            # Emit lead.status_changed if the status moved this round.
            # We notify the search owner — they're who registered the
            # webhook against their account, regardless of who edited
            # it inside the team.
            status_change = next(
                (a for a in activities if a["kind"] == "status"), None
            )
            if status_change and search is not None:
                background_tasks.add_task(
                    emit_webhook_event_sync,
                    search.user_id,
                    "lead.status_changed",
                    {
                        "lead": serialize_lead_for_webhook(lead),
                        "from_status": status_change["payload"]["from"],
                        "to_status": status_change["payload"]["to"],
                        "actor_user_id": actor_user_id,
                    },
                )
                if status_change["payload"]["to"] == "won":
                    background_tasks.add_task(
                        send_slack_notification,
                        f"Lead won: {lead.name} ({lead.website})",
                    )
            return LeadResponse.model_validate(lead)

    @app.delete("/api/v1/leads/{lead_id}")
    async def delete_lead(
        lead_id: uuid.UUID,
        forever: bool = False,
        current_user: User = Depends(get_current_user),
    ) -> dict[str, bool]:
        """Soft-delete a lead so it disappears from the CRM.

        ``forever=true`` additionally writes a row into the seen-leads
        table so future searches will treat the same place_id /
        phone / domain as already-delivered and skip it. Without
        ``forever``, the lead is just hidden — re-running a similar
        search may surface it again.

        Authorisation: caller must own the parent ``SearchQuery`` (or
        be a member of the team that owns it).
        """
        async with session_factory() as session:
            lead = await session.get(Lead, lead_id)
            if lead is None:
                raise HTTPException(status_code=404, detail="lead not found")
            search = await session.get(SearchQuery, lead.query_id)
            if search is None:
                raise HTTPException(status_code=404, detail="search not found")

            allowed = search.user_id == current_user.id
            if not allowed and search.team_id is not None:
                membership = await _membership(
                    session, search.team_id, current_user.id
                )
                allowed = membership is not None
            if not allowed:
                raise HTTPException(status_code=403, detail="forbidden")

            if lead.deleted_at is None:
                lead.deleted_at = datetime.now(timezone.utc)
            if forever:
                lead.blacklisted = True
                # Make sure the seen-leads record exists with all three
                # dedup axes filled, even if the lead came in before the
                # 0023 migration backfilled them.
                from leadgen.utils.dedup import (
                    domain_root as _domain_root,
                )
                from leadgen.utils.dedup import (
                    normalize_phone as _normalize_phone,
                )

                phone_key = _normalize_phone(lead.phone)
                domain_key = _domain_root(lead.website)

                if search.user_id != 0:
                    existing_user = (
                        await session.execute(
                            select(UserSeenLead)
                            .where(UserSeenLead.user_id == search.user_id)
                            .where(UserSeenLead.source == lead.source)
                            .where(UserSeenLead.source_id == lead.source_id)
                        )
                    ).scalar_one_or_none()
                    if existing_user is None:
                        session.add(
                            UserSeenLead(
                                user_id=search.user_id,
                                source=lead.source,
                                source_id=lead.source_id,
                                phone_e164=phone_key,
                                domain_root=domain_key,
                            )
                        )
                    else:
                        existing_user.phone_e164 = phone_key
                        existing_user.domain_root = domain_key
                if search.team_id is not None:
                    existing_team = (
                        await session.execute(
                            select(TeamSeenLead)
                            .where(TeamSeenLead.team_id == search.team_id)
                            .where(TeamSeenLead.source == lead.source)
                            .where(TeamSeenLead.source_id == lead.source_id)
                        )
                    ).scalar_one_or_none()
                    if existing_team is None:
                        session.add(
                            TeamSeenLead(
                                team_id=search.team_id,
                                source=lead.source,
                                source_id=lead.source_id,
                                phone_e164=phone_key,
                                domain_root=domain_key,
                                first_user_id=search.user_id,
                            )
                        )
                    else:
                        existing_team.phone_e164 = phone_key
                        existing_team.domain_root = domain_key

            session.add(
                LeadActivity(
                    lead_id=lead.id,
                    user_id=current_user.id,
                    team_id=search.team_id,
                    kind="deleted",
                    payload={"forever": bool(forever)},
                )
            )
            await session.commit()
        return {"ok": True, "forever": bool(forever)}

    _enriching_leads: set[str] = set()

    @app.post("/api/v1/leads/{lead_id}/re-enrich", response_model=LeadResponse)
    async def re_enrich_lead(
        lead_id: uuid.UUID,
        current_user: User = Depends(get_current_user),
    ) -> LeadResponse:
        """Trigger a fresh AI enrichment pass for a single lead.

        Returns 409 if enrichment is already running for this lead_id.
        """
        from leadgen.collectors import GooglePlacesCollector
        from leadgen.pipeline.enrichment import enrich_leads

        lead_id_str = str(lead_id)
        if lead_id_str in _enriching_leads:
            raise HTTPException(status_code=409, detail="enrichment already in progress")
        _enriching_leads.add(lead_id_str)
        try:
            async with session_factory() as session:
                lead = await session.get(Lead, lead_id)
                if lead is None:
                    raise HTTPException(status_code=404, detail="lead not found")
                search = await session.get(SearchQuery, lead.query_id)
                if search is None:
                    raise HTTPException(status_code=404, detail="search not found")
                allowed = search.user_id == current_user.id
                if not allowed and search.team_id is not None:
                    allowed = (
                        await _membership(
                            session, search.team_id, current_user.id
                        )
                    ) is not None
                if not allowed:
                    raise HTTPException(
                        status_code=404, detail="lead not found"
                    )

            collector = GooglePlacesCollector()
            await enrich_leads(
                [lead],
                collector,
                search.niche,
                search.region,
            )

            async with session_factory() as session:
                updated = await session.get(Lead, lead_id)
                if updated is None:
                    raise HTTPException(status_code=404, detail="lead not found")
                return LeadResponse.model_validate(updated)
        finally:
            _enriching_leads.discard(lead_id_str)

    # ── /api/v1/saved-searches (bookmark + recurring re-run) ───────────

    def _saved_to_schema(row: SavedSearch) -> SavedSearchSchema:
        return SavedSearchSchema(
            id=str(row.id),
            name=row.name,
            team_id=str(row.team_id) if row.team_id else None,
            niche=row.niche,
            region=row.region,
            target_languages=row.target_languages,
            scope=row.scope,
            radius_m=row.radius_m,
            max_results=row.max_results,
            schedule=row.schedule,
            next_run_at=row.next_run_at,
            last_run_at=row.last_run_at,
            last_leads_count=row.last_leads_count,
            active=row.active,
            created_at=row.created_at,
            updated_at=row.updated_at,
        )

    def _normalize_schedule(raw: str | None) -> str | None:
        """Map ``"off"`` and the empty string onto ``None`` so the
        worker query can ``WHERE schedule IS NOT NULL`` cleanly."""
        if not raw:
            return None
        value = raw.strip().lower()
        if value in ("off", "none", "manual", ""):
            return None
        from leadgen.core.services.saved_searches import VALID_SCHEDULES

        if value not in VALID_SCHEDULES:
            raise HTTPException(
                status_code=400,
                detail=(
                    "schedule must be one of off / daily / weekly / "
                    "biweekly / monthly"
                ),
            )
        return value

    @app.get(
        "/api/v1/saved-searches",
        response_model=SavedSearchListResponse,
    )
    async def list_saved_searches(
        current_user: User = Depends(get_current_user),
    ) -> SavedSearchListResponse:
        async with session_factory() as session:
            team_ids = (
                (
                    await session.execute(
                        select(TeamMembership.team_id).where(
                            TeamMembership.user_id == current_user.id
                        )
                    )
                )
                .scalars()
                .all()
            )
            stmt = (
                select(SavedSearch)
                .where(
                    sa.or_(
                        sa.and_(
                            SavedSearch.user_id == current_user.id,
                            SavedSearch.team_id.is_(None),
                        ),
                        SavedSearch.team_id.in_(team_ids)
                        if team_ids
                        else sa.false(),
                    )
                )
                .order_by(SavedSearch.created_at.desc())
            )
            rows = (await session.execute(stmt)).scalars().all()
        return SavedSearchListResponse(
            items=[_saved_to_schema(r) for r in rows]
        )

    @app.post(
        "/api/v1/saved-searches",
        response_model=SavedSearchSchema,
    )
    async def create_saved_search(
        body: SavedSearchCreate,
        current_user: User = Depends(get_current_user),
    ) -> SavedSearchSchema:
        from leadgen.core.services.saved_searches import (
            next_run_after,
        )

        team_uuid: uuid.UUID | None = None
        if body.team_id:
            try:
                team_uuid = uuid.UUID(body.team_id)
            except ValueError as exc:
                raise HTTPException(
                    status_code=400, detail="invalid team_id"
                ) from exc

        schedule = _normalize_schedule(body.schedule)
        async with session_factory() as session:
            if team_uuid is not None:
                membership = (
                    await session.execute(
                        select(TeamMembership)
                        .where(TeamMembership.user_id == current_user.id)
                        .where(TeamMembership.team_id == team_uuid)
                    )
                ).scalar_one_or_none()
                if membership is None:
                    raise HTTPException(
                        status_code=403, detail="not a team member"
                    )
            row = SavedSearch(
                user_id=current_user.id,
                team_id=team_uuid,
                name=body.name.strip(),
                niche=body.niche.strip(),
                region=body.region.strip(),
                target_languages=body.target_languages,
                scope=body.scope,
                radius_m=body.radius_m,
                max_results=body.max_results,
                schedule=schedule,
                next_run_at=next_run_after(schedule)
                if schedule
                else None,
                active=True,
            )
            session.add(row)
            await session.commit()
            await session.refresh(row)
        return _saved_to_schema(row)

    @app.patch(
        "/api/v1/saved-searches/{saved_id}",
        response_model=SavedSearchSchema,
    )
    async def update_saved_search(
        saved_id: uuid.UUID,
        body: SavedSearchUpdate,
        current_user: User = Depends(get_current_user),
    ) -> SavedSearchSchema:
        from leadgen.core.services.saved_searches import (
            next_run_after,
        )

        async with session_factory() as session:
            row = await session.get(SavedSearch, saved_id)
            if row is None or row.user_id != current_user.id:
                raise HTTPException(
                    status_code=404, detail="saved search not found"
                )
            if body.name is not None:
                row.name = body.name.strip()
            if body.schedule is not None:
                row.schedule = _normalize_schedule(body.schedule)
                row.next_run_at = (
                    next_run_after(row.schedule) if row.schedule else None
                )
            if body.active is not None:
                row.active = body.active
            if body.max_results is not None:
                row.max_results = body.max_results
            if body.radius_m is not None:
                row.radius_m = body.radius_m
            row.updated_at = datetime.now(timezone.utc)
            await session.commit()
            await session.refresh(row)
        return _saved_to_schema(row)

    @app.delete("/api/v1/saved-searches/{saved_id}")
    async def delete_saved_search(
        saved_id: uuid.UUID,
        current_user: User = Depends(get_current_user),
    ) -> dict[str, bool]:
        async with session_factory() as session:
            row = await session.get(SavedSearch, saved_id)
            if row is None or row.user_id != current_user.id:
                raise HTTPException(
                    status_code=404, detail="saved search not found"
                )
            await session.delete(row)
            await session.commit()
        return {"ok": True}

    @app.post(
        "/api/v1/saved-searches/{saved_id}/run",
        response_model=SearchCreateResponse,
    )
    async def run_saved_search_now(
        saved_id: uuid.UUID,
        current_user: User = Depends(get_current_user),
    ) -> SearchCreateResponse:
        """Manually trigger a saved search outside its schedule.

        Useful for the "Run now" button next to each saved search row.
        Reuses the same enqueue plumbing that ``POST /searches`` does so
        the SSE progress stream and the CRM lead-list are identical.
        """
        from leadgen.core.services.saved_searches import build_search_query

        async with session_factory() as session:
            row = await session.get(SavedSearch, saved_id)
            if row is None or row.user_id != current_user.id:
                raise HTTPException(
                    status_code=404, detail="saved search not found"
                )
            user = await session.get(User, current_user.id)
            new_query = build_search_query(row)
            session.add(new_query)
            row.last_run_at = datetime.now(timezone.utc)
            await session.commit()
            query_id = new_query.id

        user_profile = {
            "display_name": user.display_name or user.first_name if user else None,
            "language_code": user.language_code if user else None,
        }
        queued_id = await enqueue_search(
            query_id, chat_id=None, user_profile=user_profile
        )
        queued = bool(queued_id)
        if not queued:
            spawn(
                _run_web_search_inline(query_id, user_profile),
                name=f"convioo-web-search-{query_id}",
            )
        return SearchCreateResponse(id=query_id, queued=queued)

    # ── /api/v1/leads/{id}/custom-fields ────────────────────────────────

    @app.get(
        "/api/v1/leads/{lead_id}/custom-fields",
        response_model=LeadCustomFieldsResponse,
    )
    async def list_lead_custom_fields(
        lead_id: uuid.UUID,
        current_user: User = Depends(get_current_user),
    ) -> LeadCustomFieldsResponse:
        user_id = current_user.id
        async with session_factory() as session:
            stmt = (
                select(LeadCustomField)
                .where(LeadCustomField.lead_id == lead_id)
                .where(LeadCustomField.user_id == user_id)
                .order_by(LeadCustomField.key)
            )
            rows = (await session.execute(stmt)).scalars().all()
            items = [
                LeadCustomFieldSchema.model_validate(r) for r in rows
            ]
        return LeadCustomFieldsResponse(items=items)

    @app.put(
        "/api/v1/leads/{lead_id}/custom-fields",
        response_model=LeadCustomFieldSchema,
    )
    async def upsert_lead_custom_field(
        lead_id: uuid.UUID,
        body: LeadCustomFieldUpsert,
        current_user: User = Depends(get_current_user),
    ) -> LeadCustomFieldSchema:
        """Create or update one (key, value) pair on this lead.

        Schemaless — the user picks any key from the UI. ``value`` may
        be NULL, which acts as a soft-delete on the row (we still keep
        the row so the timeline can reference the historical key).
        """
        user_id = current_user.id
        key = body.key.strip()
        if not key:
            raise HTTPException(status_code=400, detail="key is required")
        value = body.value if body.value is None else body.value.strip()
        async with session_factory() as session:
            existing = (
                await session.execute(
                    select(LeadCustomField)
                    .where(LeadCustomField.lead_id == lead_id)
                    .where(LeadCustomField.user_id == user_id)
                    .where(LeadCustomField.key == key)
                    .limit(1)
                )
            ).scalar_one_or_none()
            now = datetime.now(timezone.utc)
            search = (
                await session.execute(
                    select(SearchQuery)
                    .join(Lead, Lead.query_id == SearchQuery.id)
                    .where(Lead.id == lead_id)
                    .limit(1)
                )
            ).scalar_one_or_none()
            team_id_for_activity = search.team_id if search else None
            if existing is None:
                existing = LeadCustomField(
                    lead_id=lead_id,
                    user_id=user_id,
                    key=key,
                    value=value,
                )
                session.add(existing)
            else:
                existing.value = value
                existing.updated_at = now
            session.add(
                LeadActivity(
                    lead_id=lead_id,
                    user_id=user_id,
                    team_id=team_id_for_activity,
                    kind="custom_field",
                    payload={"key": key, "value": value},
                )
            )
            await session.commit()
            await session.refresh(existing)
            return LeadCustomFieldSchema.model_validate(existing)

    @app.delete("/api/v1/leads/{lead_id}/custom-fields/{key}")
    async def delete_lead_custom_field(
        lead_id: uuid.UUID,
        key: str,
        current_user: User = Depends(get_current_user),
    ) -> dict[str, bool]:
        user_id = current_user.id
        async with session_factory() as session:
            row = (
                await session.execute(
                    select(LeadCustomField)
                    .where(LeadCustomField.lead_id == lead_id)
                    .where(LeadCustomField.user_id == user_id)
                    .where(LeadCustomField.key == key)
                    .limit(1)
                )
            ).scalar_one_or_none()
            if row is None:
                return {"deleted": False}
            await session.delete(row)
            await session.commit()
        return {"deleted": True}

    # ── /api/v1/leads/{id}/activity ─────────────────────────────────────

    @app.get(
        "/api/v1/leads/{lead_id}/activity",
        response_model=LeadActivityListResponse,
    )
    async def list_lead_activity(
        lead_id: uuid.UUID,
        limit: int = 50,
        current_user: User = Depends(get_current_user),
    ) -> LeadActivityListResponse:
        limit = max(1, min(limit, 200))
        async with session_factory() as session:
            await _authorise_lead_access(session, lead_id, current_user.id)
            stmt = (
                select(LeadActivity)
                .where(LeadActivity.lead_id == lead_id)
                .order_by(LeadActivity.created_at.desc())
                .limit(limit)
            )
            rows = (await session.execute(stmt)).scalars().all()
            items = [LeadActivitySchema.model_validate(r) for r in rows]
        return LeadActivityListResponse(items=items)

    # ── /api/v1/leads/{id}/tasks ────────────────────────────────────────

    @app.get(
        "/api/v1/leads/{lead_id}/tasks",
        response_model=LeadTaskListResponse,
    )
    async def list_lead_tasks(
        lead_id: uuid.UUID,
        current_user: User = Depends(get_current_user),
    ) -> LeadTaskListResponse:
        user_id = current_user.id
        async with session_factory() as session:
            stmt = (
                select(LeadTask)
                .where(LeadTask.lead_id == lead_id)
                .where(LeadTask.user_id == user_id)
                .order_by(
                    LeadTask.done_at.is_(None).desc(),
                    LeadTask.due_at.asc().nullslast(),
                    LeadTask.created_at.desc(),
                )
            )
            rows = (await session.execute(stmt)).scalars().all()
            items = [LeadTaskSchema.model_validate(r) for r in rows]
        return LeadTaskListResponse(items=items)

    @app.post(
        "/api/v1/leads/{lead_id}/tasks",
        response_model=LeadTaskSchema,
    )
    async def create_lead_task(
        lead_id: uuid.UUID,
        body: LeadTaskCreate,
        current_user: User = Depends(get_current_user),
    ) -> LeadTaskSchema:
        user_id = current_user.id
        async with session_factory() as session:
            row = LeadTask(
                lead_id=lead_id,
                user_id=user_id,
                content=body.content.strip(),
                due_at=body.due_at,
            )
            session.add(row)
            search = (
                await session.execute(
                    select(SearchQuery)
                    .join(Lead, Lead.query_id == SearchQuery.id)
                    .where(Lead.id == lead_id)
                    .limit(1)
                )
            ).scalar_one_or_none()
            session.add(
                LeadActivity(
                    lead_id=lead_id,
                    user_id=user_id,
                    team_id=search.team_id if search else None,
                    kind="task",
                    payload={
                        "content": body.content.strip()[:200],
                        "due_at": body.due_at.isoformat() if body.due_at else None,
                    },
                )
            )
            await session.commit()
            await session.refresh(row)
            return LeadTaskSchema.model_validate(row)

    @app.patch(
        "/api/v1/tasks/{task_id}",
        response_model=LeadTaskSchema,
    )
    async def update_lead_task(
        task_id: uuid.UUID,
        body: LeadTaskUpdate,
        current_user: User = Depends(get_current_user),
    ) -> LeadTaskSchema:
        user_id = current_user.id
        async with session_factory() as session:
            row = await session.get(LeadTask, task_id)
            if row is None or row.user_id != user_id:
                raise HTTPException(status_code=404, detail="task not found")
            data = body.model_dump(exclude_unset=True)
            if "content" in data and data["content"]:
                row.content = data["content"].strip()
            if "due_at" in data:
                row.due_at = data["due_at"]
            if "done" in data and data["done"] is not None:
                row.done_at = (
                    datetime.now(timezone.utc) if data["done"] else None
                )
            await session.commit()
            await session.refresh(row)
            return LeadTaskSchema.model_validate(row)

    @app.delete("/api/v1/tasks/{task_id}")
    async def delete_lead_task(
        task_id: uuid.UUID,
        current_user: User = Depends(get_current_user),
    ) -> dict[str, bool]:
        user_id = current_user.id
        async with session_factory() as session:
            row = await session.get(LeadTask, task_id)
            if row is None or row.user_id != user_id:
                return {"deleted": False}
            await session.delete(row)
            await session.commit()
        return {"deleted": True}

    @app.get(
        "/api/v1/users/me/tasks",
        response_model=LeadTaskListResponse,
    )
    async def list_my_tasks(
        open_only: bool = True,
        limit: int = 100,
        current_user: User = Depends(get_current_user),
    ) -> LeadTaskListResponse:
        """Today's-tasks widget feed: open tasks across every lead."""
        user_id = current_user.id
        limit = max(1, min(limit, 500))
        async with session_factory() as session:
            stmt = select(LeadTask).where(LeadTask.user_id == user_id)
            if open_only:
                stmt = stmt.where(LeadTask.done_at.is_(None))
            stmt = stmt.order_by(
                LeadTask.due_at.asc().nullslast(),
                LeadTask.created_at.desc(),
            ).limit(limit)
            rows = (await session.execute(stmt)).scalars().all()
            items = [LeadTaskSchema.model_validate(r) for r in rows]
        return LeadTaskListResponse(items=items)

    @app.post(
        "/api/v1/leads/{lead_id}/draft-email",
        response_model=LeadEmailDraftResponse,
    )
    async def draft_lead_email(
        lead_id: uuid.UUID,
        body: LeadEmailDraftRequest,
        current_user: User = Depends(get_current_user),
    ) -> LeadEmailDraftResponse:
        """Generate a personalised cold-email draft for one lead.

        The frontend opens the draft inline in the lead modal — the
        salesperson can copy the subject + body (or regenerate with a
        different tone) and paste into Gmail. Real send-via-Gmail
        ships once the OAuth connector lands.
        """
        async with session_factory() as session:
            lead, _search = await _authorise_lead_access(
                session, lead_id, current_user.id
            )
            user = await session.get(User, current_user.id)

        user_profile: dict[str, Any] = {}
        if user is not None:
            user_profile = {
                "display_name": user.display_name or user.first_name,
                "age_range": user.age_range,
                "gender": user.gender,
                "business_size": user.business_size,
                "profession": user.profession,
                "service_description": user.service_description,
                "home_region": user.home_region,
                "niches": list(user.niches or []),
                "language_code": user.language_code,
                "calendly_url": user.calendly_url,
                "icp_profile": user.icp_profile,
            }

        lead_payload = {
            "name": lead.name,
            "category": lead.category,
            "address": lead.address,
            "website": lead.website,
            "rating": lead.rating,
            "reviews_count": lead.reviews_count,
            "score_ai": lead.score_ai,
            "summary": lead.summary,
            "advice": lead.advice,
            "strengths": list(lead.strengths) if lead.strengths else None,
            "weaknesses": list(lead.weaknesses) if lead.weaknesses else None,
            "red_flags": list(lead.red_flags) if lead.red_flags else None,
        }

        analyzer = AIAnalyzer()

        # UI language (for the research headings shown to the user) vs
        # email language (per-draft override → UI language → ru).
        ui_lang = normalize_lang(user.language_code if user else None)
        email_language = normalize_lang(
            body.language or (user.language_code if user else None)
        )

        # Optional: deep research pass — fresh website fetch + Claude
        # extraction of notable facts. Threaded into ``extra_context``
        # so the existing email prompt naturally cites the lead's own
        # site instead of leaning on cached enrichment.
        notable_facts: list[str] = []
        recent_signal: str | None = None
        merged_extra = body.extra_context
        if body.deep_research:
            research = await analyzer.research_lead_for_outreach(
                lead_payload,
                user_profile=user_profile or None,
            )
            notable_facts = list(research.get("notable_facts") or [])
            recent_signal = research.get("recent_signal")
            opener = research.get("suggested_opener")
            research_block_parts: list[str] = []
            if notable_facts:
                research_block_parts.append(
                    locale_pick(
                        ui_lang,
                        ru="Свежие факты с сайта (можно цитировать в opener):",
                        uk="Свіжі факти з сайту (можна цитувати в opener):",
                        en="Fresh facts from the site (quotable in the opener):",
                    )
                )
                for fact in notable_facts:
                    research_block_parts.append(f"- {fact}")
            if recent_signal:
                research_block_parts.append(
                    locale_pick(
                        ui_lang,
                        ru=f"Recent signal (что-то новое у них): {recent_signal}",
                        uk=f"Recent signal (щось нове у них): {recent_signal}",
                        en=f"Recent signal (something new on their side): {recent_signal}",
                    )
                )
            if opener:
                research_block_parts.append(
                    locale_pick(
                        ui_lang,
                        ru=f"Подсказанный opener: {opener}",
                        uk=f"Підказаний opener: {opener}",
                        en=f"Suggested opener: {opener}",
                    )
                )
            if research_block_parts:
                research_block = "\n".join(research_block_parts)
                merged_extra = (
                    f"{body.extra_context}\n\n{research_block}"
                    if body.extra_context
                    else research_block
                )

        result = await analyzer.generate_cold_email(
            lead_payload,
            user_profile=user_profile or None,
            tone=body.tone,
            extra_context=merged_extra,
            language=email_language,
        )
        return LeadEmailDraftResponse(
            subject=result["subject"],
            body=result["body"],
            tone=result["tone"],
            notable_facts=notable_facts,
            recent_signal=recent_signal,
        )

    @app.post(
        "/api/v1/leads/bulk-draft",
        response_model=BulkDraftEmailResponse,
    )
    async def bulk_draft_emails(
        body: BulkDraftEmailRequest,
        current_user: User = Depends(get_current_user),
    ) -> BulkDraftEmailResponse:
        """Generate cold-email drafts for up to 20 leads in one shot.

        The salesperson selects rows on /app/leads, hits "Написать
        всем", and gets back a stitched list ready for review. Per-
        lead errors don't take the whole batch down — failed entries
        come back with ``error`` populated.

        Concurrency is throttled (3 in-flight) so a 20-lead batch
        doesn't stampede Anthropic. Authorisation is per-lead: each
        lead must belong to a search the caller owns or is a member
        of via team.
        """
        async with session_factory() as session:
            user_profile: dict[str, Any] = {
                "display_name": current_user.display_name or current_user.first_name,
                "age_range": current_user.age_range,
                "gender": current_user.gender,
                "business_size": current_user.business_size,
                "profession": current_user.profession,
                "service_description": current_user.service_description,
                "home_region": current_user.home_region,
                "niches": list(current_user.niches or []),
                "language_code": current_user.language_code,
            }
            lead_rows = (
                (
                    await session.execute(
                        select(Lead, SearchQuery)
                        .join(SearchQuery, SearchQuery.id == Lead.query_id)
                        .where(Lead.id.in_(list(body.lead_ids)))
                    )
                )
                .all()
            )
            authorised: dict[uuid.UUID, Lead] = {}
            for lead, search in lead_rows:
                if search.user_id == current_user.id:
                    authorised[lead.id] = lead
                    continue
                if search.team_id is not None and (
                    await _membership(session, search.team_id, current_user.id)
                ):
                    authorised[lead.id] = lead

        analyzer = AIAnalyzer()
        sem = asyncio.Semaphore(3)
        tone = (body.tone or "professional").strip().lower()
        # Per-batch email language: explicit override → UI language → ru.
        email_language = normalize_lang(
            body.language or current_user.language_code
        )

        async def _one(lead_id: uuid.UUID) -> BulkDraftEmailItem:
            lead = authorised.get(lead_id)
            if lead is None:
                return BulkDraftEmailItem(
                    lead_id=lead_id, error="not authorised"
                )
            payload = {
                "name": lead.name,
                "category": lead.category,
                "address": lead.address,
                "website": lead.website,
                "rating": lead.rating,
                "reviews_count": lead.reviews_count,
                "score_ai": lead.score_ai,
                "summary": lead.summary,
                "advice": lead.advice,
                "strengths": list(lead.strengths) if lead.strengths else None,
                "weaknesses": list(lead.weaknesses) if lead.weaknesses else None,
                "red_flags": list(lead.red_flags) if lead.red_flags else None,
            }
            async with sem:
                try:
                    result = await analyzer.generate_cold_email(
                        payload,
                        user_profile=user_profile,
                        tone=tone,
                        extra_context=body.extra_context,
                        language=email_language,
                    )
                    return BulkDraftEmailItem(
                        lead_id=lead_id,
                        subject=result.get("subject"),
                        body=result.get("body"),
                    )
                except Exception as exc:  # noqa: BLE001
                    logger.exception(
                        "bulk-draft: failed for lead %s", lead_id
                    )
                    return BulkDraftEmailItem(
                        lead_id=lead_id, error=str(exc)[:200]
                    )

        items = await asyncio.gather(*(_one(lid) for lid in body.lead_ids))
        return BulkDraftEmailResponse(items=list(items))

    @app.put("/api/v1/leads/{lead_id}/mark", response_model=LeadResponse)
    async def set_lead_mark(
        lead_id: uuid.UUID,
        body: LeadMarkRequest,
        current_user: User = Depends(get_current_user),
    ) -> LeadResponse:
        """Set or clear the caller's private colour mark on a lead.

        Pass ``color: null`` to remove. The mark is only ever visible
        to the caller; teammates see their own marks (or none).
        """
        async with session_factory() as session:
            lead, _search = await _authorise_lead_access(
                session, lead_id, current_user.id
            )

            existing = (
                await session.execute(
                    select(LeadMark)
                    .where(LeadMark.user_id == current_user.id)
                    .where(LeadMark.lead_id == lead_id)
                    .limit(1)
                )
            ).scalar_one_or_none()

            color = (body.color or "").strip() or None
            if color is None:
                if existing is not None:
                    await session.delete(existing)
                final_color: str | None = None
            elif existing is None:
                session.add(
                    LeadMark(
                        user_id=current_user.id,
                        lead_id=lead_id,
                        color=color,
                    )
                )
                final_color = color
            else:
                existing.color = color
                existing.updated_at = datetime.now(timezone.utc)
                final_color = color

            await session.commit()
            await session.refresh(lead)
            return _to_lead_response(lead, final_color)

    @app.get(
        "/api/v1/teams/{team_id}/members-summary",
        response_model=list[TeamMemberSummary],
    )
    async def team_members_summary(
        team_id: uuid.UUID,
        current_user: User = Depends(get_current_user),
    ) -> list[TeamMemberSummary]:
        """Owner-only roll-up: per-member sessions/leads/hot counts.

        Powers the "see each teammate's CRM" panel — the owner picks a
        row and the workspace switches to viewing that member via
        ``member_user_id`` on the list endpoints.
        """
        async with session_factory() as session:
            caller = await _membership(session, team_id, current_user.id)
            if caller is None or _normalize_role(caller.role) != _ROLE_OWNER:
                raise HTTPException(
                    status_code=403,
                    detail="only the team owner can see the per-member summary",
                )

            rows = (
                await session.execute(
                    select(TeamMembership, User)
                    .join(User, User.id == TeamMembership.user_id)
                    .where(TeamMembership.team_id == team_id)
                    .order_by(TeamMembership.created_at)
                )
            ).all()

            results: list[TeamMemberSummary] = []
            for membership, member in rows:
                sessions_total = int(
                    (
                        await session.execute(
                            select(func.count(SearchQuery.id))
                            .where(SearchQuery.team_id == team_id)
                            .where(SearchQuery.user_id == member.id)
                        )
                    ).scalar()
                    or 0
                )
                lead_scores = [
                    s
                    for s, in (
                        await session.execute(
                            select(Lead.score_ai)
                            .join(SearchQuery, SearchQuery.id == Lead.query_id)
                            .where(SearchQuery.team_id == team_id)
                            .where(SearchQuery.user_id == member.id)
                        )
                    ).all()
                ]
                hot = sum(1 for s in lead_scores if s is not None and s >= 75)
                display = (
                    member.display_name
                    or " ".join(filter(None, [member.first_name, member.last_name]))
                    or f"User {member.id}"
                )
                results.append(
                    TeamMemberSummary(
                        user_id=member.id,
                        name=display,
                        role=membership.role,
                        sessions_total=sessions_total,
                        leads_total=len(lead_scores),
                        hot_total=hot,
                    )
                )
            return results

    @app.get("/api/v1/stats", response_model=DashboardStats)
    async def dashboard_stats(
        team_id: uuid.UUID | None = None,
        member_user_id: int | None = None,
        current_user: User = Depends(get_current_user),
    ) -> DashboardStats:
        user_id = current_user.id
        async with session_factory() as session:
            query_stmt = (
                select(SearchQuery).where(SearchQuery.source == "web")
            )
            lead_stmt = (
                select(Lead.score_ai)
                .join(SearchQuery, SearchQuery.id == Lead.query_id)
                .where(SearchQuery.source == "web")
            )
            if team_id is not None:
                target_user = await _resolve_team_view(
                    session, team_id, user_id, member_user_id
                )
                query_stmt = query_stmt.where(
                    SearchQuery.team_id == team_id
                ).where(SearchQuery.user_id == target_user)
                lead_stmt = lead_stmt.where(
                    SearchQuery.team_id == team_id
                ).where(SearchQuery.user_id == target_user)
            else:
                query_stmt = query_stmt.where(SearchQuery.user_id == user_id).where(
                    SearchQuery.team_id.is_(None)
                )
                lead_stmt = lead_stmt.where(SearchQuery.user_id == user_id).where(
                    SearchQuery.team_id.is_(None)
                )

            searches = list((await session.execute(query_stmt)).scalars().all())
            scores = [row[0] for row in (await session.execute(lead_stmt)).all()]

        hot = sum(1 for s in scores if s is not None and s >= 75)
        warm = sum(1 for s in scores if s is not None and 50 <= s < 75)
        cold = sum(1 for s in scores if s is not None and s < 50)
        running = sum(1 for s in searches if s.status == "running")

        return DashboardStats(
            sessions_total=len(searches),
            sessions_running=running,
            leads_total=len(scores),
            hot_total=hot,
            warm_total=warm,
            cold_total=cold,
        )

    # GET /api/v1/team (flat all-users roster) was removed: nothing in
    # the frontend called it and it leaked every registered user's name
    # to any caller. Team rosters live at /api/v1/teams/{team_id}.

    @app.get("/api/v1/queue/status", include_in_schema=False)
    async def queue_status() -> dict[str, bool]:
        return {"queue_enabled": is_queue_enabled()}

    # /api/v1/tags moved to routes/tags.py

    # /api/v1/segments moved to routes/segments.py

    # ── /api/v1/integrations/notion ────────────────────────────────────
    #
    # Two connection paths:
    # 1. Public OAuth — user clicks "Connect Notion", authorizes via
    #    Notion's consent screen, callback saves the access_token.
    #    503-safe when NOTION_OAUTH_CLIENT_ID / _SECRET are unset.
    # 2. Internal integration token (legacy) — user pastes a token
    #    from notion.so/my-integrations. Still works for power users.
    #
    # Either way, the user must set a database_id via PATCH before
    # export works.

    def _notion_oauth_configured() -> bool:
        s = get_settings()
        return bool(s.notion_oauth_client_id and s.notion_oauth_client_secret)

    def _notion_oauth_unavailable() -> HTTPException:
        return HTTPException(
            status_code=503,
            detail=(
                "Notion public OAuth is not configured on this deployment. "
                "Set NOTION_OAUTH_CLIENT_ID, NOTION_OAUTH_CLIENT_SECRET and "
                "NOTION_OAUTH_REDIRECT_URI to enable OAuth. You can still "
                "connect via an internal integration token using PUT."
            ),
        )

    @app.get(
        "/api/v1/integrations/notion/authorize",
        response_model=NotionAuthorizeResponse,
    )
    async def notion_authorize(
        current_user: User = Depends(get_current_user),
    ) -> NotionAuthorizeResponse:
        """Return the Notion consent URL for the public OAuth flow."""
        if not _notion_oauth_configured():
            raise _notion_oauth_unavailable()
        from leadgen.integrations.notion_oauth import (
            StateValidationError,
            build_authorize_url,
            issue_state,
        )

        settings = get_settings()
        try:
            state = issue_state(
                current_user.id, secret=settings.auth_jwt_secret
            )
        except StateValidationError as exc:
            # Misconfigured deployment (no AUTH_JWT_SECRET). Surface as
            # 503 so ops sees the missing env var instead of a generic
            # 500.
            raise HTTPException(status_code=503, detail=str(exc)) from exc
        url = build_authorize_url(
            client_id=settings.notion_oauth_client_id,
            redirect_uri=settings.notion_oauth_redirect_uri,
            state=state,
        )
        return NotionAuthorizeResponse(url=url, state=state)

    @app.get("/api/v1/integrations/notion/callback")
    async def notion_callback(
        code: str = Query(..., min_length=10, max_length=512),
        state: str = Query(..., min_length=1, max_length=256),
        error: str | None = Query(default=None),
    ) -> Response:
        """OAuth callback — exchanges the code and saves the access token.

        On success, redirects to /app/settings?notion=connected so the
        user sees the "set your database" prompt.
        """
        settings = get_settings()
        return_base = settings.public_app_url.rstrip("/") + "/app/settings"

        if error:
            return Response(
                status_code=302,
                content="redirecting",
                headers={
                    "Location": f"{return_base}?notion=error&reason={error}"
                },
            )

        if not _notion_oauth_configured():
            raise _notion_oauth_unavailable()

        from leadgen.core.services.secrets_vault import encrypt
        from leadgen.integrations.notion_oauth import (
            NotionOAuthError,
            StateValidationError,
            exchange_code_for_token,
            verify_state,
        )

        try:
            user_id = verify_state(
                state, secret=settings.auth_jwt_secret
            )
        except StateValidationError as exc:
            # Malformed / forged / expired state. Don't reveal which —
            # uniform 400 prevents oracles. Logging captures the reason
            # for ops without leaking it to the caller.
            logger.warning(
                "notion_oauth: rejected callback state reason=%s",
                str(exc),
            )
            raise HTTPException(
                status_code=400, detail="invalid state"
            ) from exc

        try:
            token_data = await exchange_code_for_token(
                code,
                client_id=settings.notion_oauth_client_id,
                client_secret=settings.notion_oauth_client_secret,
                redirect_uri=settings.notion_oauth_redirect_uri,
            )
        except NotionOAuthError as exc:
            raise HTTPException(
                status_code=400, detail=f"oauth: {exc}"
            ) from exc

        ciphertext = encrypt(token_data.access_token)
        config: dict[str, Any] = {
            "workspace_id": token_data.workspace_id,
            "workspace_name": token_data.workspace_name,
            "auth_type": "oauth",
        }
        if token_data.owner_email:
            config["owner_email"] = token_data.owner_email

        async with session_factory() as session:
            existing = (
                await session.execute(
                    select(UserIntegrationCredential)
                    .where(UserIntegrationCredential.user_id == user_id)
                    .where(UserIntegrationCredential.provider == "notion")
                )
            ).scalar_one_or_none()
            if existing is None:
                row = UserIntegrationCredential(
                    user_id=user_id,
                    provider="notion",
                    token_ciphertext=ciphertext,
                    config=config,
                )
                session.add(row)
            else:
                existing.token_ciphertext = ciphertext
                # Preserve existing database_id if the user already set one.
                if existing.config and existing.config.get("database_id"):
                    config["database_id"] = existing.config["database_id"]
                existing.config = config
                existing.updated_at = datetime.now(timezone.utc)
            await session.commit()

        return Response(
            status_code=302,
            content="redirecting",
            headers={"Location": f"{return_base}?notion=connected"},
        )

    @app.get(
        "/api/v1/integrations/notion/databases",
        response_model=NotionDatabaseList,
    )
    async def list_notion_databases(
        current_user: User = Depends(get_current_user),
    ) -> NotionDatabaseList:
        """Surface databases the connected workspace has shared with us.

        Powers the in-Settings picker the SPA shows after OAuth. The
        legacy internal-token flow can call it too — by then the user
        already pasted ``database_id`` so it's mostly a courtesy there.
        """
        from leadgen.core.services.secrets_vault import decrypt
        from leadgen.integrations.notion import NotionClient, NotionError

        async with session_factory() as session:
            cred = (
                await session.execute(
                    select(UserIntegrationCredential)
                    .where(
                        UserIntegrationCredential.user_id == current_user.id
                    )
                    .where(UserIntegrationCredential.provider == "notion")
                )
            ).scalar_one_or_none()
        if cred is None:
            raise HTTPException(
                status_code=400, detail="Notion is not connected"
            )
        try:
            token = decrypt(cred.token_ciphertext)
        except ValueError as exc:
            raise HTTPException(
                status_code=400,
                detail="Saved Notion credentials are unreadable; reconnect.",
            ) from exc
        try:
            async with NotionClient(token) as client:
                results = await client.list_databases()
        except NotionError as exc:
            raise HTTPException(
                status_code=502, detail=f"notion: {exc}"
            ) from exc

        items: list[NotionDatabase] = []
        for db in results:
            db_id = db.get("id")
            if not db_id:
                continue
            title_blocks = db.get("title") or []
            title = "".join(
                str(b.get("plain_text") or "") for b in title_blocks
            ).strip() or "Без названия"
            icon_block = db.get("icon") or {}
            icon = icon_block.get("emoji") or (
                (icon_block.get("external") or {}).get("url")
                if icon_block.get("type") == "external"
                else None
            )
            items.append(
                NotionDatabase(
                    id=db_id,
                    title=title,
                    icon=icon,
                    url=db.get("url"),
                )
            )
        return NotionDatabaseList(items=items)

    @app.patch(
        "/api/v1/integrations/notion/database",
        response_model=NotionIntegrationStatus,
    )
    async def set_notion_database(
        body: NotionSetDatabaseRequest,
        current_user: User = Depends(get_current_user),
    ) -> NotionIntegrationStatus:
        """Set (or update) the database_id for an already-connected Notion account.

        Used after the OAuth flow completes — the token is already saved
        but the user hasn't chosen a target database yet.  Validates
        the database_id against the stored token before saving.
        """
        from leadgen.core.services.secrets_vault import decrypt, mask_token
        from leadgen.integrations.notion import NotionClient, NotionError

        async with session_factory() as session:
            row = (
                await session.execute(
                    select(UserIntegrationCredential)
                    .where(
                        UserIntegrationCredential.user_id == current_user.id
                    )
                    .where(UserIntegrationCredential.provider == "notion")
                )
            ).scalar_one_or_none()

        if row is None:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Notion is not connected yet. Connect via OAuth or "
                    "supply an internal token first."
                ),
            )

        try:
            token = decrypt(row.token_ciphertext)
        except ValueError as exc:
            raise HTTPException(
                status_code=400,
                detail="Saved token is unreadable — please reconnect Notion.",
            ) from exc

        database_id = body.database_id.strip()
        try:
            async with NotionClient(token) as client:
                schema = await client.get_database(database_id)
        except NotionError as exc:
            raise HTTPException(
                status_code=400,
                detail=locale_pick(
                    current_user.language_code,
                    ru=(
                        "Notion отказал в доступе к базе. Убедитесь что "
                        "интеграция/подключение имеет доступ к этой базе. "
                        f"Подробности: {exc}"
                    ),
                    uk=(
                        "Notion відмовив у доступі до бази. Переконайтеся, "
                        "що інтеграція/підключення має доступ до цієї бази. "
                        f"Деталі: {exc}"
                    ),
                    en=(
                        "Notion denied access to the database. Make sure "
                        "the integration/connection has access to it. "
                        f"Details: {exc}"
                    ),
                ),
            ) from exc

        db_title = (schema.get("title") or [{}])[0].get("plain_text") or None
        config = dict(row.config or {})
        config["database_id"] = database_id
        if db_title and not config.get("workspace_name"):
            config["workspace_name"] = db_title

        async with session_factory() as session:
            existing = (
                await session.execute(
                    select(UserIntegrationCredential)
                    .where(
                        UserIntegrationCredential.user_id == current_user.id
                    )
                    .where(UserIntegrationCredential.provider == "notion")
                )
            ).scalar_one_or_none()
            if existing is not None:
                existing.config = config
                existing.updated_at = datetime.now(timezone.utc)
                await session.commit()
                await session.refresh(existing)
                row = existing

        return NotionIntegrationStatus(
            connected=True,
            token_preview=mask_token(token),
            database_id=database_id,
            workspace_name=config.get("workspace_name"),
            updated_at=row.updated_at,
        )

    @app.get(
        "/api/v1/integrations/notion",
        response_model=NotionIntegrationStatus,
    )
    async def get_notion_integration(
        current_user: User = Depends(get_current_user),
    ) -> NotionIntegrationStatus:
        async with session_factory() as session:
            row = (
                await session.execute(
                    select(UserIntegrationCredential)
                    .where(UserIntegrationCredential.user_id == current_user.id)
                    .where(UserIntegrationCredential.provider == "notion")
                )
            ).scalar_one_or_none()
        if row is None:
            return NotionIntegrationStatus(connected=False)
        from leadgen.core.services.secrets_vault import decrypt, mask_token

        try:
            preview = mask_token(decrypt(row.token_ciphertext))
        except ValueError:
            preview = None  # key rotated; UI will offer reconnect
        config = row.config or {}
        return NotionIntegrationStatus(
            connected=True,
            token_preview=preview,
            database_id=config.get("database_id"),
            workspace_name=config.get("workspace_name"),
            owner_email=config.get("owner_email"),
            auth_type=config.get("auth_type", "internal"),
            updated_at=row.updated_at,
        )

    @app.put(
        "/api/v1/integrations/notion",
        response_model=NotionIntegrationStatus,
    )
    async def connect_notion(
        body: NotionConnectRequest,
        current_user: User = Depends(get_current_user),
    ) -> NotionIntegrationStatus:
        """Save (or replace) the user's Notion credentials.

        We immediately probe the database to validate the token has
        access — saving an unworkable credential would just give the
        user a misleading "connected" badge.
        """
        from leadgen.core.services.secrets_vault import encrypt, mask_token
        from leadgen.integrations.notion import NotionClient, NotionError

        token = body.token.strip()
        database_id = body.database_id.strip()
        try:
            async with NotionClient(token) as client:
                schema = await client.get_database(database_id)
        except NotionError as exc:
            raise HTTPException(
                status_code=400,
                detail=locale_pick(
                    current_user.language_code,
                    ru=(
                        "Notion отказал в доступе к базе. Проверьте что "
                        "интеграция share-нута на эту базу и токен "
                        f"актуален. Подробности: {exc}"
                    ),
                    uk=(
                        "Notion відмовив у доступі до бази. Перевірте, що "
                        "інтеграцію розшарено на цю базу і токен "
                        f"актуальний. Деталі: {exc}"
                    ),
                    en=(
                        "Notion denied access to the database. Check that "
                        "the integration is shared with this database and "
                        f"the token is valid. Details: {exc}"
                    ),
                ),
            ) from exc

        workspace_name = (schema.get("title") or [{}])[0].get(
            "plain_text"
        ) or None
        ciphertext = encrypt(token)
        config: dict[str, Any] = {
            "database_id": database_id,
            "workspace_name": workspace_name,
        }
        async with session_factory() as session:
            existing = (
                await session.execute(
                    select(UserIntegrationCredential)
                    .where(
                        UserIntegrationCredential.user_id == current_user.id
                    )
                    .where(UserIntegrationCredential.provider == "notion")
                )
            ).scalar_one_or_none()
            if existing is None:
                row = UserIntegrationCredential(
                    user_id=current_user.id,
                    provider="notion",
                    token_ciphertext=ciphertext,
                    config=config,
                )
                session.add(row)
            else:
                existing.token_ciphertext = ciphertext
                existing.config = config
                existing.updated_at = datetime.now(timezone.utc)
                row = existing
            await session.commit()
            await session.refresh(row)

        return NotionIntegrationStatus(
            connected=True,
            token_preview=mask_token(token),
            database_id=database_id,
            workspace_name=workspace_name,
            updated_at=row.updated_at,
        )

    @app.delete("/api/v1/integrations/notion")
    async def disconnect_notion(
        current_user: User = Depends(get_current_user),
    ) -> dict[str, bool]:
        async with session_factory() as session:
            row = (
                await session.execute(
                    select(UserIntegrationCredential)
                    .where(
                        UserIntegrationCredential.user_id == current_user.id
                    )
                    .where(UserIntegrationCredential.provider == "notion")
                )
            ).scalar_one_or_none()
            if row is not None:
                await session.delete(row)
                await session.commit()
        return {"ok": True}

    @app.post(
        "/api/v1/leads/export-to-notion",
        response_model=NotionExportResponse,
    )
    async def export_leads_to_notion(
        body: NotionExportRequest,
        current_user: User = Depends(get_current_user),
    ) -> NotionExportResponse:
        """Push a batch of selected leads as new pages in the user's database.

        Authorisation is per-lead — only leads the caller owns (or can
        see via team membership) get pushed. Per-lead failures inline
        as ``error`` so a misconfigured property doesn't sink the
        whole batch.
        """
        from leadgen.core.services.secrets_vault import decrypt
        from leadgen.integrations.notion import (
            NotionClient,
            NotionError,
            NotionExportRow,
            resolve_property_map,
            row_to_properties,
        )

        async with session_factory() as session:
            cred = (
                await session.execute(
                    select(UserIntegrationCredential)
                    .where(
                        UserIntegrationCredential.user_id == current_user.id
                    )
                    .where(UserIntegrationCredential.provider == "notion")
                )
            ).scalar_one_or_none()
            if cred is None:
                raise HTTPException(
                    status_code=400,
                    detail="Notion is not connected. Connect it in Settings → Интеграции.",
                )
            try:
                token = decrypt(cred.token_ciphertext)
            except ValueError as exc:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Saved Notion credentials are unreadable "
                        "(encryption key rotated). Reconnect in Settings."
                    ),
                ) from exc
            database_id = (cred.config or {}).get("database_id")
            if not database_id:
                raise HTTPException(
                    status_code=400,
                    detail="Notion database is not set; reconnect in Settings.",
                )

            # Lead authorisation join (same pattern as bulk-draft).
            lead_rows = (
                (
                    await session.execute(
                        select(Lead, SearchQuery)
                        .join(SearchQuery, SearchQuery.id == Lead.query_id)
                        .where(Lead.id.in_(list(body.lead_ids)))
                    )
                )
                .all()
            )
            authorised: dict[uuid.UUID, tuple[Lead, SearchQuery]] = {}
            for lead, search in lead_rows:
                if search.user_id == current_user.id:
                    authorised[lead.id] = (lead, search)
                    continue
                if search.team_id is not None and (
                    await _membership(session, search.team_id, current_user.id)
                ):
                    authorised[lead.id] = (lead, search)
            tags_by_lead = await _tags_by_lead(session, list(authorised))

        items: list[NotionExportItem] = []
        async with NotionClient(token) as client:
            try:
                schema = await client.get_database(database_id)
            except NotionError as exc:
                raise HTTPException(
                    status_code=502,
                    detail=f"Notion database is unreachable: {exc}",
                ) from exc
            mapping = resolve_property_map(schema)
            if "name" not in mapping:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Notion database must have a Title column. "
                        "Add one in Notion and try again."
                    ),
                )

            for lead_id in body.lead_ids:
                pair = authorised.get(lead_id)
                if pair is None:
                    items.append(
                        NotionExportItem(
                            lead_id=lead_id, error="not authorised"
                        )
                    )
                    continue
                lead, search = pair
                row = NotionExportRow(
                    name=lead.name or "(unnamed)",
                    score=int(round(lead.score_ai)) if lead.score_ai else None,
                    status=lead.lead_status,
                    rating=lead.rating,
                    reviews=lead.reviews_count,
                    phone=lead.phone,
                    website=lead.website,
                    address=lead.address,
                    category=lead.category,
                    notes=lead.notes,
                    niche=search.niche,
                    region=search.region,
                    tags=tuple(
                        tag.name
                        for tag in tags_by_lead.get(lead_id, ())
                    ),
                )
                properties = row_to_properties(row, mapping)
                try:
                    page = await client.create_page(
                        database_id=database_id, properties=properties
                    )
                    items.append(
                        NotionExportItem(
                            lead_id=lead_id,
                            notion_url=page.get("url"),
                        )
                    )
                except NotionError as exc:
                    logger.exception(
                        "notion export: failed for lead %s", lead_id
                    )
                    items.append(
                        NotionExportItem(lead_id=lead_id, error=str(exc)[:200])
                    )

        successes = sum(1 for it in items if it.notion_url)
        return NotionExportResponse(
            items=items,
            success_count=successes,
            failure_count=len(items) - successes,
        )

    # ── /api/v1/track — email open pixel (no auth) ─────────────────────

    _gif_pixel = bytes([
        0x47, 0x49, 0x46, 0x38, 0x39, 0x61, 0x01, 0x00, 0x01, 0x00,
        0x80, 0x00, 0x00, 0xff, 0xff, 0xff, 0x00, 0x00, 0x00, 0x21,
        0xf9, 0x04, 0x00, 0x00, 0x00, 0x00, 0x00, 0x2c, 0x00, 0x00,
        0x00, 0x00, 0x01, 0x00, 0x01, 0x00, 0x00, 0x02, 0x02, 0x44,
        0x01, 0x00, 0x3b,
    ])

    @app.get("/api/v1/track/{token}", include_in_schema=False)
    async def track_email_open(
        token: str,
        lead_id: str,
        user_id: int,
    ) -> Response:
        try:
            from leadgen.core.services.tracking import verify_track_token

            if verify_track_token(token, lead_id, str(user_id)):
                async with session_factory() as session:
                    lead = await session.get(Lead, uuid.UUID(lead_id))
                    if lead is not None and lead.deleted_at is None:
                        session.add(
                            LeadActivity(
                                lead_id=uuid.UUID(lead_id),
                                user_id=user_id,
                                kind="email_opened",
                                payload={},
                            )
                        )
                        await session.commit()
        except Exception:
            pass
        return Response(content=_gif_pixel, media_type="image/gif")

    # ── /api/v1/oauth/gmail (OAuth flow + send-as-user) ────────────────
    #
    # Stage-mode: empty GOOGLE_OAUTH_CLIENT_ID / _SECRET makes
    # /authorize, /callback and /leads/{id}/send-email respond 503,
    # leaving the rest of the API healthy.

    def _gmail_oauth_configured() -> bool:
        s = get_settings()
        return bool(s.google_oauth_client_id and s.google_oauth_client_secret)

    def _gmail_unavailable() -> HTTPException:
        return HTTPException(
            status_code=503,
            detail=(
                "Gmail OAuth is not configured on this deployment. "
                "Set GOOGLE_OAUTH_CLIENT_ID, GOOGLE_OAUTH_CLIENT_SECRET "
                "and GOOGLE_OAUTH_REDIRECT_URI to enable Gmail send."
            ),
        )

    @app.get(
        "/api/v1/oauth/gmail",
        response_model=GmailIntegrationStatus,
    )
    async def gmail_status(
        current_user: User = Depends(get_current_user),
    ) -> GmailIntegrationStatus:
        async with session_factory() as session:
            cred = (
                await session.execute(
                    select(OAuthCredential)
                    .where(OAuthCredential.user_id == current_user.id)
                    .where(OAuthCredential.provider == "gmail")
                )
            ).scalar_one_or_none()
        if cred is None:
            return GmailIntegrationStatus(connected=False)
        return GmailIntegrationStatus(
            connected=True,
            account_email=cred.account_email,
            scope=cred.scope,
            expires_at=cred.expires_at,
        )

    @app.get(
        "/api/v1/oauth/gmail/authorize",
        response_model=GmailAuthorizeResponse,
    )
    async def gmail_authorize(
        current_user: User = Depends(get_current_user),
    ) -> GmailAuthorizeResponse:
        """Mint a consent-screen URL the SPA redirects the user to.

        ``state`` is signed (HMAC-SHA256 over user_id + nonce + ts)
        with ``AUTH_JWT_SECRET`` so the callback can verify the user
        identity without a session-side nonce store. The shared helper
        in ``core.services.oauth_state`` is also used by Notion and
        Outlook.
        """
        if not _gmail_oauth_configured():
            raise _gmail_unavailable()
        from leadgen.core.services.oauth_state import (
            StateValidationError,
            issue_state,
        )
        from leadgen.integrations.gmail import build_authorize_url

        settings = get_settings()
        try:
            state = issue_state(
                current_user.id, secret=settings.auth_jwt_secret
            )
        except StateValidationError as exc:
            # Misconfigured deployment (no AUTH_JWT_SECRET). Surface
            # as 503 so ops sees the missing env var instead of a
            # generic 500.
            raise HTTPException(status_code=503, detail=str(exc)) from exc
        url = build_authorize_url(
            client_id=settings.google_oauth_client_id,
            redirect_uri=settings.google_oauth_redirect_uri,
            state=state,
        )
        return GmailAuthorizeResponse(url=url, state=state)

    @app.get("/api/v1/oauth/gmail/callback")
    async def gmail_callback(
        code: str = Query(..., min_length=10, max_length=512),
        state: str = Query(..., min_length=1, max_length=512),
    ) -> Response:
        """Receive Google's callback, exchange the code, store tokens.

        We don't go through ``get_current_user`` here because Google
        bounces back without our session cookie when the consent
        happens in a fresh browser context. The user-id is recovered
        from ``state``, which is HMAC-signed — so a forged
        ``"<victim_id>:..."`` callback can't write the attacker's
        Gmail token under the victim's account.
        """
        if not _gmail_oauth_configured():
            raise _gmail_unavailable()
        from leadgen.core.services.oauth_state import (
            StateValidationError,
            verify_state,
        )
        from leadgen.core.services.oauth_store import save_tokens
        from leadgen.integrations.gmail import (
            GmailError,
            exchange_code_for_tokens,
            fetch_account_email,
        )

        settings = get_settings()
        try:
            user_id = verify_state(
                state, secret=settings.auth_jwt_secret
            )
        except StateValidationError as exc:
            logger.warning(
                "gmail_oauth: rejected callback state reason=%s",
                str(exc),
            )
            raise HTTPException(
                status_code=400, detail="invalid state"
            ) from exc

        try:
            tokens = await exchange_code_for_tokens(
                code,
                client_id=settings.google_oauth_client_id,
                client_secret=settings.google_oauth_client_secret,
                redirect_uri=settings.google_oauth_redirect_uri,
            )
        except GmailError as exc:
            raise HTTPException(
                status_code=400, detail=f"oauth: {exc}"
            ) from exc

        account_email = await fetch_account_email(tokens.access_token)
        async with session_factory() as session:
            await save_tokens(
                session,
                user_id=user_id,
                provider="gmail",
                tokens=tokens,
                account_email=account_email,
            )

        # Bounce back to the Settings page where the user kicked the
        # flow off. ``PUBLIC_APP_URL`` is the canonical front-end
        # origin so the redirect lands in their browser tab.
        return_to = (
            settings.public_app_url.rstrip("/") + "/app/settings?gmail=connected"
        )
        return Response(
            status_code=302,
            content="redirecting",
            headers={"Location": return_to},
        )

    @app.delete("/api/v1/oauth/gmail")
    async def gmail_disconnect(
        current_user: User = Depends(get_current_user),
    ) -> dict[str, bool]:
        async with session_factory() as session:
            cred = (
                await session.execute(
                    select(OAuthCredential)
                    .where(OAuthCredential.user_id == current_user.id)
                    .where(OAuthCredential.provider == "gmail")
                )
            ).scalar_one_or_none()
            if cred is not None:
                await session.delete(cred)
                await session.commit()
        return {"ok": True}

    @app.post(
        "/api/v1/leads/{lead_id}/send-email",
        response_model=GmailSendResponse,
    )
    async def gmail_send_email(
        lead_id: uuid.UUID,
        body: GmailSendRequest,
        current_user: User = Depends(get_current_user),
    ) -> GmailSendResponse:
        """Send an email through the user's Gmail or Outlook account.

        Provider is selected via ``body.provider`` (default: gmail).
        Both providers log a ``LeadActivity`` of kind="email_sent" so
        the timeline on the lead modal shows the message went out.
        Body is truncated to 4000 chars in the activity record so the
        JSONB column doesn't bloat over time.
        """
        provider = (body.provider or "gmail").lower()
        if provider == "gmail" and not _gmail_oauth_configured():
            raise _gmail_unavailable()
        if provider == "outlook" and not _outlook_oauth_configured():
            raise _outlook_unavailable()
        from leadgen.core.services.oauth_store import (
            OAuthStoreError,
            ensure_fresh_token,
        )

        async with session_factory() as session:
            lead = await session.get(Lead, lead_id)
            if lead is None or lead.deleted_at is not None:
                raise HTTPException(status_code=404, detail="lead not found")
            # Use the explicit override or pull the first email out of
            # the website-meta blob; fail loudly if neither is set.
            recipient = body.to or _extract_lead_email(lead)
            if not recipient:
                raise HTTPException(
                    status_code=400,
                    detail="lead has no email address on file",
                )

            try:
                fresh = await ensure_fresh_token(
                    session, user_id=current_user.id, provider=provider
                )
            except OAuthStoreError as exc:
                raise HTTPException(
                    status_code=400, detail=str(exc)
                ) from exc

            from_addr = fresh.account_email or current_user.email or ""
            if not from_addr:
                raise HTTPException(
                    status_code=400,
                    detail="cannot determine sender address",
                )

            from leadgen.core.services.tracking import generate_track_token

            _track_token = generate_track_token(
                str(lead_id), str(current_user.id)
            )
            _base = get_settings().public_app_url.rstrip("/")
            _pixel_url = (
                f"{_base}/api/v1/track/{_track_token}"
                f"?lead_id={lead_id}&user_id={current_user.id}"
            )
            _html_body = (
                f"<p>{body.body}</p>"
                f'<img src="{_pixel_url}" width="1" height="1"'
                f' style="display:none" alt="">'
            )

            message_id: str | None = None
            thread_id: str | None = None
            if provider == "gmail":
                from leadgen.integrations.gmail import (
                    GmailError,
                    build_raw_message,
                    send_message,
                )

                raw = build_raw_message(
                    from_addr=from_addr,
                    to_addr=recipient,
                    subject=body.subject,
                    body=body.body,
                    html_body=_html_body,
                )
                try:
                    resp = await send_message(
                        access_token=fresh.access_token, raw_message=raw
                    )
                except GmailError as exc:
                    raise HTTPException(
                        status_code=502,
                        detail=f"gmail send failed: {exc}",
                    ) from exc
                message_id = resp.get("id")
                thread_id = resp.get("threadId")
            else:  # outlook
                from leadgen.integrations.outlook import (
                    OutlookError,
                )
                from leadgen.integrations.outlook import (
                    send_message as outlook_send,
                )

                try:
                    await outlook_send(
                        access_token=fresh.access_token,
                        from_addr=from_addr,
                        to_addr=recipient,
                        subject=body.subject,
                        body=body.body,
                        html_body=_html_body,
                    )
                except OutlookError as exc:
                    raise HTTPException(
                        status_code=502,
                        detail=f"outlook send failed: {exc}",
                    ) from exc
                # Microsoft Graph's sendMail returns 202 + empty body —
                # we don't get a message id back. Stamp the activity
                # with a synthetic provider-prefixed sentinel so the
                # reply tracker can still tell which provider sent it
                # even without a real message id.
                message_id = None
                thread_id = None

            now = datetime.now(timezone.utc)
            activity = LeadActivity(
                lead_id=lead_id,
                user_id=current_user.id,
                kind="email_sent",
                payload={
                    "to": recipient,
                    "subject": body.subject[:255],
                    "body": body.body[:4000],
                    "message_id": message_id,
                    "thread_id": thread_id,
                    "provider": provider,
                },
                created_at=now,
            )
            session.add(activity)
            lead.last_touched_at = now
            if lead.lead_status == "new":
                lead.lead_status = "contacted"
            await session.commit()

        return GmailSendResponse(
            message_id=message_id or "",
            thread_id=thread_id,
            sent_at=now,
        )

    @app.post("/api/v1/leads/bulk-send-email")
    async def bulk_send_email(
        data: BulkSendRequest,
        current_user: User = Depends(get_current_user),
    ) -> dict:
        """Send personalized emails to multiple leads via Henry.

        Rate-limited to 1 email per 2 seconds. Max 50 leads per call.
        Returns summary of sent/failed counts.
        """
        provider = (data.provider or "gmail").lower()
        if provider == "gmail" and not _gmail_oauth_configured():
            raise _gmail_unavailable()
        if provider == "outlook" and not _outlook_oauth_configured():
            raise _outlook_unavailable()

        from leadgen.core.services.oauth_store import (
            OAuthStoreError,
            ensure_fresh_token,
        )

        lead_ids = data.lead_ids[:50]
        if not lead_ids:
            raise HTTPException(status_code=400, detail="No lead IDs provided")

        analyzer = AIAnalyzer()
        sent = 0
        failed = 0
        errors: list[str] = []

        async with session_factory() as session:
            try:
                fresh = await ensure_fresh_token(
                    session, user_id=current_user.id, provider=provider
                )
            except OAuthStoreError as exc:
                raise HTTPException(
                    status_code=400, detail=str(exc)
                ) from exc

            from_addr = fresh.account_email or current_user.email or ""
            if not from_addr:
                raise HTTPException(
                    status_code=400,
                    detail="cannot determine sender address",
                )

            for lead_id_str in lead_ids:
                try:
                    try:
                        lead_uuid = uuid.UUID(lead_id_str)
                    except (ValueError, AttributeError):
                        failed += 1
                        errors.append(f"{lead_id_str}: invalid id")
                        continue

                    lead = await session.get(Lead, lead_uuid)
                    if lead is None or lead.deleted_at is not None:
                        failed += 1
                        continue

                    # Authorize via the parent search query owner.
                    sq = await session.get(SearchQuery, lead.query_id)
                    if sq is None or sq.user_id != current_user.id:
                        failed += 1
                        continue

                    recipient = _extract_lead_email(lead)
                    if not recipient:
                        failed += 1
                        errors.append(
                            locale_pick(
                                current_user.language_code,
                                ru=f"{lead.name}: нет email",
                                uk=f"{lead.name}: немає email",
                                en=f"{lead.name}: no email",
                            )
                        )
                        continue

                    email_draft = await analyzer.generate_cold_email(
                        lead={
                            "name": lead.name,
                            "category": lead.category,
                            "address": lead.address,
                            "website": lead.website,
                            "rating": lead.rating,
                            "reviews_count": lead.reviews_count,
                            "tags": lead.tags or [],
                            "summary": lead.summary,
                            "advice": lead.advice,
                        },
                        user_profile={
                            "display_name": current_user.display_name,
                            "email": current_user.email,
                            "language_code": current_user.language_code,
                            "calendly_url": getattr(
                                current_user, "calendly_url", None
                            ),
                            "icp_profile": getattr(
                                current_user, "icp_profile", None
                            ),
                        },
                    )
                    subject = (email_draft.get("subject") or "").strip() or (
                        locale_pick(
                            current_user.language_code,
                            ru=f"Привет от {current_user.display_name or 'нас'}",
                            uk=f"Привіт від {current_user.display_name or 'нас'}",
                            en=f"Hello from {current_user.display_name or 'us'}",
                        )
                    )
                    body_text = email_draft.get("body") or ""
                    html_body = f"<p>{body_text}</p>"

                    if provider == "gmail":
                        from leadgen.integrations.gmail import (
                            GmailError,
                            build_raw_message,
                            send_message,
                        )

                        raw = build_raw_message(
                            from_addr=from_addr,
                            to_addr=recipient,
                            subject=subject,
                            body=body_text,
                            html_body=html_body,
                        )
                        try:
                            await send_message(
                                access_token=fresh.access_token,
                                raw_message=raw,
                            )
                        except GmailError as exc:
                            failed += 1
                            errors.append(f"{lead.name}: {str(exc)[:50]}")
                            continue
                    else:
                        from leadgen.integrations.outlook import (
                            OutlookError,
                        )
                        from leadgen.integrations.outlook import (
                            send_message as outlook_send,
                        )

                        try:
                            await outlook_send(
                                access_token=fresh.access_token,
                                from_addr=from_addr,
                                to_addr=recipient,
                                subject=subject,
                                body=body_text,
                                html_body=html_body,
                            )
                        except OutlookError as exc:
                            failed += 1
                            errors.append(f"{lead.name}: {str(exc)[:50]}")
                            continue

                    now = datetime.now(timezone.utc)
                    session.add(
                        LeadActivity(
                            lead_id=lead.id,
                            user_id=current_user.id,
                            kind="email_sent",
                            payload={
                                "to": recipient,
                                "subject": subject[:255],
                                "body": body_text[:4000],
                                "provider": provider,
                                "bulk": True,
                            },
                            created_at=now,
                        )
                    )
                    lead.last_touched_at = now
                    if lead.lead_status == "new":
                        lead.lead_status = "contacted"
                    sent += 1
                    await asyncio.sleep(2)
                except Exception as exc:  # noqa: BLE001
                    logger.warning(
                        "bulk_send: failed lead_id=%s err=%s",
                        lead_id_str,
                        exc,
                    )
                    failed += 1
                    errors.append(f"{lead_id_str}: {str(exc)[:50]}")

            await session.commit()

        return {"sent": sent, "failed": failed, "errors": errors[:10]}

    # ── /api/v1/oauth/outlook (OAuth flow + send-as-user mirror) ───────
    #
    # Same shape as Gmail: status / authorize / callback / delete. The
    # send endpoint is the same /leads/{id}/send-email — it picks the
    # provider from the body. 503-safe when Outlook env vars are unset.

    def _outlook_oauth_configured() -> bool:
        s = get_settings()
        return bool(
            s.outlook_oauth_client_id and s.outlook_oauth_client_secret
        )

    def _outlook_unavailable() -> HTTPException:
        return HTTPException(
            status_code=503,
            detail=(
                "Outlook OAuth is not configured on this deployment. "
                "Set OUTLOOK_OAUTH_CLIENT_ID, OUTLOOK_OAUTH_CLIENT_SECRET "
                "and OUTLOOK_OAUTH_REDIRECT_URI to enable Outlook send."
            ),
        )

    @app.get(
        "/api/v1/oauth/outlook",
        response_model=OutlookIntegrationStatus,
    )
    async def outlook_status(
        current_user: User = Depends(get_current_user),
    ) -> OutlookIntegrationStatus:
        async with session_factory() as session:
            cred = (
                await session.execute(
                    select(OAuthCredential)
                    .where(OAuthCredential.user_id == current_user.id)
                    .where(OAuthCredential.provider == "outlook")
                )
            ).scalar_one_or_none()
        if cred is None:
            return OutlookIntegrationStatus(connected=False)
        return OutlookIntegrationStatus(
            connected=True,
            account_email=cred.account_email,
            scope=cred.scope,
            expires_at=cred.expires_at,
        )

    @app.get(
        "/api/v1/oauth/outlook/authorize",
        response_model=OutlookAuthorizeResponse,
    )
    async def outlook_authorize(
        current_user: User = Depends(get_current_user),
    ) -> OutlookAuthorizeResponse:
        """Mint a Microsoft consent-screen URL.

        Uses the shared HMAC-signed ``oauth_state`` helper so the
        callback can verify the state parameter without a DB-backed
        nonce table — and forged ``"<victim_id>:..."`` callbacks are
        rejected.
        """
        if not _outlook_oauth_configured():
            raise _outlook_unavailable()
        from leadgen.core.services.oauth_state import (
            StateValidationError,
            issue_state,
        )
        from leadgen.integrations.outlook import build_authorize_url

        settings = get_settings()
        try:
            state = issue_state(
                current_user.id, secret=settings.auth_jwt_secret
            )
        except StateValidationError as exc:
            raise HTTPException(status_code=503, detail=str(exc)) from exc
        url = build_authorize_url(
            client_id=settings.outlook_oauth_client_id,
            redirect_uri=settings.outlook_oauth_redirect_uri,
            state=state,
        )
        return OutlookAuthorizeResponse(url=url, state=state)

    @app.get("/api/v1/oauth/outlook/callback")
    async def outlook_callback(
        code: str = Query(..., min_length=10, max_length=2048),
        state: str = Query(..., min_length=1, max_length=512),
        error: str | None = Query(default=None),
    ) -> Response:
        """Microsoft Graph callback — exchanges code, stores tokens.

        On success redirects to /app/settings/integrations?outlook=connected
        so the SPA can render the post-connect state. On state-mismatch
        or token-exchange failure redirects with an error flag.
        """
        settings = get_settings()
        return_base = (
            settings.public_app_url.rstrip("/")
            + "/app/settings/integrations"
        )

        if error:
            return Response(
                status_code=302,
                content="redirecting",
                headers={
                    "Location": f"{return_base}?outlook=error&reason={error}"
                },
            )

        if not _outlook_oauth_configured():
            raise _outlook_unavailable()

        from leadgen.core.services.oauth_state import (
            StateValidationError,
            verify_state,
        )
        from leadgen.core.services.oauth_store import save_tokens
        from leadgen.integrations.gmail import TokenSet  # shared shape
        from leadgen.integrations.outlook import (
            OutlookError,
            exchange_code_for_tokens,
            fetch_account_email,
        )

        try:
            user_id = verify_state(
                state, secret=settings.auth_jwt_secret
            )
        except StateValidationError as exc:
            logger.warning(
                "outlook_oauth: rejected callback state reason=%s",
                str(exc),
            )
            raise HTTPException(
                status_code=400, detail="invalid state"
            ) from exc

        try:
            ms_tokens = await exchange_code_for_tokens(
                code,
                client_id=settings.outlook_oauth_client_id,
                client_secret=settings.outlook_oauth_client_secret,
                redirect_uri=settings.outlook_oauth_redirect_uri,
            )
        except OutlookError as exc:
            raise HTTPException(
                status_code=400, detail=f"oauth: {exc}"
            ) from exc

        account_email = await fetch_account_email(ms_tokens.access_token)

        # Re-shape into the shared TokenSet so save_tokens stays
        # provider-agnostic. The two dataclasses have identical fields;
        # this is a typing nicety, not a behaviour change.
        unified = TokenSet(
            access_token=ms_tokens.access_token,
            refresh_token=ms_tokens.refresh_token,
            expires_at=ms_tokens.expires_at,
            scope=ms_tokens.scope,
        )
        async with session_factory() as session:
            await save_tokens(
                session,
                user_id=user_id,
                provider="outlook",
                tokens=unified,
                account_email=account_email,
            )

        return Response(
            status_code=302,
            content="redirecting",
            headers={"Location": f"{return_base}?outlook=connected"},
        )

    @app.delete("/api/v1/oauth/outlook")
    async def outlook_disconnect(
        current_user: User = Depends(get_current_user),
    ) -> dict[str, bool]:
        async with session_factory() as session:
            cred = (
                await session.execute(
                    select(OAuthCredential)
                    .where(OAuthCredential.user_id == current_user.id)
                    .where(OAuthCredential.provider == "outlook")
                )
            ).scalar_one_or_none()
            if cred is not None:
                await session.delete(cred)
                await session.commit()
        return {"ok": True}

    # ── /api/v1/integrations/hubspot (OAuth + push-to-CRM) ─────────────
    #
    # Stage-mode: empty HUBSPOT_OAUTH_CLIENT_ID / _SECRET makes every
    # endpoint below respond 503 — the rest of the API stays usable.

    def _hubspot_oauth_configured() -> bool:
        s = get_settings()
        return bool(
            s.hubspot_oauth_client_id and s.hubspot_oauth_client_secret
        )

    def _hubspot_unavailable() -> HTTPException:
        return HTTPException(
            status_code=503,
            detail=(
                "HubSpot OAuth is not configured on this deployment. "
                "Set HUBSPOT_OAUTH_CLIENT_ID, HUBSPOT_OAUTH_CLIENT_SECRET "
                "and HUBSPOT_OAUTH_REDIRECT_URI to enable HubSpot."
            ),
        )

    @app.get(
        "/api/v1/integrations/hubspot",
        response_model=HubspotIntegrationStatus,
    )
    async def hubspot_status(
        current_user: User = Depends(get_current_user),
    ) -> HubspotIntegrationStatus:
        async with session_factory() as session:
            cred = (
                await session.execute(
                    select(OAuthCredential)
                    .where(OAuthCredential.user_id == current_user.id)
                    .where(OAuthCredential.provider == "hubspot")
                )
            ).scalar_one_or_none()
        if cred is None:
            return HubspotIntegrationStatus(connected=False)
        # Portal id is appended to the scope string on connect as
        # ``portal:<id>`` so we don't have to widen the OAuth schema
        # for one provider; recover it on read.
        portal_id: int | None = None
        for token in (cred.scope or "").split():
            if token.startswith("portal:"):
                try:
                    portal_id = int(token.split(":", 1)[1])
                except (ValueError, IndexError):
                    portal_id = None
                break
        return HubspotIntegrationStatus(
            connected=True,
            portal_id=portal_id,
            account_email=cred.account_email,
            scope=cred.scope,
            expires_at=cred.expires_at,
        )

    @app.get(
        "/api/v1/integrations/hubspot/authorize",
        response_model=HubspotAuthorizeResponse,
    )
    async def hubspot_authorize(
        current_user: User = Depends(get_current_user),
    ) -> HubspotAuthorizeResponse:
        if not _hubspot_oauth_configured():
            raise _hubspot_unavailable()
        from leadgen.core.services.oauth_state import (
            StateValidationError,
            issue_state,
        )
        from leadgen.integrations.hubspot import build_authorize_url

        settings = get_settings()
        try:
            state = issue_state(
                current_user.id, secret=settings.auth_jwt_secret
            )
        except StateValidationError as exc:
            raise HTTPException(status_code=503, detail=str(exc)) from exc
        url = build_authorize_url(
            client_id=settings.hubspot_oauth_client_id,
            redirect_uri=settings.hubspot_oauth_redirect_uri,
            state=state,
        )
        return HubspotAuthorizeResponse(url=url, state=state)

    @app.get("/api/v1/integrations/hubspot/callback")
    async def hubspot_callback(
        code: str = Query(..., min_length=10, max_length=512),
        state: str = Query(..., min_length=1, max_length=512),
    ) -> Response:
        if not _hubspot_oauth_configured():
            raise _hubspot_unavailable()
        from leadgen.core.services.oauth_state import (
            StateValidationError,
            verify_state,
        )
        from leadgen.core.services.oauth_store import save_tokens
        from leadgen.integrations.gmail import TokenSet
        from leadgen.integrations.hubspot import (
            HubspotError,
            exchange_code_for_tokens,
            fetch_token_info,
        )

        settings = get_settings()
        try:
            user_id = verify_state(
                state, secret=settings.auth_jwt_secret
            )
        except StateValidationError as exc:
            logger.warning(
                "hubspot_oauth: rejected callback state reason=%s",
                str(exc),
            )
            raise HTTPException(
                status_code=400, detail="invalid state"
            ) from exc
        try:
            tokens = await exchange_code_for_tokens(
                code,
                client_id=settings.hubspot_oauth_client_id,
                client_secret=settings.hubspot_oauth_client_secret,
                redirect_uri=settings.hubspot_oauth_redirect_uri,
            )
        except HubspotError as exc:
            raise HTTPException(
                status_code=400, detail=f"oauth: {exc}"
            ) from exc

        # Try to enrich with portal id + user email; failure is fine.
        portal_id = tokens.portal_id
        account_email: str | None = None
        try:
            info = await fetch_token_info(tokens.access_token)
            portal_id = portal_id or info.get("hub_id")
            account_email = info.get("user")
        except HubspotError:
            pass

        # Stuff portal id into the scope string so we don't have to
        # widen the OAuthCredential schema for one provider.
        scope_with_portal = tokens.scope or ""
        if portal_id is not None:
            scope_with_portal = (
                f"{scope_with_portal} portal:{portal_id}".strip()
            )
        save_payload = TokenSet(
            access_token=tokens.access_token,
            refresh_token=tokens.refresh_token,
            expires_at=tokens.expires_at,
            scope=scope_with_portal or None,
        )
        async with session_factory() as session:
            await save_tokens(
                session,
                user_id=user_id,
                provider="hubspot",
                tokens=save_payload,
                account_email=account_email,
            )

        return_to = (
            settings.public_app_url.rstrip("/")
            + "/app/settings?hubspot=connected"
        )
        return Response(
            status_code=302,
            content="redirecting",
            headers={"Location": return_to},
        )

    @app.delete("/api/v1/integrations/hubspot")
    async def hubspot_disconnect(
        current_user: User = Depends(get_current_user),
    ) -> dict[str, bool]:
        async with session_factory() as session:
            cred = (
                await session.execute(
                    select(OAuthCredential)
                    .where(OAuthCredential.user_id == current_user.id)
                    .where(OAuthCredential.provider == "hubspot")
                )
            ).scalar_one_or_none()
            if cred is not None:
                await session.delete(cred)
                await session.commit()
        return {"ok": True}

    @app.post(
        "/api/v1/leads/export-to-hubspot",
        response_model=HubspotExportResponse,
    )
    async def export_leads_to_hubspot(
        body: HubspotExportRequest,
        current_user: User = Depends(get_current_user),
    ) -> HubspotExportResponse:
        """Push a batch of leads into the user's HubSpot portal as contacts.

        Authorisation matches the Notion export: each lead must belong
        to the caller (or to a team they're a member of). Per-lead
        failures inline as ``error`` strings so a single bad row
        doesn't sink the whole batch.
        """
        if not _hubspot_oauth_configured():
            raise _hubspot_unavailable()
        from leadgen.core.services.oauth_store import (
            OAuthStoreError,
            ensure_fresh_token,
        )
        from leadgen.integrations.hubspot import (
            HubspotClient,
            HubspotContactInput,
            HubspotError,
            split_full_name,
        )

        async with session_factory() as session:
            try:
                fresh = await ensure_fresh_token(
                    session, user_id=current_user.id, provider="hubspot"
                )
            except OAuthStoreError as exc:
                raise HTTPException(
                    status_code=400, detail=str(exc)
                ) from exc

            lead_rows = (
                await session.execute(
                    select(Lead, SearchQuery)
                    .join(SearchQuery, SearchQuery.id == Lead.query_id)
                    .where(Lead.id.in_(list(body.lead_ids)))
                )
            ).all()
            authorised: dict[uuid.UUID, tuple[Lead, SearchQuery]] = {}
            for lead, search in lead_rows:
                if search.user_id == current_user.id:
                    authorised[lead.id] = (lead, search)
                    continue
                if search.team_id is not None and (
                    await _membership(
                        session, search.team_id, current_user.id
                    )
                ):
                    authorised[lead.id] = (lead, search)

        items: list[HubspotExportItem] = []
        async with HubspotClient(fresh.access_token) as client:
            for lead_id in body.lead_ids:
                pair = authorised.get(lead_id)
                if pair is None:
                    items.append(
                        HubspotExportItem(
                            lead_id=lead_id, error="not authorised"
                        )
                    )
                    continue
                lead, search = pair
                email = _extract_lead_email(lead)
                if not email:
                    items.append(
                        HubspotExportItem(
                            lead_id=lead_id,
                            error="lead has no email on file",
                        )
                    )
                    continue
                first, last = split_full_name(lead.name)
                contact = HubspotContactInput(
                    email=email,
                    firstname=first,
                    lastname=last,
                    phone=lead.phone,
                    company=lead.name,
                    website=lead.website,
                    city=search.region,
                    convioo_score=lead.score_ai,
                    convioo_status=lead.lead_status,
                )
                try:
                    contact_id = await client.upsert_contact(contact)
                    items.append(
                        HubspotExportItem(
                            lead_id=lead_id, contact_id=contact_id
                        )
                    )
                except HubspotError as exc:
                    logger.exception(
                        "hubspot export: failed for lead %s", lead_id
                    )
                    items.append(
                        HubspotExportItem(
                            lead_id=lead_id, error=str(exc)[:200]
                        )
                    )

        successes = sum(1 for it in items if it.contact_id)

        # LeadActivity rows so the timeline shows the export.
        if successes:
            now = datetime.now(timezone.utc)
            async with session_factory() as session:
                for it in items:
                    if not it.contact_id:
                        continue
                    session.add(
                        LeadActivity(
                            lead_id=it.lead_id,
                            user_id=current_user.id,
                            kind="exported_hubspot",
                            payload={"contact_id": it.contact_id},
                            created_at=now,
                        )
                    )
                await session.commit()

        return HubspotExportResponse(
            items=items,
            success_count=successes,
            failure_count=len(items) - successes,
        )

    # ── /api/v1/integrations/pipedrive (OAuth + push-to-CRM) ───────────
    #
    # Stage-mode: empty PIPEDRIVE_OAUTH_CLIENT_ID / _SECRET makes every
    # endpoint below respond 503 — the rest of the API stays usable.

    def _pipedrive_oauth_configured() -> bool:
        s = get_settings()
        return bool(
            s.pipedrive_oauth_client_id and s.pipedrive_oauth_client_secret
        )

    def _pipedrive_unavailable() -> HTTPException:
        return HTTPException(
            status_code=503,
            detail=(
                "Pipedrive OAuth is not configured on this deployment. "
                "Set PIPEDRIVE_OAUTH_CLIENT_ID, "
                "PIPEDRIVE_OAUTH_CLIENT_SECRET and "
                "PIPEDRIVE_OAUTH_REDIRECT_URI to enable Pipedrive."
            ),
        )

    async def _pipedrive_credential(
        session, user_id: int
    ) -> tuple[OAuthCredential, UserIntegrationCredential | None]:
        oauth = (
            await session.execute(
                select(OAuthCredential)
                .where(OAuthCredential.user_id == user_id)
                .where(OAuthCredential.provider == "pipedrive")
            )
        ).scalar_one_or_none()
        if oauth is None:
            return None, None  # type: ignore[return-value]
        cfg = (
            await session.execute(
                select(UserIntegrationCredential)
                .where(
                    UserIntegrationCredential.user_id == user_id
                )
                .where(UserIntegrationCredential.provider == "pipedrive")
            )
        ).scalar_one_or_none()
        return oauth, cfg

    def _pipedrive_api_domain(scope: str | None) -> str | None:
        for token in (scope or "").split():
            if token.startswith("api_domain:"):
                return token.split(":", 1)[1]
        return None

    @app.get(
        "/api/v1/integrations/pipedrive",
        response_model=PipedriveIntegrationStatus,
    )
    async def pipedrive_status(
        current_user: User = Depends(get_current_user),
    ) -> PipedriveIntegrationStatus:
        async with session_factory() as session:
            oauth, cfg = await _pipedrive_credential(
                session, current_user.id
            )
        if oauth is None:
            return PipedriveIntegrationStatus(connected=False)
        config = (cfg.config if cfg is not None else {}) or {}
        return PipedriveIntegrationStatus(
            connected=True,
            api_domain=_pipedrive_api_domain(oauth.scope),
            account_email=oauth.account_email,
            scope=oauth.scope,
            expires_at=oauth.expires_at,
            default_pipeline_id=config.get("default_pipeline_id"),
            default_stage_id=config.get("default_stage_id"),
        )

    @app.get(
        "/api/v1/integrations/pipedrive/authorize",
        response_model=PipedriveAuthorizeResponse,
    )
    async def pipedrive_authorize(
        current_user: User = Depends(get_current_user),
    ) -> PipedriveAuthorizeResponse:
        if not _pipedrive_oauth_configured():
            raise _pipedrive_unavailable()
        from leadgen.core.services.oauth_state import (
            StateValidationError,
            issue_state,
        )
        from leadgen.integrations.pipedrive import build_authorize_url

        settings = get_settings()
        try:
            state = issue_state(
                current_user.id, secret=settings.auth_jwt_secret
            )
        except StateValidationError as exc:
            raise HTTPException(status_code=503, detail=str(exc)) from exc
        url = build_authorize_url(
            client_id=settings.pipedrive_oauth_client_id,
            redirect_uri=settings.pipedrive_oauth_redirect_uri,
            state=state,
        )
        return PipedriveAuthorizeResponse(url=url, state=state)

    @app.get("/api/v1/integrations/pipedrive/callback")
    async def pipedrive_callback(
        code: str = Query(..., min_length=10, max_length=512),
        state: str = Query(..., min_length=1, max_length=512),
    ) -> Response:
        if not _pipedrive_oauth_configured():
            raise _pipedrive_unavailable()
        from leadgen.core.services.oauth_state import (
            StateValidationError,
            verify_state,
        )
        from leadgen.core.services.oauth_store import save_tokens
        from leadgen.integrations.gmail import TokenSet
        from leadgen.integrations.pipedrive import (
            PipedriveError,
            exchange_code_for_tokens,
        )

        settings = get_settings()
        try:
            user_id = verify_state(
                state, secret=settings.auth_jwt_secret
            )
        except StateValidationError as exc:
            logger.warning(
                "pipedrive_oauth: rejected callback state reason=%s",
                str(exc),
            )
            raise HTTPException(
                status_code=400, detail="invalid state"
            ) from exc
        try:
            tokens = await exchange_code_for_tokens(
                code,
                client_id=settings.pipedrive_oauth_client_id,
                client_secret=settings.pipedrive_oauth_client_secret,
                redirect_uri=settings.pipedrive_oauth_redirect_uri,
            )
        except PipedriveError as exc:
            raise HTTPException(
                status_code=400, detail=f"oauth: {exc}"
            ) from exc

        scope_with_domain = tokens.scope or ""
        if tokens.api_domain:
            scope_with_domain = (
                f"{scope_with_domain} api_domain:{tokens.api_domain}".strip()
            )
        save_payload = TokenSet(
            access_token=tokens.access_token,
            refresh_token=tokens.refresh_token,
            expires_at=tokens.expires_at,
            scope=scope_with_domain or None,
        )
        async with session_factory() as session:
            await save_tokens(
                session,
                user_id=user_id,
                provider="pipedrive",
                tokens=save_payload,
                account_email=tokens.account_email,
            )

        return_to = (
            settings.public_app_url.rstrip("/")
            + "/app/settings?pipedrive=connected"
        )
        return Response(
            status_code=302,
            content="redirecting",
            headers={"Location": return_to},
        )

    @app.delete("/api/v1/integrations/pipedrive")
    async def pipedrive_disconnect(
        current_user: User = Depends(get_current_user),
    ) -> dict[str, bool]:
        async with session_factory() as session:
            oauth, cfg = await _pipedrive_credential(
                session, current_user.id
            )
            if oauth is not None:
                await session.delete(oauth)
            if cfg is not None:
                await session.delete(cfg)
            await session.commit()
        return {"ok": True}

    @app.put(
        "/api/v1/integrations/pipedrive/config",
        response_model=PipedriveIntegrationStatus,
    )
    async def pipedrive_set_config(
        body: PipedriveConfigUpdate,
        current_user: User = Depends(get_current_user),
    ) -> PipedriveIntegrationStatus:
        async with session_factory() as session:
            oauth, cfg = await _pipedrive_credential(
                session, current_user.id
            )
            if oauth is None:
                raise HTTPException(
                    status_code=400, detail="pipedrive is not connected"
                )
            payload = {
                "default_pipeline_id": int(body.default_pipeline_id),
                "default_stage_id": int(body.default_stage_id),
            }
            if cfg is None:
                from leadgen.core.services.secrets_vault import encrypt

                cfg = UserIntegrationCredential(
                    user_id=current_user.id,
                    provider="pipedrive",
                    token_ciphertext=encrypt("pipedrive-config"),
                    config=payload,
                )
                session.add(cfg)
            else:
                cfg.config = payload
                cfg.updated_at = datetime.now(timezone.utc)
            await session.commit()
        return PipedriveIntegrationStatus(
            connected=True,
            api_domain=_pipedrive_api_domain(oauth.scope),
            account_email=oauth.account_email,
            scope=oauth.scope,
            expires_at=oauth.expires_at,
            default_pipeline_id=payload["default_pipeline_id"],
            default_stage_id=payload["default_stage_id"],
        )

    @app.get(
        "/api/v1/integrations/pipedrive/pipelines",
        response_model=PipedrivePipelinesResponse,
    )
    async def pipedrive_list_pipelines(
        current_user: User = Depends(get_current_user),
    ) -> PipedrivePipelinesResponse:
        if not _pipedrive_oauth_configured():
            raise _pipedrive_unavailable()
        from leadgen.core.services.oauth_store import (
            OAuthStoreError,
            ensure_fresh_token,
        )
        from leadgen.integrations.pipedrive import (
            PipedriveClient,
            PipedriveError,
        )

        async with session_factory() as session:
            oauth, _ = await _pipedrive_credential(
                session, current_user.id
            )
            if oauth is None:
                raise HTTPException(
                    status_code=400,
                    detail="pipedrive is not connected",
                )
            api_domain = _pipedrive_api_domain(oauth.scope)
            if not api_domain:
                raise HTTPException(
                    status_code=400,
                    detail="pipedrive api_domain unknown — reconnect",
                )
            try:
                fresh = await ensure_fresh_token(
                    session,
                    user_id=current_user.id,
                    provider="pipedrive",
                )
            except OAuthStoreError as exc:
                raise HTTPException(
                    status_code=400, detail=str(exc)
                ) from exc

        try:
            async with PipedriveClient(
                fresh.access_token, api_domain
            ) as client:
                pipelines = await client.list_pipelines()
        except PipedriveError as exc:
            raise HTTPException(
                status_code=502, detail=f"pipedrive: {exc}"
            ) from exc

        items = [
            PipedrivePipelineView(
                id=p.id,
                name=p.name,
                stages=[
                    PipedriveStageView(
                        id=s.id,
                        name=s.name,
                        pipeline_id=s.pipeline_id,
                        order_nr=s.order_nr,
                    )
                    for s in p.stages
                ],
            )
            for p in pipelines
        ]
        return PipedrivePipelinesResponse(items=items)

    @app.post(
        "/api/v1/leads/export-to-pipedrive",
        response_model=PipedriveExportResponse,
    )
    async def export_leads_to_pipedrive(
        body: PipedriveExportRequest,
        current_user: User = Depends(get_current_user),
    ) -> PipedriveExportResponse:
        """Push selected leads into Pipedrive as Person + Deal pairs."""
        if not _pipedrive_oauth_configured():
            raise _pipedrive_unavailable()
        from leadgen.core.services.oauth_store import (
            OAuthStoreError,
            ensure_fresh_token,
        )
        from leadgen.integrations.pipedrive import (
            PipedriveClient,
            PipedriveError,
            PipedrivePersonInput,
        )

        async with session_factory() as session:
            oauth, cfg = await _pipedrive_credential(
                session, current_user.id
            )
            if oauth is None:
                raise HTTPException(
                    status_code=400,
                    detail="pipedrive is not connected",
                )
            api_domain = _pipedrive_api_domain(oauth.scope)
            if not api_domain:
                raise HTTPException(
                    status_code=400,
                    detail="pipedrive api_domain unknown — reconnect",
                )
            config = (cfg.config if cfg is not None else {}) or {}
            pipeline_id = config.get("default_pipeline_id")
            stage_id = config.get("default_stage_id")
            if not (pipeline_id and stage_id):
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "pick a pipeline + stage in Settings → "
                        "Pipedrive before exporting leads"
                    ),
                )

            try:
                fresh = await ensure_fresh_token(
                    session,
                    user_id=current_user.id,
                    provider="pipedrive",
                )
            except OAuthStoreError as exc:
                raise HTTPException(
                    status_code=400, detail=str(exc)
                ) from exc

            lead_rows = (
                await session.execute(
                    select(Lead, SearchQuery)
                    .join(SearchQuery, SearchQuery.id == Lead.query_id)
                    .where(Lead.id.in_(list(body.lead_ids)))
                )
            ).all()
            authorised: dict[uuid.UUID, tuple[Lead, SearchQuery]] = {}
            for lead, search in lead_rows:
                if search.user_id == current_user.id:
                    authorised[lead.id] = (lead, search)
                    continue
                if search.team_id is not None and (
                    await _membership(
                        session, search.team_id, current_user.id
                    )
                ):
                    authorised[lead.id] = (lead, search)

        items: list[PipedriveExportItem] = []
        async with PipedriveClient(fresh.access_token, api_domain) as client:
            for lead_id in body.lead_ids:
                pair = authorised.get(lead_id)
                if pair is None:
                    items.append(
                        PipedriveExportItem(
                            lead_id=lead_id, error="not authorised"
                        )
                    )
                    continue
                lead, search = pair
                person_name = lead.name or "(unnamed)"
                email = _extract_lead_email(lead)
                person = PipedrivePersonInput(
                    name=person_name,
                    email=email,
                    phone=lead.phone,
                    org_name=lead.name,
                )
                try:
                    person_id = await client.upsert_person(person)
                    deal_id = await client.create_deal(
                        person_id=person_id,
                        title=f"{search.niche} · {person_name}"[:255],
                        pipeline_id=int(pipeline_id),
                        stage_id=int(stage_id),
                    )
                    items.append(
                        PipedriveExportItem(
                            lead_id=lead_id,
                            person_id=person_id,
                            deal_id=deal_id,
                        )
                    )
                except PipedriveError as exc:
                    logger.exception(
                        "pipedrive export: failed for lead %s", lead_id
                    )
                    items.append(
                        PipedriveExportItem(
                            lead_id=lead_id, error=str(exc)[:200]
                        )
                    )

        successes = sum(1 for it in items if it.deal_id)
        if successes:
            now = datetime.now(timezone.utc)
            async with session_factory() as session:
                for it in items:
                    if not it.deal_id:
                        continue
                    session.add(
                        LeadActivity(
                            lead_id=it.lead_id,
                            user_id=current_user.id,
                            kind="exported_pipedrive",
                            payload={
                                "person_id": it.person_id,
                                "deal_id": it.deal_id,
                            },
                            created_at=now,
                        )
                    )
                await session.commit()

        return PipedriveExportResponse(
            items=items,
            success_count=successes,
            failure_count=len(items) - successes,
        )

    # /api/v1/billing/* moved to routes/billing.py

    # ── /api/v1/niches (public taxonomy autocomplete) ──────────────────

    @app.get("/api/v1/niches", response_model=NicheTaxonomyResponse)
    async def list_niches(
        q: str | None = Query(default=None, max_length=80),
        lang: str | None = Query(default=None, max_length=8),
        limit: int = Query(default=12, ge=1, le=50),
    ) -> NicheTaxonomyResponse:
        """Static taxonomy lookup for the search-form niche combobox.

        Empty ``q`` returns the curated top-of-list so the dropdown
        can prefill on focus. Anything else does a substring/prefix
        match across labels + aliases in any supported language —
        a Russian-speaking user typing "дантист" lands on
        ``dentists`` even though the canonical English label says
        "Dentists".
        """
        from leadgen.data.niches import (
            DEFAULT_LANGUAGE,
            SUPPORTED_LANGUAGES,
        )
        from leadgen.data.niches import (
            suggest as suggest_niches_taxonomy,
        )

        language = lang or DEFAULT_LANGUAGE
        if language not in SUPPORTED_LANGUAGES:
            language = DEFAULT_LANGUAGE
        matches = suggest_niches_taxonomy(q, language=language, limit=limit)
        return NicheTaxonomyResponse(
            items=[
                NicheTaxonomyEntry(
                    id=m.id,
                    label=m.label(language),
                    category=m.category,
                )
                for m in matches
            ],
            query=(q or ""),
            language=language,
        )

    # ── /api/v1/cities (public city autocomplete) ──────────────────────

    @app.get("/api/v1/cities", response_model=CityListResponse)
    async def list_cities(
        q: str | None = Query(default=None, max_length=80),
        country: str | None = Query(default=None, max_length=4),
        lang: str | None = Query(default=None, max_length=8),
        limit: int = Query(default=12, ge=1, le=50),
    ) -> CityListResponse:
        """Curated city catalogue for the search-form region combobox.

        Empty ``q`` returns the population-sorted top so the dropdown
        prefills on focus. ``country`` (ISO2) narrows to a single
        country once the SPA knows the user picked scope=country.
        Anything outside the catalogue is still typeable by hand —
        the combobox is purely additive.
        """
        from leadgen.data.cities import suggest as suggest_cities

        language = (lang or "en").lower()
        matches = suggest_cities(
            q, country=country, language=language, limit=limit
        )
        return CityListResponse(
            items=[
                CityEntryResponse(
                    id=c.id,
                    name=c.label(language),
                    country=c.country,
                    lat=c.lat,
                    lon=c.lon,
                    population=c.population,
                )
                for c in matches
            ],
            query=(q or ""),
            language=language,
        )

    # /api/v1/admin/* (overview / sources/health / quality) moved
    # to routes/admin.py.

    # ── /api/v1/teams/{team_id}/analytics (per-team analytics) ────────

    @app.get(
        "/api/v1/teams/{team_id}/analytics",
        response_model=TeamAnalytics,
    )
    async def team_analytics(
        team_id: uuid.UUID,
        from_: datetime | None = Query(default=None, alias="from"),
        to: datetime | None = None,
        current_user: User = Depends(get_current_user),
    ) -> TeamAnalytics:
        """Owner-only per-team analytics for ``/app/team/analytics``.

        Returns a single payload with everything the page renders:
        totals, status breakdown, top source / member / niche, and a
        per-day timeseries of searches + leads. Defaults to the last
        30 days when the caller doesn't pass an explicit window.
        """
        now = datetime.now(timezone.utc)
        period_to = to or now
        period_from = from_ or (period_to - timedelta(days=30))
        if period_to.tzinfo is None:
            period_to = period_to.replace(tzinfo=timezone.utc)
        if period_from.tzinfo is None:
            period_from = period_from.replace(tzinfo=timezone.utc)

        async with session_factory() as session:
            team = await session.get(Team, team_id)
            if team is None:
                raise HTTPException(status_code=404, detail="team not found")
            membership = await _membership(session, team_id, current_user.id)
            if membership is None or not _can_manage_members(membership.role):
                raise HTTPException(
                    status_code=403,
                    detail="only owner or admin can view analytics",
                )

            base_searches = (
                select(SearchQuery)
                .where(SearchQuery.team_id == team_id)
                .where(SearchQuery.created_at >= period_from)
                .where(SearchQuery.created_at <= period_to)
            )
            search_rows = (
                await session.execute(base_searches)
            ).scalars().all()

            search_ids = [s.id for s in search_rows]
            lead_rows: list[Lead] = []
            if search_ids:
                lead_rows = (
                    await session.execute(
                        select(Lead).where(Lead.query_id.in_(search_ids))
                    )
                ).scalars().all()

            # Aggregations -------------------------------------------------
            scores = [
                float(lead.score_ai)
                for lead in lead_rows
                if lead.score_ai is not None
            ]
            avg_score = round(sum(scores) / len(scores), 1) if scores else None

            # Per-status (use whatever string is on Lead.lead_status to
            # stay compatible with custom team statuses from PR #30).
            status_counts: dict[str, int] = {}
            for lead in lead_rows:
                key = lead.lead_status or "new"
                status_counts[key] = status_counts.get(key, 0) + 1
            status_breakdown = [
                TeamAnalyticsStatusBucket(status=k, leads_count=v)
                for k, v in sorted(
                    status_counts.items(), key=lambda kv: kv[1], reverse=True
                )
            ]

            # Per-source.
            source_counts: dict[str, int] = {}
            for lead in lead_rows:
                key = lead.source or "unknown"
                source_counts[key] = source_counts.get(key, 0) + 1
            sources = [
                TeamAnalyticsSourceBucket(source=k, leads_count=v)
                for k, v in sorted(
                    source_counts.items(), key=lambda kv: kv[1], reverse=True
                )
            ]
            top_source = sources[0] if sources else None

            # Per-niche (search-level).
            niche_counts: dict[str, int] = {}
            for sq in search_rows:
                key = (sq.niche or "").strip().lower() or "—"
                niche_counts[key] = niche_counts.get(key, 0) + 1
            niches = [
                TeamAnalyticsNicheBucket(niche=k, searches_total=v)
                for k, v in sorted(
                    niche_counts.items(), key=lambda kv: kv[1], reverse=True
                )
            ]
            top_niche = niches[0] if niches else None

            # Per-member (active users in this team during the window).
            members_rows = (
                await session.execute(
                    select(TeamMembership, User)
                    .join(User, User.id == TeamMembership.user_id)
                    .where(TeamMembership.team_id == team_id)
                )
            ).all()
            search_user_map = {sq.id: sq.user_id for sq in search_rows}
            searches_by_user: dict[int, int] = {}
            for sq in search_rows:
                searches_by_user[sq.user_id] = searches_by_user.get(sq.user_id, 0) + 1
            leads_by_user: dict[int, list[Lead]] = {}
            for lead in lead_rows:
                uid = search_user_map.get(lead.query_id)
                if uid is None:
                    continue
                leads_by_user.setdefault(uid, []).append(lead)

            members: list[TeamAnalyticsMemberBucket] = []
            for _ms, u in members_rows:
                user_leads = leads_by_user.get(u.id, [])
                user_scores = [
                    float(lead.score_ai)
                    for lead in user_leads
                    if lead.score_ai is not None
                ]
                hot = sum(1 for s in user_scores if s >= 75)
                avg = (
                    round(sum(user_scores) / len(user_scores), 1)
                    if user_scores
                    else None
                )
                display = (
                    u.display_name
                    or " ".join(filter(None, [u.first_name, u.last_name]))
                    or f"User {u.id}"
                )
                members.append(
                    TeamAnalyticsMemberBucket(
                        user_id=u.id,
                        name=display,
                        searches_total=searches_by_user.get(u.id, 0),
                        leads_total=len(user_leads),
                        hot_leads=hot,
                        avg_score=avg,
                    )
                )
            members.sort(key=lambda m: m.leads_total, reverse=True)
            top_member = (
                members[0] if members and members[0].leads_total > 0 else None
            )

            # Day timeseries (UTC dates).
            day_searches: dict[str, int] = {}
            day_leads: dict[str, int] = {}
            for sq in search_rows:
                d = sq.created_at
                if d.tzinfo is None:
                    d = d.replace(tzinfo=timezone.utc)
                key = d.date().isoformat()
                day_searches[key] = day_searches.get(key, 0) + 1
            for lead in lead_rows:
                d = lead.created_at
                if d.tzinfo is None:
                    d = d.replace(tzinfo=timezone.utc)
                key = d.date().isoformat()
                day_leads[key] = day_leads.get(key, 0) + 1
            day_keys = sorted(set(day_searches) | set(day_leads))
            timeseries = [
                TeamAnalyticsTimepoint(
                    date=k,
                    searches_total=day_searches.get(k, 0),
                    leads_total=day_leads.get(k, 0),
                )
                for k in day_keys
            ]

            # Lead-cost approximation: same Haiku per-call estimate as
            # the admin dashboard, applied to every analyzed lead in
            # the window. One enriched lead ≈ one analysis call.
            enriched_leads = sum(1 for lead in lead_rows if lead.enriched)
            avg_lead_cost = (
                round((enriched_leads * 0.005) / max(len(lead_rows), 1), 4)
                if lead_rows
                else None
            )

        return TeamAnalytics(
            team_id=str(team_id),
            period_from=period_from,
            period_to=period_to,
            searches_total=len(search_rows),
            leads_total=len(lead_rows),
            avg_lead_score=avg_score,
            avg_lead_cost_usd=avg_lead_cost,
            status_breakdown=status_breakdown,
            top_source=top_source,
            top_member=top_member,
            top_niche=top_niche,
            members=members,
            sources=sources,
            niches=niches[:10],
            timeseries=timeseries,
        )

    # ── /api/v1/affiliate (per-user partner dashboard) ─────────────────

    @app.get("/api/v1/affiliate", response_model=AffiliateOverview)
    async def get_affiliate_overview(
        current_user: User = Depends(get_current_user),
    ) -> AffiliateOverview:
        async with session_factory() as session:
            codes = list(
                (
                    await session.execute(
                        select(AffiliateCode)
                        .where(AffiliateCode.owner_user_id == current_user.id)
                        .order_by(AffiliateCode.created_at.asc())
                    )
                )
                .scalars()
                .all()
            )
            counts: dict[str, tuple[int, int]] = {c.code: (0, 0) for c in codes}
            if codes:
                rows = (
                    await session.execute(
                        select(
                            Referral.code,
                            func.count(Referral.id),
                            func.count(Referral.first_paid_at),
                        )
                        .where(
                            Referral.code.in_([c.code for c in codes])
                        )
                        .group_by(Referral.code)
                    )
                ).all()
                for code, total, paid in rows:
                    counts[code] = (int(total or 0), int(paid or 0))

        items = [
            AffiliateCodeSchema(
                code=c.code,
                name=c.name,
                percent_share=c.percent_share,
                active=c.active,
                created_at=c.created_at,
                referrals_count=counts.get(c.code, (0, 0))[0],
                paid_referrals_count=counts.get(c.code, (0, 0))[1],
            )
            for c in codes
        ]
        return AffiliateOverview(
            codes=items,
            total_referrals=sum(i.referrals_count for i in items),
            total_paid_referrals=sum(i.paid_referrals_count for i in items),
        )

    @app.post(
        "/api/v1/affiliate/codes", response_model=AffiliateCodeSchema
    )
    async def create_affiliate_code(
        body: AffiliateCodeCreateRequest,
        current_user: User = Depends(get_current_user),
    ) -> AffiliateCodeSchema:
        """Create or claim an affiliate slug.

        Empty ``code`` → generate ~8-char URL-safe random slug. Caller-
        chosen slugs are normalised lowercase + restricted to
        ``[a-z0-9_-]`` so the public ``/r/{code}`` URL stays clean.
        """
        raw = (body.code or "").strip().lower()
        if raw:
            cleaned = "".join(
                ch for ch in raw if ch.isalnum() or ch in "-_"
            )
            if len(cleaned) < 3:
                raise HTTPException(
                    status_code=400,
                    detail="code must be at least 3 alphanumeric chars",
                )
            slug = cleaned[:64]
        else:
            slug = secrets.token_urlsafe(6).lower().replace("_", "").replace("-", "")[:8]
            if len(slug) < 3:
                slug = secrets.token_hex(4)
        async with session_factory() as session:
            existing = await session.get(AffiliateCode, slug)
            if existing is not None:
                raise HTTPException(
                    status_code=409, detail="this code is already taken"
                )
            row = AffiliateCode(
                code=slug,
                owner_user_id=current_user.id,
                name=(body.name or "").strip() or None,
            )
            session.add(row)
            await session.commit()
            await session.refresh(row)
        return AffiliateCodeSchema(
            code=row.code,
            name=row.name,
            percent_share=row.percent_share,
            active=row.active,
            created_at=row.created_at,
        )

    @app.patch(
        "/api/v1/affiliate/codes/{code}",
        response_model=AffiliateCodeSchema,
    )
    async def update_affiliate_code(
        code: str,
        body: AffiliateCodeUpdate,
        current_user: User = Depends(get_current_user),
    ) -> AffiliateCodeSchema:
        async with session_factory() as session:
            row = await session.get(AffiliateCode, code.lower())
            if row is None or row.owner_user_id != current_user.id:
                raise HTTPException(status_code=404, detail="code not found")
            if body.name is not None:
                row.name = body.name.strip() or None
            if body.active is not None:
                row.active = bool(body.active)
            await session.commit()
            await session.refresh(row)
        return AffiliateCodeSchema(
            code=row.code,
            name=row.name,
            percent_share=row.percent_share,
            active=row.active,
            created_at=row.created_at,
        )

    @app.delete("/api/v1/affiliate/codes/{code}")
    async def delete_affiliate_code(
        code: str,
        current_user: User = Depends(get_current_user),
    ) -> dict[str, bool]:
        async with session_factory() as session:
            row = await session.get(AffiliateCode, code.lower())
            if row is None or row.owner_user_id != current_user.id:
                raise HTTPException(status_code=404, detail="code not found")
            await session.delete(row)
            await session.commit()
        return {"ok": True}

    # ── /api/v1/teams/{team_id}/statuses (custom CRM pipeline) ─────────

    @app.get(
        "/api/v1/teams/{team_id}/statuses",
        response_model=LeadStatusListResponse,
    )
    async def list_lead_statuses(
        team_id: uuid.UUID,
        current_user: User = Depends(get_current_user),
    ) -> LeadStatusListResponse:
        async with session_factory() as session:
            membership = await _membership(session, team_id, current_user.id)
            if membership is None:
                raise HTTPException(status_code=403, detail="forbidden")
            rows = (
                (
                    await session.execute(
                        select(LeadStatus)
                        .where(LeadStatus.team_id == team_id)
                        .order_by(
                            LeadStatus.order_index.asc(),
                            LeadStatus.created_at.asc(),
                        )
                    )
                )
                .scalars()
                .all()
            )
        return LeadStatusListResponse(
            items=[_status_to_schema(s) for s in rows]
        )

    @app.post(
        "/api/v1/teams/{team_id}/statuses",
        response_model=LeadStatusSchema,
    )
    async def create_lead_status(
        team_id: uuid.UUID,
        body: LeadStatusCreate,
        current_user: User = Depends(get_current_user),
    ) -> LeadStatusSchema:
        async with session_factory() as session:
            membership = await _membership(session, team_id, current_user.id)
            if membership is None:
                raise HTTPException(status_code=403, detail="forbidden")
            key = body.key.strip().lower()
            cleaned = "".join(
                ch for ch in key if ch.isalnum() or ch in "-_"
            )
            if len(cleaned) < 1:
                raise HTTPException(
                    status_code=400, detail="key must be alphanumeric"
                )
            existing_keys = (
                await session.execute(
                    select(LeadStatus.key, func.max(LeadStatus.order_index))
                    .where(LeadStatus.team_id == team_id)
                    .group_by(LeadStatus.key)
                )
            ).all()
            keys = {k for k, _ in existing_keys}
            if cleaned in keys:
                raise HTTPException(
                    status_code=409, detail="key already exists in this team"
                )
            max_order = max(
                (m for _, m in existing_keys), default=-1
            )
            row = LeadStatus(
                team_id=team_id,
                key=cleaned[:32],
                label=body.label.strip()[:64],
                color=(body.color or "slate").strip().lower()[:16],
                order_index=int(max_order) + 1,
                is_terminal=bool(body.is_terminal),
            )
            session.add(row)
            await session.commit()
            await session.refresh(row)
        return _status_to_schema(row)

    @app.patch(
        "/api/v1/teams/{team_id}/statuses/{status_id}",
        response_model=LeadStatusSchema,
    )
    async def update_lead_status(
        team_id: uuid.UUID,
        status_id: uuid.UUID,
        body: LeadStatusUpdate,
        current_user: User = Depends(get_current_user),
    ) -> LeadStatusSchema:
        async with session_factory() as session:
            membership = await _membership(session, team_id, current_user.id)
            if membership is None:
                raise HTTPException(status_code=403, detail="forbidden")
            row = await session.get(LeadStatus, status_id)
            if row is None or row.team_id != team_id:
                raise HTTPException(status_code=404, detail="status not found")
            if body.label is not None:
                row.label = body.label.strip()[:64] or row.label
            if body.color is not None:
                row.color = body.color.strip().lower()[:16] or "slate"
            if body.order_index is not None:
                row.order_index = int(body.order_index)
            if body.is_terminal is not None:
                row.is_terminal = bool(body.is_terminal)
            await session.commit()
            await session.refresh(row)
        return _status_to_schema(row)

    @app.delete("/api/v1/teams/{team_id}/statuses/{status_id}")
    async def delete_lead_status(
        team_id: uuid.UUID,
        status_id: uuid.UUID,
        current_user: User = Depends(get_current_user),
    ) -> dict[str, bool]:
        async with session_factory() as session:
            membership = await _membership(session, team_id, current_user.id)
            if membership is None:
                raise HTTPException(status_code=403, detail="forbidden")
            row = await session.get(LeadStatus, status_id)
            if row is None or row.team_id != team_id:
                raise HTTPException(status_code=404, detail="status not found")
            # Refuse to delete a status still attached to live leads —
            # cascading them silently to "archived" surprises the user.
            in_use = (
                await session.execute(
                    select(func.count(Lead.id))
                    .join(SearchQuery, SearchQuery.id == Lead.query_id)
                    .where(SearchQuery.team_id == team_id)
                    .where(Lead.lead_status == row.key)
                    .where(Lead.deleted_at.is_(None))
                )
            ).scalar_one()
            if int(in_use or 0) > 0:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"{in_use} lead(s) still use this status. "
                        "Move them to another status first, then delete."
                    ),
                )
            await session.delete(row)
            await session.commit()
        return {"ok": True}

    @app.post(
        "/api/v1/teams/{team_id}/statuses/reorder",
        response_model=LeadStatusListResponse,
    )
    async def reorder_lead_statuses(
        team_id: uuid.UUID,
        body: LeadStatusReorderRequest,
        current_user: User = Depends(get_current_user),
    ) -> LeadStatusListResponse:
        async with session_factory() as session:
            membership = await _membership(session, team_id, current_user.id)
            if membership is None:
                raise HTTPException(status_code=403, detail="forbidden")
            owned_ids = set(
                (
                    await session.execute(
                        select(LeadStatus.id).where(
                            LeadStatus.team_id == team_id
                        )
                    )
                )
                .scalars()
                .all()
            )
            updates = [
                {"id": sid, "order_index": index}
                for index, sid in enumerate(body.ordered_ids)
                if sid in owned_ids
            ]
            if updates:
                # Single round-trip executemany — replaces the previous
                # for-loop that issued one UPDATE per status row.
                await session.execute(sa.update(LeadStatus), updates)
            await session.commit()
            rows = list(
                (
                    await session.execute(
                        select(LeadStatus)
                        .where(LeadStatus.team_id == team_id)
                        .order_by(
                            LeadStatus.order_index.asc(),
                            LeadStatus.created_at.asc(),
                        )
                    )
                )
                .scalars()
                .all()
            )
        return LeadStatusListResponse(
            items=[_status_to_schema(r) for r in rows]
        )

    # ── SSE: live search progress ───────────────────────────────────────

    @app.get("/api/v1/searches/{search_id}/progress")
    async def search_progress(
        search_id: uuid.UUID,
        api_key: str | None = Query(default=None, alias="api_key"),
    ) -> StreamingResponse:
        """Server-Sent Events stream of progress beats.

        Auth: if WEB_API_KEY is configured, require it as ``?api_key=``.
        Otherwise (open-demo mode), stream unauthenticated — connections
        are short-lived and the broker auto-closes on search completion.
        """
        expected = get_settings().web_api_key
        if expected and api_key != expected:
            raise HTTPException(status_code=401, detail="invalid api_key")

        # SSE streams hold a connection (and a subscriber slot in the
        # broker) for the lifetime of the search. Cap each stream at
        # 10 min and send a comment heartbeat every 15s so idle proxies
        # and load balancers don't silently drop the socket mid-search.
        stream_max_seconds = 600.0
        heartbeat_interval = 15.0

        async def event_stream() -> asyncio.AsyncIterator[bytes]:
            yield b"retry: 5000\n\n"
            sub = default_broker.subscribe(search_id)
            loop = asyncio.get_running_loop()
            deadline = loop.time() + stream_max_seconds
            try:
                while True:
                    remaining = deadline - loop.time()
                    if remaining <= 0:
                        yield b"event: timeout\ndata: {}\n\n"
                        return
                    try:
                        event = await asyncio.wait_for(
                            sub.__anext__(),
                            timeout=min(heartbeat_interval, remaining),
                        )
                    except TimeoutError:
                        # SSE comment line — keepalive that the client
                        # does not surface as an event.
                        yield b": heartbeat\n\n"
                        continue
                    except StopAsyncIteration:
                        break
                    payload = json.dumps({"kind": event.kind, **event.data})
                    yield f"event: {event.kind}\ndata: {payload}\n\n".encode()
                yield b"event: done\ndata: {}\n\n"
            finally:
                await sub.aclose()

        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache, no-transform",
                "X-Accel-Buffering": "no",
                "Connection": "keep-alive",
            },
        )

    # ── Legacy /users/{user_id}/* path redirects ───────────────────────
    #
    # Cookie sessions know who the caller is, so the user_id path
    # parameter is now redundant. New callers MUST use /users/me/*;
    # old clients (older SPA bundles, third-party integrations) get a
    # 308 redirect to the canonical path. The 403 on user_id mismatch
    # also closes the historical IDOR where a path id was trusted
    # without comparing it to the session user.
    legacy_user_suffixes: list[tuple[str, list[str]]] = [
        ("", ["GET", "PATCH", "DELETE"]),
        ("/change-email", ["POST"]),
        ("/change-password", ["POST"]),
        ("/audit-log", ["GET"]),
        ("/export", ["GET"]),
        ("/assistant-memory", ["GET", "DELETE"]),
        ("/weekly-checkin", ["GET"]),
        ("/suggest-niches", ["POST"]),
        ("/tasks", ["GET"]),
    ]

    def _make_legacy_redirect(suffix: str):
        async def _redirect(
            user_id: int,
            request: Request,
            current_user: User = Depends(get_current_user),
        ) -> Response:
            if user_id != current_user.id:
                raise HTTPException(status_code=403, detail="forbidden")
            qs = ("?" + request.url.query) if request.url.query else ""
            return RedirectResponse(
                url=f"/api/v1/users/me{suffix}{qs}", status_code=308
            )

        return _redirect

    # Per-domain APIRouter modules carved out of this monolith. Each
    # one is self-contained (uses module-level dependencies, doesn't
    # capture create_app() locals). Adding a new domain = drop a file
    # in routes/ and one include_router line here.
    from leadgen.adapters.web_api.routes import admin as _admin
    from leadgen.adapters.web_api.routes import assistant as _assistant
    from leadgen.adapters.web_api.routes import audit as _audit
    from leadgen.adapters.web_api.routes import auth as _auth
    from leadgen.adapters.web_api.routes import billing as _billing
    from leadgen.adapters.web_api.routes import leads as _leads
    from leadgen.adapters.web_api.routes import (
        notifications as _notifications,
    )
    from leadgen.adapters.web_api.routes import search as _search
    from leadgen.adapters.web_api.routes import segments as _segments
    from leadgen.adapters.web_api.routes import sequences as _sequences
    from leadgen.adapters.web_api.routes import tags as _tags
    from leadgen.adapters.web_api.routes import teams as _teams
    from leadgen.adapters.web_api.routes import templates as _templates
    from leadgen.adapters.web_api.routes import users as _users
    from leadgen.adapters.web_api.routes import webhooks as _webhooks

    # IMPORTANT: include the routers FIRST so the literal /users/me
    # routes win over the legacy /users/{user_id} catch-all below —
    # otherwise FastAPI tries to parse "me" as an int and the SPA's
    # GET /api/v1/users/me dies with 422 ("Input should be a valid
    # integer, unable to parse string as an integer").
    app.include_router(_admin.router)
    app.include_router(_audit.router)
    app.include_router(_assistant.router)
    app.include_router(_auth.router)
    app.include_router(_billing.router)
    app.include_router(_leads.router)
    app.include_router(_notifications.router)
    app.include_router(_search.router)
    app.include_router(_segments.router)
    app.include_router(_sequences.router)
    app.include_router(_tags.router)
    app.include_router(_teams.router)
    app.include_router(_templates.router)
    app.include_router(_users.router)
    app.include_router(_webhooks.router)

    for suffix, methods in legacy_user_suffixes:
        app.add_api_route(
            f"/api/v1/users/{{user_id}}{suffix}",
            _make_legacy_redirect(suffix),
            methods=methods,
            deprecated=True,
            include_in_schema=False,
        )

    return app


async def _run_web_search_inline(
    query_id: uuid.UUID, user_profile: dict[str, Any] | None
) -> None:
    """Fallback in-process runner when no Redis worker is available.

    Wraps ``run_search_with_sinks`` with a WebDeliverySink + a
    BrokerProgressSink so the SSE endpoint has something to stream.
    Any exception is swallowed here — the pipeline itself marks the
    SearchQuery as failed, and a crash in this task shouldn't take
    down the API server.
    """
    try:
        progress = BrokerProgressSink(default_broker, query_id)
        delivery = WebDeliverySink(query_id)
        await run_search_with_timeout(
            query_id=query_id,
            progress=progress,
            delivery=delivery,
            user_profile=user_profile,
        )
    except Exception:  # noqa: BLE001
        logger.exception("inline web search crashed for %s", query_id)


# How often the in-process scheduler scans the saved_searches table
# when Redis isn't available. 60s lines up with the `daily` recurrence
# resolution, which is the finest grain we expose in the UI.
_SCHEDULER_TICK_SEC = 60


async def _saved_search_scheduler_loop() -> None:
    """Background task: fire due saved searches every ~60 seconds.

    Intentionally does *no* error propagation — the task is
    long-lived and any per-row failure is logged inside
    ``dispatch_due``. A persistent crash here would silence further
    scheduling, so we wrap the whole loop and continue.
    """
    from leadgen.core.services.saved_searches import (
        build_search_query,
        dispatch_due,
    )

    async def _run_one(saved: SavedSearch, session) -> uuid.UUID | None:
        new_query = build_search_query(saved)
        session.add(new_query)
        await session.commit()
        # Mirror the per-search profile lookup used by POST /searches —
        # cheap and lets Henry's tone match the owner.
        user = await session.get(User, saved.user_id)
        profile = (
            {
                "display_name": user.display_name or user.first_name,
                "language_code": user.language_code,
            }
            if user is not None
            else None
        )
        queued = await enqueue_search(
            new_query.id, chat_id=None, user_profile=profile
        )
        if not queued:
            spawn(
                _run_web_search_inline(new_query.id, profile),
                name=f"convioo-saved-{new_query.id}",
            )
        return new_query.id

    while True:
        try:
            async with session_factory() as session:
                count = await dispatch_due(session, run_search=_run_one)
                if count:
                    logger.info(
                        "saved-search scheduler: dispatched %d searches",
                        count,
                    )
        except Exception:  # noqa: BLE001
            logger.exception("saved-search scheduler tick crashed")
        await asyncio.sleep(_SCHEDULER_TICK_SEC)


def _to_summary(query: SearchQuery) -> SearchSummary:
    insights: str | None = None
    if isinstance(query.analysis_summary, dict):
        raw = query.analysis_summary.get("insights")
        if isinstance(raw, str):
            insights = raw
    return SearchSummary(
        id=query.id,
        user_id=query.user_id,
        niche=query.niche,
        region=query.region,
        status=query.status,
        source=query.source,
        created_at=query.created_at,
        finished_at=query.finished_at,
        leads_count=query.leads_count,
        avg_score=query.avg_score,
        hot_leads_count=query.hot_leads_count,
        error=query.error,
        insights=insights,
    )


async def _marks_for_user(
    session, user_id: int, lead_ids: list[uuid.UUID]
) -> dict[uuid.UUID, str]:
    """Return ``lead_id -> color`` for one user across many leads."""
    if not lead_ids:
        return {}
    rows = (
        await session.execute(
            select(LeadMark.lead_id, LeadMark.color)
            .where(LeadMark.user_id == user_id)
            .where(LeadMark.lead_id.in_(lead_ids))
        )
    ).all()
    return {lead_id: color for lead_id, color in rows}


def _to_lead_response(
    lead: Lead,
    mark_color: str | None,
    user_tags: list[LeadTagSchema] | None = None,
) -> LeadResponse:
    payload = LeadResponse.model_validate(lead)
    payload.mark_color = mark_color
    if user_tags:
        payload.user_tags = list(user_tags)
    return payload


async def _tags_by_lead(
    session, lead_ids: list[uuid.UUID]
) -> dict[uuid.UUID, list[LeadTagSchema]]:
    """Eager-load every tag chip attached to ``lead_ids``.

    Returns ``{}`` for an empty input so callers can blindly merge it
    into the listing result.
    """
    if not lead_ids:
        return {}
    rows = (
        await session.execute(
            select(LeadTagAssignment.lead_id, LeadTag)
            .join(LeadTag, LeadTag.id == LeadTagAssignment.tag_id)
            .where(LeadTagAssignment.lead_id.in_(lead_ids))
            .order_by(LeadTag.created_at.asc())
        )
    ).all()
    out: dict[uuid.UUID, list[LeadTagSchema]] = {}
    for lead_id, tag in rows:
        out.setdefault(lead_id, []).append(
            LeadTagSchema(
                id=tag.id, name=tag.name, color=tag.color, team_id=tag.team_id
            )
        )
    return out


def _temp(score: float | None) -> str:
    """Bucket a 0–100 AI score into prototype temperature tiers."""
    if score is None:
        return "cold"
    if score >= 75:
        return "hot"
    if score >= 50:
        return "warm"
    return "cold"


def _extract_lead_email(lead: Lead) -> str | None:
    """Pluck the first usable email out of the website-meta payload.

    The website scraper stores discovered addresses under
    ``website_meta["emails"]`` after filtering out generic ones
    (info@, support@…). We pick the first; if none, we look at
    ``website_meta["primary_email"]`` for the rare case the scraper
    surfaced one but stripped the array.
    """
    meta = lead.website_meta or {}
    candidates = meta.get("emails") or []
    if isinstance(candidates, list) and candidates:
        first = candidates[0]
        if isinstance(first, str) and "@" in first:
            return first
    primary = meta.get("primary_email")
    if isinstance(primary, str) and "@" in primary:
        return primary
    return None


_password_hasher = PasswordHasher()


def _hash_password(plain: str) -> str:
    return _password_hasher.hash(plain)


def _verify_password(plain: str, hashed: str) -> bool:
    try:
        return _password_hasher.verify(hashed, plain)
    except VerifyMismatchError:
        return False
    except Exception:  # noqa: BLE001
        return False


async def _record_audit(
    session,
    *,
    user_id: int,
    action: str,
    request: Request | None = None,
    payload: dict[str, Any] | None = None,
) -> None:
    """Append an entry to ``user_audit_logs``.

    Best-effort: callers must commit the session themselves. Failures
    are logged but never raised so an audit hiccup can't break a real
    user-facing operation.
    """
    try:
        ua = (
            request.headers.get("user-agent")[:256]
            if request is not None and request.headers.get("user-agent")
            else None
        )
        session.add(
            UserAuditLog(
                user_id=user_id,
                action=action[:64],
                ip=request_ip(request),
                user_agent=ua,
                payload=payload,
            )
        )
    except Exception:  # noqa: BLE001
        logger.exception("failed to record audit log entry")


async def _issue_and_send_verification(session, user: User) -> None:
    """Mint a fresh verification token and email the user.

    Invalidates earlier outstanding tokens so there's only one live
    link at a time. Email dispatch failures don't bubble — the
    log-only fallback in send_email keeps signups working without a
    real provider.
    """
    settings = get_settings()
    await session.execute(
        update(EmailVerificationToken)
        .where(EmailVerificationToken.user_id == user.id)
        .where(EmailVerificationToken.kind == "verify")
        .where(EmailVerificationToken.used_at.is_(None))
        .values(used_at=datetime.now(timezone.utc))
    )
    token = secrets.token_urlsafe(32)
    expires = datetime.now(timezone.utc) + timedelta(hours=24)
    session.add(
        EmailVerificationToken(
            user_id=user.id,
            kind="verify",
            token=token,
            expires_at=expires,
        )
    )
    await session.commit()

    base = settings.public_app_url.rstrip("/")
    verify_url = f"{base}/verify-email/{token}"
    name = (
        user.first_name
        or user.display_name
        or (user.email.split("@")[0] if user.email else "")
        or "там"
    )
    html, text = render_verification_email(name=name, verify_url=verify_url)
    if user.email:
        await send_email(
            to=user.email,
            subject="Подтвердите email — Convioo",
            html=html,
            text=text,
        )


async def _issue_and_send_change_email(
    session, user: User, new_email: str
) -> None:
    """Mint a change-email token addressed to the *new* mailbox.

    The existing email keeps working until the user clicks the link;
    only then ``users.email`` is rewritten to the pending value.
    Earlier outstanding change-email tokens are invalidated so the
    user can't end up confirming a stale request.
    """
    settings = get_settings()
    await session.execute(
        update(EmailVerificationToken)
        .where(EmailVerificationToken.user_id == user.id)
        .where(EmailVerificationToken.kind == "change_email")
        .where(EmailVerificationToken.used_at.is_(None))
        .values(used_at=datetime.now(timezone.utc))
    )
    token = secrets.token_urlsafe(32)
    expires = datetime.now(timezone.utc) + timedelta(hours=24)
    session.add(
        EmailVerificationToken(
            user_id=user.id,
            kind="change_email",
            token=token,
            pending_email=new_email,
            expires_at=expires,
        )
    )
    await session.commit()

    base = settings.public_app_url.rstrip("/")
    verify_url = f"{base}/verify-email/{token}"
    name = (
        user.first_name
        or user.display_name
        or new_email.split("@")[0]
        or "там"
    )
    html, text = render_verification_email(name=name, verify_url=verify_url)
    await send_email(
        to=new_email,
        subject="Подтвердите новый email — Convioo",
        html=html,
        text=text,
    )


def _is_onboarded(user: User) -> bool:
    """Web onboarding gate.

    The web flow only requires a confirmed identity (a name + the
    onboarded_at stamp set at registration). What the user sells, the
    niches they target and their home region are filled later from
    the workspace (manually on /app/profile or via Henry) — they no
    longer block access. The Telegram bot keeps its own stricter
    check because its conversational onboarding still owns those
    fields end-to-end before letting the user search.
    """
    return user.onboarded_at is not None and bool(
        user.first_name or user.display_name
    )


def _invite_expired(invite: TeamInvite) -> bool:
    expires = invite.expires_at
    if expires.tzinfo is None:
        expires = expires.replace(tzinfo=timezone.utc)
    return datetime.now(timezone.utc) >= expires


async def _load_invite(session, token: str) -> tuple[TeamInvite, Team]:
    result = await session.execute(
        select(TeamInvite, Team)
        .join(Team, Team.id == TeamInvite.team_id)
        .where(TeamInvite.token == token)
        .limit(1)
    )
    row = result.first()
    if row is None:
        raise HTTPException(status_code=404, detail="invite not found")
    return row[0], row[1]


async def _authorise_lead_access(
    session, lead_id: uuid.UUID, user_id: int
) -> tuple[Lead, SearchQuery | None]:
    """Load a lead and verify the caller may touch it.

    Allowed when the caller owns the parent search, or is a member of
    the team that search belongs to. Any failure — missing lead,
    missing search, foreign owner — answers 404 (never 403) so lead
    ids can't be probed for existence across accounts.
    """
    lead = await session.get(Lead, lead_id)
    if lead is None:
        raise HTTPException(status_code=404, detail="lead not found")
    search = await session.get(SearchQuery, lead.query_id)
    allowed = search is not None and search.user_id == user_id
    if not allowed and search is not None and search.team_id is not None:
        allowed = (
            await _membership(session, search.team_id, user_id)
        ) is not None
    if not allowed:
        raise HTTPException(status_code=404, detail="lead not found")
    return lead, search


async def _resolve_team_view(
    session,
    team_id: uuid.UUID,
    caller_user_id: int,
    member_user_id: int | None,
) -> int:
    """Decide whose data the caller is allowed to read in a team view.

    Members only ever see their own. The owner can pass an explicit
    ``member_user_id`` to drill into a teammate's CRM; everyone else
    gets a 403 if they try the same.
    """
    caller = await _membership(session, team_id, caller_user_id)
    if caller is None:
        raise HTTPException(status_code=403, detail="not a team member")

    if member_user_id is None or member_user_id == caller_user_id:
        return caller_user_id

    if not _can_manage_members(caller.role):
        raise HTTPException(
            status_code=403,
            detail="only owner or admin can view another member",
        )
    target = await _membership(session, team_id, member_user_id)
    if target is None:
        raise HTTPException(status_code=404, detail="that user isn't a team member")
    return member_user_id


async def _team_prior_searches(
    session,
    team_id: uuid.UUID,
    niche: str,
    region: str,
) -> list[PriorTeamSearch]:
    """Return earlier completed searches in this team that already
    used the same (niche, region) pair, normalised case-insensitively
    and trimmed. Empty list = combo is fresh, OK to launch.
    """
    n = (niche or "").strip().lower()
    r = (region or "").strip().lower()
    if not n or not r:
        return []

    rows = (
        await session.execute(
            select(SearchQuery, User)
            .join(User, User.id == SearchQuery.user_id)
            .where(SearchQuery.team_id == team_id)
            .where(func.lower(func.trim(SearchQuery.niche)) == n)
            .where(func.lower(func.trim(SearchQuery.region)) == r)
            .where(SearchQuery.status.in_(["running", "done", "pending"]))
            .order_by(SearchQuery.created_at.desc())
        )
    ).all()

    out: list[PriorTeamSearch] = []
    for sq, user in rows:
        display = (
            user.display_name
            or " ".join(filter(None, [user.first_name, user.last_name]))
            or f"User {user.id}"
        )
        out.append(
            PriorTeamSearch(
                search_id=sq.id,
                user_id=sq.user_id,
                user_name=display,
                niche=sq.niche,
                region=sq.region,
                leads_count=sq.leads_count,
                created_at=sq.created_at,
            )
        )
    return out


def _status_to_schema(row: LeadStatus) -> LeadStatusSchema:
    return LeadStatusSchema(
        id=row.id,
        key=row.key,
        label=row.label,
        color=row.color,
        order_index=row.order_index,
        is_terminal=row.is_terminal,
    )


# ``_membership`` and ``_can_manage_tag`` were lifted into
# ``routes/_helpers.py`` so the extracted route modules don't import
# back into this file (which would be a cycle). The module-level
# aliases keep the rest of ``app.py`` working unchanged.
from leadgen.adapters.web_api.routes._helpers import (  # noqa: E402
    membership as _membership,
)


async def _team_detail(
    session, team: Team, viewer_user_id: int
) -> TeamDetailResponse:
    membership = await _membership(session, team.id, viewer_user_id)
    if membership is None:
        raise HTTPException(status_code=403, detail="not a team member")

    rows = (
        await session.execute(
            select(TeamMembership, User)
            .join(User, User.id == TeamMembership.user_id)
            .where(TeamMembership.team_id == team.id)
            .order_by(TeamMembership.created_at)
        )
    ).all()

    members: list[TeamMemberResponse] = []
    for i, (m, user) in enumerate(rows):
        display = (
            user.display_name
            or " ".join(filter(None, [user.first_name, user.last_name]))
            or f"User {user.id}"
        )
        initials = "".join(
            part[:1].upper()
            for part in display.split()
            if part
        )[:2] or display[:1].upper()
        members.append(
            TeamMemberResponse(
                id=user.id,
                name=display,
                role=m.role,
                description=m.description,
                initials=initials,
                color=_DEMO_TEAM_COLORS[i % len(_DEMO_TEAM_COLORS)],
                email=None,
            )
        )

    return TeamDetailResponse(
        id=team.id,
        name=team.name,
        description=team.description,
        plan=team.plan,
        created_at=team.created_at,
        role=membership.role,
        members=members,
    )


def _to_profile(user: User) -> UserProfile:
    recovery = getattr(user, "recovery_email", None)
    return UserProfile(
        user_id=user.id,
        first_name=user.first_name or "",
        last_name=user.last_name or "",
        display_name=user.display_name,
        age_range=user.age_range,
        gender=user.gender,
        business_size=user.business_size,
        profession=user.profession,
        service_description=user.service_description,
        home_region=user.home_region,
        niches=list(user.niches) if user.niches else None,
        language_code=user.language_code,
        calendly_url=user.calendly_url,
        onboarded=_is_onboarded(user),
        onboarding_tour_completed=user.onboarding_completed_at is not None,
        email=user.email,
        email_verified=user.email_verified_at is not None,
        recovery_email_masked=mask_email(recovery) if recovery else None,
        queries_used=int(user.queries_used or 0),
        queries_limit=int(user.queries_limit or 0),
    )


async def _summarise_and_store(
    user_id: int,
    team_id: uuid.UUID | None,
    history: list[dict[str, str]],
    user_profile: dict[str, Any] | None,
    existing_memories: list[dict[str, Any]],
) -> None:
    """Background task: distill the dialogue, persist summary + facts.

    Run from ``asyncio.create_task`` so the user-facing chat reply
    isn't blocked on the second LLM call. Failures are swallowed —
    memory is best-effort, the chat itself is the source of truth
    for the current turn.
    """
    try:
        analyzer = AIAnalyzer()
        result = await analyzer.summarize_session(
            history, user_profile, existing_memories=existing_memories
        )
        summary = result.get("summary")
        facts = result.get("facts") or []
        if not summary and not facts:
            return
        async with session_factory() as session:
            if summary:
                await record_memory(
                    session,
                    user_id,
                    team_id,
                    kind="summary",
                    content=summary,
                    meta={"messages": len(history)},
                )
            for fact in facts:
                await record_memory(
                    session,
                    user_id,
                    team_id,
                    kind="fact",
                    content=fact,
                )
            await prune_old(session, user_id, team_id)
            await session.commit()
    except Exception:  # noqa: BLE001
        logger.exception(
            "summarise_and_store failed for user_id=%s team=%s",
            user_id,
            team_id,
        )


# ── Henry confirm-before-write plumbing ─────────────────────────────

# The confirm/refuse keyword regexes (ru + uk + en) live in
# ``routes/_helpers.py`` — single source of truth. The alias keeps the
# legacy ``app.py`` name importable.
from leadgen.adapters.web_api.routes._helpers import (  # noqa: E402, F401, I001
    detect_confirmation as _detect_confirmation,
)


_PROFILE_FIELDS_WHITELIST = {
    "display_name",
    "age_range",
    "business_size",
    "service_description",
    "home_region",
    "niches",
}


async def _apply_pending_actions(
    session,
    user: User | None,
    team_context: dict[str, Any] | None,
    actions: list[PendingAction],
) -> list[PendingAction]:
    """Apply a list of confirmed actions, return what was applied.

    Each action is validated against the kind's whitelist and the
    caller's permissions (owner-only for team / member descriptions).
    Failures are logged and the action is silently skipped — the
    rest of the batch still goes through.
    """
    is_owner = bool(team_context and team_context.get("is_owner"))
    raw_team_id = (team_context or {}).get("team_id")
    team_id: uuid.UUID | None
    if isinstance(raw_team_id, uuid.UUID):
        team_id = raw_team_id
    elif isinstance(raw_team_id, str):
        try:
            team_id = uuid.UUID(raw_team_id)
        except ValueError:
            team_id = None
    else:
        team_id = None

    applied: list[PendingAction] = []
    profile_dirty = False

    for action in actions:
        try:
            kind = action.kind
            payload = action.payload or {}

            if kind == "profile_patch" and user is not None:
                changed = False
                for key, val in payload.items():
                    if key not in _PROFILE_FIELDS_WHITELIST:
                        continue
                    if key == "niches":
                        if isinstance(val, list):
                            cleaned = [
                                n.strip()
                                for n in val
                                if isinstance(n, str) and n.strip()
                            ]
                            user.niches = cleaned[:7] or None
                            changed = True
                    elif key == "service_description":
                        raw = (val or "").strip()
                        if raw:
                            user.service_description = raw
                            try:
                                user.profession = (
                                    await asyncio.wait_for(
                                        AIAnalyzer().normalize_profession(raw),
                                        timeout=8.0,
                                    )
                                ) or raw
                            except Exception:  # noqa: BLE001
                                logger.exception(
                                    "normalize_profession failed in apply"
                                )
                                user.profession = raw
                        else:
                            user.service_description = None
                            user.profession = None
                        changed = True
                    else:
                        text_val = val if val is None else str(val).strip() or None
                        setattr(user, key, text_val)
                        changed = True
                if changed:
                    profile_dirty = True
                    applied.append(action)

            elif (
                kind == "team_description"
                and is_owner
                and team_id is not None
            ):
                description = (payload.get("description") or "").strip() or None
                team = await session.get(Team, team_id)
                if team is not None:
                    team.description = (
                        description[:2000] if description else None
                    )
                    applied.append(action)

            elif (
                kind == "member_description"
                and is_owner
                and team_id is not None
            ):
                target_user_id = payload.get("user_id")
                description = (payload.get("description") or "").strip() or None
                if isinstance(target_user_id, int):
                    membership = await _membership(
                        session, team_id, target_user_id
                    )
                    if membership is not None:
                        membership.description = (
                            description[:1000] if description else None
                        )
                        applied.append(action)

            elif kind == "launch_search" and user is not None:
                niche = (payload.get("niche") or "").strip()
                region = (payload.get("region") or "").strip()
                if not niche or not region:
                    continue
                ideal_customer = (
                    payload.get("ideal_customer") or ""
                ).strip() or None
                exclusions = (payload.get("exclusions") or "").strip() or None
                # Compose the profession blob the search pipeline reads
                # exactly the way /app/search builds it.
                offer_parts: list[str] = []
                base_offer = (user.profession or user.service_description or "").strip()
                if base_offer:
                    offer_parts.append(base_offer)
                if ideal_customer:
                    offer_parts.append(f"Идеальный клиент: {ideal_customer}")
                if exclusions:
                    offer_parts.append(f"Исключения: {exclusions}")
                profession_blob = ". ".join(offer_parts) or None

                new_query = SearchQuery(
                    user_id=user.id,
                    team_id=team_id,
                    niche=niche[:256],
                    region=region[:256],
                    source="web",
                )
                session.add(new_query)
                try:
                    await session.commit()
                except Exception:  # noqa: BLE001
                    await session.rollback()
                    logger.exception(
                        "launch_search via Henry: insert failed (likely "
                        "duplicate niche+region in team)"
                    )
                    continue
                await session.refresh(new_query)

                user_profile_for_run: dict[str, Any] = {
                    "display_name": user.display_name or user.first_name,
                    "age_range": user.age_range,
                    "gender": user.gender,
                    "business_size": user.business_size,
                    "profession": profession_blob or user.profession,
                    "service_description": user.service_description,
                    "home_region": user.home_region,
                    "niches": list(user.niches or []),
                    "language_code": user.language_code,
                }

                # Try the queue first; fall through to inline runner if
                # Redis isn't configured. Same path /api/v1/searches uses.
                queued_id = await enqueue_search(
                    new_query.id,
                    chat_id=None,
                    user_profile=user_profile_for_run,
                )
                if not queued_id:
                    spawn(
                        _run_web_search_inline(
                            new_query.id, user_profile_for_run
                        ),
                        name=f"convioo-henry-search-{new_query.id}",
                    )

                # Echo the new search_id back into the payload so the
                # frontend can render a "Open session" CTA on the
                # applied-action card.
                applied_payload = dict(action.payload)
                applied_payload["search_id"] = str(new_query.id)
                applied.append(
                    PendingAction(
                        kind=action.kind,
                        summary=action.summary,
                        payload=applied_payload,
                    )
                )
        except Exception:  # noqa: BLE001
            logger.exception(
                "apply_pending_action failed for kind=%s", action.kind
            )
            continue

    if applied or profile_dirty:
        await session.commit()
    return applied


def _result_to_pending_actions(
    result: dict[str, Any], mode: str
) -> list[PendingAction]:
    """Translate Henry's raw JSON output to PendingAction items.

    The LLM still emits ``profile_suggestion`` / ``team_suggestion``
    in its JSON because those shapes are easy for the model to fill;
    this helper flattens them into the user-facing pending_actions
    list (one card per action) the frontend renders.
    """
    out: list[PendingAction] = []
    summary_text = (result.get("suggestion_summary") or "").strip()

    if mode == "personal":
        ps = result.get("profile_suggestion")
        if isinstance(ps, dict):
            cleaned = {
                k: v for k, v in ps.items() if k in _PROFILE_FIELDS_WHITELIST and v
            }
            if cleaned:
                out.append(
                    PendingAction(
                        kind="profile_patch",
                        summary=summary_text or "Записать в профиль",
                        payload=cleaned,
                    )
                )

    if mode == "team_owner":
        ts = result.get("team_suggestion")
        if isinstance(ts, dict):
            description = (ts.get("description") or "").strip()
            if description:
                out.append(
                    PendingAction(
                        kind="team_description",
                        summary=summary_text or "Записать описание команды",
                        payload={"description": description},
                    )
                )
            for md in ts.get("member_descriptions") or []:
                if (
                    isinstance(md, dict)
                    and isinstance(md.get("user_id"), int)
                    and (md.get("description") or "").strip()
                ):
                    out.append(
                        PendingAction(
                            kind="member_description",
                            summary=(
                                f"Записать описание для участника "
                                f"#{md['user_id']}"
                            ),
                            payload={
                                "user_id": md["user_id"],
                                "description": md["description"].strip(),
                            },
                        )
                    )

    # Henry-can-search: LLM emits ``search_request`` with the same
    # shape ``consult_search`` slot dict has. Personal mode only —
    # team mode launches go through the regular form.
    if mode == "personal":
        sr = result.get("search_request")
        if isinstance(sr, dict):
            niche = (sr.get("niche") or "").strip()
            region = (sr.get("region") or "").strip()
            if niche and region:
                payload: dict[str, Any] = {
                    "niche": niche,
                    "region": region,
                }
                ic = (sr.get("ideal_customer") or "").strip()
                if ic:
                    payload["ideal_customer"] = ic
                ex = (sr.get("exclusions") or "").strip()
                if ex:
                    payload["exclusions"] = ex
                out.append(
                    PendingAction(
                        kind="launch_search",
                        summary=f"Run search: {niche} in {region}",
                        payload=payload,
                    )
                )

    return out
