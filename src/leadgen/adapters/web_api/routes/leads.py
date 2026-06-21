"""``/api/v1/leads/*``, ``/api/v1/saved-searches/*``, ``/api/v1/tasks/*`` — CRM core.

Carved from ``app.py`` Wave 9. Covers the full lead lifecycle:
list, export (CSV + XLSX), bulk-update, update, delete, re-enrich,
saved-searches (CRUD + manual run), custom-fields, activities, tasks,
draft-email, bulk-draft, and mark.
"""

from __future__ import annotations

import asyncio
import io
import logging
import uuid
from collections.abc import AsyncIterator
from datetime import datetime, timedelta, timezone
from typing import Any

import sqlalchemy as sa
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import func, select, update

from leadgen.adapters.web_api.auth import get_current_user
from leadgen.adapters.web_api.routes._helpers import (
    LEGACY_LEAD_STATUS_KEYS,
    marks_for_user,
    membership,
    resolve_team_view,
    to_lead_response,
)
from leadgen.adapters.web_api.routes._helpers import (
    run_web_search_inline as _run_web_search_inline,
)
from leadgen.adapters.web_api.routes._helpers import (
    tags_by_lead as _tags_by_lead,
)
from leadgen.adapters.web_api.schemas import (
    BulkDraftEmailItem,
    BulkDraftEmailRequest,
    BulkDraftEmailResponse,
    LeadActivityListResponse,
    LeadBulkUpdateRequest,
    LeadBulkUpdateResponse,
    LeadCustomFieldsResponse,
    LeadCustomFieldUpsert,
    LeadEmailDraftRequest,
    LeadEmailDraftResponse,
    LeadListResponse,
    LeadMarkRequest,
    LeadResponse,
    LeadTaskCreate,
    LeadTaskListResponse,
    LeadTaskUpdate,
    LeadUpdate,
    SavedSearchCreate,
    SavedSearchListResponse,
    SavedSearchSchema,
    SavedSearchUpdate,
    SearchCreateResponse,
)
from leadgen.adapters.web_api.schemas import LeadActivity as LeadActivitySchema
from leadgen.adapters.web_api.schemas import LeadCustomField as LeadCustomFieldSchema
from leadgen.adapters.web_api.schemas import LeadTask as LeadTaskSchema
from leadgen.analysis.ai_analyzer import AIAnalyzer
from leadgen.core.services.webhooks import emit_event_sync as emit_webhook_event_sync
from leadgen.core.services.webhooks import serialize_lead as serialize_lead_for_webhook
from leadgen.db.models import (
    Lead,
    LeadActivity,
    LeadCustomField,
    LeadMark,
    LeadStatus,
    LeadTagAssignment,
    LeadTask,
    SavedSearch,
    SearchQuery,
    TeamMembership,
    TeamSeenLead,
    User,
    UserSeenLead,
)
from leadgen.db.session import session_factory
from leadgen.integrations.slack import send_slack_notification
from leadgen.queue import enqueue_search
from leadgen.utils import spawn
from leadgen.utils.locale_text import normalize_lang
from leadgen.utils.locale_text import pick as locale_pick

logger = logging.getLogger(__name__)

router = APIRouter(tags=["leads"])


async def _authorise_lead_access(
    session, lead_id: uuid.UUID, user_id: int
) -> tuple[Lead, SearchQuery | None]:
    """Load a lead and verify the caller may touch it.

    Allowed when the caller owns the parent search, or is a member of
    the team that search belongs to. Any failure — missing lead,
    missing search, foreign owner — answers 404 (never 403) so lead
    ids can't be probed for existence across accounts.
    """
    lead = await session.get(Lead, lead_id)
    if lead is None:
        raise HTTPException(status_code=404, detail="lead not found")
    search = await session.get(SearchQuery, lead.query_id)
    allowed = search is not None and search.user_id == user_id
    if not allowed and search is not None and search.team_id is not None:
        allowed = (
            await membership(session, search.team_id, user_id)
        ) is not None
    if not allowed:
        raise HTTPException(status_code=404, detail="lead not found")
    return lead, search



# ── archive / unarchive (existing routes) ──────────────────────────────


async def _authorise_lead(
    session,
    lead_id: uuid.UUID,
    current_user: User,
) -> tuple[Lead, SearchQuery]:
    """Load + authorise a lead. Raises 404 / 403 directly on failure."""
    lead = await session.get(Lead, lead_id)
    if lead is None:
        raise HTTPException(status_code=404, detail="lead not found")
    search = await session.get(SearchQuery, lead.query_id)
    if search is None:
        raise HTTPException(status_code=404, detail="search not found")

    allowed = search.user_id == current_user.id
    if not allowed and search.team_id is not None:
        ms = await membership(session, search.team_id, current_user.id)
        allowed = ms is not None
    if not allowed:
        raise HTTPException(status_code=403, detail="forbidden")
    return lead, search


@router.post("/api/v1/leads/{lead_id}/archive")
async def archive_lead_endpoint(
    lead_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
) -> dict[str, bool]:
    """Move a lead to the Archive zone."""
    async with session_factory() as session:
        lead, search = await _authorise_lead(session, lead_id, current_user)
        from leadgen.core.services.lead_archive import archive_lead
        await archive_lead(session, lead, search)
        session.add(
            LeadActivity(
                lead_id=lead.id,
                user_id=current_user.id,
                team_id=search.team_id,
                kind="archived",
                payload={},
            )
        )
        await session.commit()
    return {"ok": True}


@router.post("/api/v1/leads/{lead_id}/unarchive")
async def unarchive_lead_endpoint(
    lead_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
) -> dict[str, bool]:
    """Restore a lead from the Archive zone back to the active CRM."""
    async with session_factory() as session:
        lead, search = await _authorise_lead(session, lead_id, current_user)
        from leadgen.core.services.lead_archive import unarchive_lead
        await unarchive_lead(lead)
        session.add(
            LeadActivity(
                lead_id=lead.id,
                user_id=current_user.id,
                team_id=search.team_id,
                kind="unarchived",
                payload={},
            )
        )
        await session.commit()
    return {"ok": True}


# ── module-level state ───────────────────────────────────────────────────
@router.get("/api/v1/leads", response_model=LeadListResponse)
async def list_all_leads(
    team_id: uuid.UUID | None = None,
    member_user_id: int | None = None,
    lead_status: str | None = None,
    temp: str | None = None,
    created_after: datetime | None = None,
    untouched_days: int | None = None,
    tag_id: uuid.UUID | None = None,
    archived: bool = False,
    limit: int = 200,
    current_user: User = Depends(get_current_user),
) -> LeadListResponse:
    """Cross-session CRM listing.

    Personal mode → caller's own leads. Team mode → caller's own
    leads inside that team by default. Team owners can pass
    ``member_user_id`` to inspect a specific teammate's CRM.

    Filter knobs the frontend's smart-filter chips lean on:
    - ``temp`` ∈ {"hot","warm","cold"} → filters by score buckets
      (hot ≥ 75, warm 50-74, cold < 50).
    - ``created_after`` → ISO timestamp; "новые сегодня" / "за неделю".
    - ``untouched_days`` → leads whose ``last_touched_at`` is older
      than N days (or never touched at all). "Без касания 14+ дней".

    ``mark_color`` on each row is always the *caller's* private
    mark (never the viewed-as user's), so an owner browsing a
    teammate's CRM still sees their own colour codes.
    """
    user_id = current_user.id
    limit = max(1, min(limit, 500))
    # ``archived`` flag splits the list into two zones — active
    # CRM (default) vs the Archive tab. Both still hide soft-deleted
    # rows, only ``archived_at`` flips between IS NULL / IS NOT NULL.
    archived_predicate = (
        Lead.archived_at.is_not(None) if archived else Lead.archived_at.is_(None)
    )
    async with session_factory() as session:
        stmt = (
            select(Lead, SearchQuery.niche, SearchQuery.region)
            .join(SearchQuery, SearchQuery.id == Lead.query_id)
            .where(SearchQuery.source == "web")
            .where(Lead.deleted_at.is_(None))
            .where(archived_predicate)
            .order_by(Lead.score_ai.desc().nullslast(), Lead.created_at.desc())
            .limit(limit)
        )
        total_stmt = (
            select(func.count(Lead.id))
            .join(SearchQuery, SearchQuery.id == Lead.query_id)
            .where(SearchQuery.source == "web")
            .where(Lead.deleted_at.is_(None))
            .where(archived_predicate)
        )
        if team_id is not None:
            target_user = await resolve_team_view(
                session, team_id, user_id, member_user_id
            )
            stmt = stmt.where(SearchQuery.team_id == team_id).where(
                SearchQuery.user_id == target_user
            )
            total_stmt = total_stmt.where(
                SearchQuery.team_id == team_id
            ).where(SearchQuery.user_id == target_user)
        else:
            stmt = stmt.where(SearchQuery.user_id == user_id).where(
                SearchQuery.team_id.is_(None)
            )
            total_stmt = total_stmt.where(
                SearchQuery.user_id == user_id
            ).where(SearchQuery.team_id.is_(None))
        if lead_status:
            stmt = stmt.where(Lead.lead_status == lead_status)
            total_stmt = total_stmt.where(Lead.lead_status == lead_status)
        if temp == "hot":
            stmt = stmt.where(Lead.score_ai >= 75)
            total_stmt = total_stmt.where(Lead.score_ai >= 75)
        elif temp == "warm":
            stmt = stmt.where(Lead.score_ai >= 50).where(Lead.score_ai < 75)
            total_stmt = total_stmt.where(Lead.score_ai >= 50).where(
                Lead.score_ai < 75
            )
        elif temp == "cold":
            stmt = stmt.where(
                (Lead.score_ai < 50) | (Lead.score_ai.is_(None))
            )
            total_stmt = total_stmt.where(
                (Lead.score_ai < 50) | (Lead.score_ai.is_(None))
            )
        if created_after is not None:
            stmt = stmt.where(Lead.created_at >= created_after)
            total_stmt = total_stmt.where(Lead.created_at >= created_after)
        if untouched_days and untouched_days > 0:
            cutoff = datetime.now(timezone.utc) - timedelta(
                days=untouched_days
            )
            stmt = stmt.where(
                (Lead.last_touched_at < cutoff)
                | (Lead.last_touched_at.is_(None))
            )
            total_stmt = total_stmt.where(
                (Lead.last_touched_at < cutoff)
                | (Lead.last_touched_at.is_(None))
            )
        if tag_id is not None:
            tagged_subq = (
                select(LeadTagAssignment.lead_id)
                .where(LeadTagAssignment.tag_id == tag_id)
                .subquery()
            )
            stmt = stmt.where(Lead.id.in_(select(tagged_subq.c.lead_id)))
            total_stmt = total_stmt.where(
                Lead.id.in_(select(tagged_subq.c.lead_id))
            )
        rows = (await session.execute(stmt)).all()

        lead_ids = [lead.id for lead, _n, _r in rows]
        total = int((await session.execute(total_stmt)).scalar() or 0)
        marks = await marks_for_user(session, user_id, lead_ids)
        tags_by_lead = await _tags_by_lead(session, lead_ids)

    leads: list[LeadResponse] = []
    sessions_by_id: dict[str, dict[str, Any]] = {}
    for lead, niche, region in rows:
        leads.append(
            to_lead_response(
                lead, marks.get(lead.id), tags_by_lead.get(lead.id)
            )
        )
        sessions_by_id[str(lead.query_id)] = {"niche": niche, "region": region}
    return LeadListResponse(leads=leads, total=total, sessions_by_id=sessions_by_id)

@router.get("/api/v1/leads/export.csv", include_in_schema=False)
async def export_leads_csv(
    team_id: uuid.UUID | None = None,
    member_user_id: int | None = None,
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    """Export the caller's CRM rows as a CSV file.

    Streamed in 500-row chunks so the response starts before the
    whole result set is in memory. Mirrors the same scoping as the
    JSON list endpoint (personal / team / view-as) but ignores the
    smart-filter knobs — export is always "everything in this
    scope" so the file is the complete copy.
    """
    user_id = current_user.id
    import csv as _csv
    import io as _io

    # Hand-rolled CSV — keeps the deps tight (no openpyxl/pandas in
    # the request path). Columns are intentionally narrow: the
    # things you'd actually paste into another CRM.
    header = [
        "name",
        "niche",
        "region",
        "score",
        "lead_status",
        "rating",
        "reviews_count",
        "phone",
        "website",
        "address",
        "category",
        "notes",
        "last_touched_at",
        "created_at",
    ]

    def _row_bytes(values: list[Any]) -> bytes:
        buf = _io.StringIO()
        _csv.writer(buf, quoting=_csv.QUOTE_MINIMAL).writerow(values)
        return buf.getvalue().encode("utf-8")

    async def generate() -> AsyncIterator[bytes]:
        # UTF-8 BOM so Excel on Windows opens Cyrillic columns cleanly.
        yield b"\xef\xbb\xbf"
        yield _row_bytes(header)
        async with session_factory() as session:
            stmt = (
                select(Lead, SearchQuery.niche, SearchQuery.region)
                .join(SearchQuery, SearchQuery.id == Lead.query_id)
                .where(SearchQuery.source == "web")
                .where(Lead.deleted_at.is_(None))
                .order_by(
                    Lead.score_ai.desc().nullslast(),
                    Lead.created_at.desc(),
                )
                .limit(50_000)
            )
            if team_id is not None:
                target_user = await resolve_team_view(
                    session, team_id, user_id, member_user_id
                )
                stmt = stmt.where(
                    SearchQuery.team_id == team_id
                ).where(SearchQuery.user_id == target_user)
            else:
                stmt = stmt.where(SearchQuery.user_id == user_id).where(
                    SearchQuery.team_id.is_(None)
                )
            result = await session.stream(stmt.execution_options(yield_per=500))
            async for lead, niche, region in result:
                yield _row_bytes(
                    [
                        lead.name or "",
                        niche or "",
                        region or "",
                        ""
                        if lead.score_ai is None
                        else int(round(lead.score_ai)),
                        lead.lead_status or "",
                        "" if lead.rating is None else lead.rating,
                        ""
                        if lead.reviews_count is None
                        else lead.reviews_count,
                        lead.phone or "",
                        lead.website or "",
                        lead.address or "",
                        lead.category or "",
                        (lead.notes or "").replace("\n", " "),
                        lead.last_touched_at.isoformat()
                        if lead.last_touched_at
                        else "",
                        lead.created_at.isoformat()
                        if lead.created_at
                        else "",
                    ]
                )

    filename = f"convioo-leads-{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        generate(),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )

@router.get(
    "/api/v1/searches/{query_id}/export.xlsx", include_in_schema=False
)
async def export_session_xlsx(
    query_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
) -> Response:
    """Export one search session as a styled Excel workbook.

    One sheet, header row formatted bold, frozen first row, columns
    auto-fit-ish. The deliberately narrow column set matches the CSV
    export so the user gets the same shape they're used to plus the
    extra polish (cell types, no BOM hack) that Excel users expect.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    async with session_factory() as session:
        query = await session.get(SearchQuery, query_id)
        if query is None:
            raise HTTPException(status_code=404, detail="search not found")
        # Cross-user access answers 404 (not 403) so the export URL
        # can't be used to probe which session ids exist.
        allowed = query.user_id == current_user.id
        if not allowed and query.team_id is not None:
            allowed = (
                await membership(session, query.team_id, current_user.id)
            ) is not None
        if not allowed:
            raise HTTPException(status_code=404, detail="search not found")
        rows = list(
            (
                await session.execute(
                    select(Lead)
                    .where(Lead.query_id == query_id)
                    .where(Lead.deleted_at.is_(None))
                    .order_by(
                        Lead.score_ai.desc().nullslast(),
                        Lead.created_at.desc(),
                    )
                )
            )
            .scalars()
            .all()
        )

    _xlsx_lang = normalize_lang(current_user.language_code)
    _xlsx_headers = {
        "ru": [
            "Название",
            "Скор",
            "Статус",
            "Рейтинг",
            "Отзывов",
            "Телефон",
            "Сайт",
            "Адрес",
            "Категория",
            "Заметки",
            "Последнее касание",
            "Создан",
        ],
        "uk": [
            "Назва",
            "Скор",
            "Статус",
            "Рейтинг",
            "Відгуків",
            "Телефон",
            "Сайт",
            "Адреса",
            "Категорія",
            "Нотатки",
            "Останній контакт",
            "Створено",
        ],
        "en": [
            "Name",
            "Score",
            "Status",
            "Rating",
            "Reviews",
            "Phone",
            "Website",
            "Address",
            "Category",
            "Notes",
            "Last touched",
            "Created",
        ],
    }
    headers = _xlsx_headers[_xlsx_lang]

    # openpyxl is pure-Python and CPU-bound; running it inline blocks
    # the event loop for the entire workbook build + zip. Offload to
    # the default thread pool so other requests keep flowing while
    # the export builds.
    def _build_workbook() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = (query.niche or "leads")[:30]
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="3D5AFE",
            end_color="3D5AFE",
            fill_type="solid",
        )
        header_align = Alignment(vertical="center")
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for lead in rows:
            ws.append(
                [
                    lead.name or "",
                    ""
                    if lead.score_ai is None
                    else int(round(lead.score_ai)),
                    lead.lead_status or "",
                    "" if lead.rating is None else lead.rating,
                    ""
                    if lead.reviews_count is None
                    else lead.reviews_count,
                    lead.phone or "",
                    lead.website or "",
                    lead.address or "",
                    lead.category or "",
                    (lead.notes or "").replace("\n", " "),
                    lead.last_touched_at.isoformat()
                    if lead.last_touched_at
                    else "",
                    lead.created_at.isoformat()
                    if lead.created_at
                    else "",
                ]
            )

        widths = [32, 8, 12, 8, 10, 18, 36, 36, 22, 40, 22, 22]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    body = await asyncio.to_thread(_build_workbook)
    slug = (query.niche or "session").replace(" ", "-").lower()[:40]
    date = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"convioo-{slug}-{date}.xlsx"
    return Response(
        content=body,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )

# NOTE: registered BEFORE the /api/v1/leads/{lead_id} routes —
# Starlette matches in registration order and the literal
# "bulk" segment would otherwise be captured by {lead_id}.
@router.patch(
    "/api/v1/leads/bulk", response_model=LeadBulkUpdateResponse
)
async def bulk_update_leads(
    body: LeadBulkUpdateRequest,
    current_user: User = Depends(get_current_user),
) -> LeadBulkUpdateResponse:
    """Apply ``lead_status`` and/or the caller's mark to many leads
    in one round-trip. The CRM bulk-toolbar uses this so the user
    can sweep dozens of rows in one click.

    Only leads the caller owns (or shares a team with via the
    parent search) are touched — foreign ids in the payload are
    silently dropped from the update set.
    """
    if not body.lead_status and not body.set_mark_color:
        raise HTTPException(
            status_code=400, detail="nothing to update"
        )

    async with session_factory() as session:
        # Authorise per-lead: keep only ids whose parent search the
        # caller owns or can reach through a team membership.
        owner_rows = (
            await session.execute(
                select(Lead.id, SearchQuery.user_id, SearchQuery.team_id)
                .join(SearchQuery, SearchQuery.id == Lead.query_id)
                .where(Lead.id.in_(body.lead_ids))
            )
        ).all()
        team_member_cache: dict[uuid.UUID, bool] = {}
        allowed_ids: list[uuid.UUID] = []
        for row_lead_id, owner_id, owner_team_id in owner_rows:
            if owner_id == current_user.id:
                allowed_ids.append(row_lead_id)
                continue
            if owner_team_id is None:
                continue
            is_member = team_member_cache.get(owner_team_id)
            if is_member is None:
                is_member = (
                    await membership(
                        session, owner_team_id, current_user.id
                    )
                ) is not None
                team_member_cache[owner_team_id] = is_member
            if is_member:
                allowed_ids.append(row_lead_id)
        # Permissive validation: accept if matches a legacy key OR
        # any team's custom palette. Bulk operations span teams so
        # a strict per-team check would block mixed selections.
        if (
            body.lead_status
            and body.lead_status not in LEGACY_LEAD_STATUS_KEYS
        ):
            custom = (
                await session.execute(
                    select(LeadStatus.key).where(
                        LeadStatus.key == body.lead_status
                    ).limit(1)
                )
            ).scalar_one_or_none()
            if custom is None:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "lead_status is not a valid key in any "
                        "team palette or the default set"
                    ),
                )
        if body.lead_status and allowed_ids:
            await session.execute(
                update(Lead)
                .where(Lead.id.in_(allowed_ids))
                .values(
                    lead_status=body.lead_status,
                    last_touched_at=datetime.now(timezone.utc),
                )
            )

        if body.set_mark_color and allowed_ids:
            color = (body.mark_color or "").strip() or None
            if color is None:
                await session.execute(
                    sa.delete(LeadMark)
                    .where(LeadMark.user_id == current_user.id)
                    .where(LeadMark.lead_id.in_(allowed_ids))
                )
            else:
                # Per-row upsert. Postgres ON CONFLICT keeps it cheap;
                # SQLite (test harness) iterates Python-side.
                from sqlalchemy.dialects.postgresql import (
                    insert as pg_insert,
                )

                rows = [
                    {
                        "user_id": current_user.id,
                        "lead_id": lid,
                        "color": color,
                        "updated_at": datetime.now(timezone.utc),
                    }
                    for lid in allowed_ids
                ]
                if session.bind.dialect.name == "postgresql":
                    stmt = pg_insert(LeadMark).values(rows)
                    stmt = stmt.on_conflict_do_update(
                        index_elements=["user_id", "lead_id"],
                        set_={
                            "color": color,
                            "updated_at": datetime.now(timezone.utc),
                        },
                    )
                    await session.execute(stmt)
                else:
                    for r in rows:
                        existing = (
                            await session.execute(
                                select(LeadMark)
                                .where(LeadMark.user_id == r["user_id"])
                                .where(LeadMark.lead_id == r["lead_id"])
                            )
                        ).scalar_one_or_none()
                        if existing:
                            existing.color = color
                            existing.updated_at = r["updated_at"]
                        else:
                            session.add(LeadMark(**r))

        await session.commit()

        # Touched rows = requested ids that exist AND passed the
        # ownership filter.
        return LeadBulkUpdateResponse(updated=len(allowed_ids))

@router.patch("/api/v1/leads/{lead_id}", response_model=LeadResponse)
async def update_lead(
    lead_id: uuid.UUID,
    body: LeadUpdate,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
) -> LeadResponse:
    """Partial update: status, owner, notes. Touches last_touched_at.

    Writes ``lead_activities`` rows per changed field so the
    timeline + team feed have something to render. The actor is
    always the authenticated user — used to point at a query-param
    default which broke the lead_activities FK.
    """
    actor_user_id = current_user.id
    async with session_factory() as session:
        lead = await session.get(Lead, lead_id)
        if lead is None:
            raise HTTPException(status_code=404, detail="lead not found")

        # Ownership: the caller must own the parent search or be a
        # member of the team it belongs to. Cross-user access gets
        # a 404 (not 403) so lead ids can't be probed for existence.
        search = await session.get(SearchQuery, lead.query_id)
        allowed = search is not None and search.user_id == actor_user_id
        if not allowed and search is not None and search.team_id is not None:
            allowed = (
                await membership(session, search.team_id, actor_user_id)
            ) is not None
        if not allowed:
            raise HTTPException(status_code=404, detail="lead not found")

        # Lead-status validation: team-mode searches use the
        # team's custom palette; personal-mode searches keep the
        # legacy hard-coded keys. Either way an unknown key fails.
        if body.lead_status is not None:
            search_for_status = search
            valid_keys: set[str] | frozenset[str]
            if search_for_status and search_for_status.team_id is not None:
                valid_keys = {
                    k for (k,) in (
                        await session.execute(
                            select(LeadStatus.key).where(
                                LeadStatus.team_id == search_for_status.team_id
                            )
                        )
                    ).all()
                }
                # Defensive fallback — if the team's palette wasn't
                # seeded for some reason, accept the legacy keys
                # rather than rejecting every drag-and-drop.
                if not valid_keys:
                    valid_keys = set(LEGACY_LEAD_STATUS_KEYS)
            else:
                valid_keys = LEGACY_LEAD_STATUS_KEYS
            if body.lead_status not in valid_keys:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "lead_status must be one of "
                        + ", ".join(sorted(valid_keys))
                    ),
                )

        # Capture before/after so we can write meaningful activity
        # rows. The fields list mirrors what LeadUpdate exposes —
        # if a new field gets added there, add it here too.
        activities: list[dict[str, Any]] = []
        now = datetime.now(timezone.utc)

        if body.lead_status is not None and body.lead_status != lead.lead_status:
            activities.append(
                {
                    "kind": "status",
                    "payload": {
                        "from": lead.lead_status,
                        "to": body.lead_status,
                    },
                }
            )
            lead.lead_status = body.lead_status
        if "owner_user_id" in body.model_fields_set:
            if body.owner_user_id != lead.owner_user_id:
                activities.append(
                    {
                        "kind": "assigned",
                        "payload": {
                            "from": lead.owner_user_id,
                            "to": body.owner_user_id,
                        },
                    }
                )
            lead.owner_user_id = body.owner_user_id
        if "deal_value" in body.model_fields_set:
            lead.deal_value = body.deal_value

        if body.notes is not None and body.notes != (lead.notes or ""):
            activities.append(
                {
                    "kind": "notes",
                    "payload": {"len": len(body.notes)},
                }
            )
            lead.notes = body.notes

        if not activities and (
            body.lead_status is None
            and body.notes is None
            and "owner_user_id" not in body.model_fields_set
            and "deal_value" not in body.model_fields_set
        ):
            raise HTTPException(status_code=400, detail="no fields to update")

        lead.last_touched_at = now

        # Pull team_id off the parent search query so the activity
        # row can land in the team feed when the lead is shared.
        team_id_for_activity = search.team_id if search else None

        for act in activities:
            session.add(
                LeadActivity(
                    lead_id=lead.id,
                    user_id=actor_user_id,
                    team_id=team_id_for_activity,
                    kind=act["kind"],
                    payload=act["payload"],
                )
            )
        await session.commit()
        await session.refresh(lead)

        # Emit lead.status_changed if the status moved this round.
        # We notify the search owner — they're who registered the
        # webhook against their account, regardless of who edited
        # it inside the team.
        status_change = next(
            (a for a in activities if a["kind"] == "status"), None
        )
        if status_change and search is not None:
            background_tasks.add_task(
                emit_webhook_event_sync,
                search.user_id,
                "lead.status_changed",
                {
                    "lead": serialize_lead_for_webhook(lead),
                    "from_status": status_change["payload"]["from"],
                    "to_status": status_change["payload"]["to"],
                    "actor_user_id": actor_user_id,
                },
            )
            if status_change["payload"]["to"] == "won":
                background_tasks.add_task(
                    send_slack_notification,
                    f"Lead won: {lead.name} ({lead.website})",
                )
        return LeadResponse.model_validate(lead)

@router.delete("/api/v1/leads/{lead_id}")
async def delete_lead(
    lead_id: uuid.UUID,
    forever: bool = False,
    current_user: User = Depends(get_current_user),
) -> dict[str, bool]:
    """Soft-delete a lead so it disappears from the CRM.

    ``forever=true`` additionally writes a row into the seen-leads
    table so future searches will treat the same place_id /
    phone / domain as already-delivered and skip it. Without
    ``forever``, the lead is just hidden — re-running a similar
    search may surface it again.

    Authorisation: caller must own the parent ``SearchQuery`` (or
    be a member of the team that owns it).
    """
    async with session_factory() as session:
        lead = await session.get(Lead, lead_id)
        if lead is None:
            raise HTTPException(status_code=404, detail="lead not found")
        search = await session.get(SearchQuery, lead.query_id)
        if search is None:
            raise HTTPException(status_code=404, detail="search not found")

        allowed = search.user_id == current_user.id
        if not allowed and search.team_id is not None:
            ms = await membership(
                session, search.team_id, current_user.id
            )
            allowed = ms is not None
        if not allowed:
            raise HTTPException(status_code=403, detail="forbidden")

        if lead.deleted_at is None:
            lead.deleted_at = datetime.now(timezone.utc)
        if forever:
            lead.blacklisted = True
            # Make sure the seen-leads record exists with all three
            # dedup axes filled, even if the lead came in before the
            # 0023 migration backfilled them.
            from leadgen.utils.dedup import (
                domain_root as _domain_root,
            )
            from leadgen.utils.dedup import (
                normalize_phone as _normalize_phone,
            )

            phone_key = _normalize_phone(lead.phone)
            domain_key = _domain_root(lead.website)

            if search.user_id != 0:
                existing_user = (
                    await session.execute(
                        select(UserSeenLead)
                        .where(UserSeenLead.user_id == search.user_id)
                        .where(UserSeenLead.source == lead.source)
                        .where(UserSeenLead.source_id == lead.source_id)
                    )
                ).scalar_one_or_none()
                if existing_user is None:
                    session.add(
                        UserSeenLead(
                            user_id=search.user_id,
                            source=lead.source,
                            source_id=lead.source_id,
                            phone_e164=phone_key,
                            domain_root=domain_key,
                        )
                    )
                else:
                    existing_user.phone_e164 = phone_key
                    existing_user.domain_root = domain_key
            if search.team_id is not None:
                existing_team = (
                    await session.execute(
                        select(TeamSeenLead)
                        .where(TeamSeenLead.team_id == search.team_id)
                        .where(TeamSeenLead.source == lead.source)
                        .where(TeamSeenLead.source_id == lead.source_id)
                    )
                ).scalar_one_or_none()
                if existing_team is None:
                    session.add(
                        TeamSeenLead(
                            team_id=search.team_id,
                            source=lead.source,
                            source_id=lead.source_id,
                            phone_e164=phone_key,
                            domain_root=domain_key,
                            first_user_id=search.user_id,
                        )
                    )
                else:
                    existing_team.phone_e164 = phone_key
                    existing_team.domain_root = domain_key

        session.add(
            LeadActivity(
                lead_id=lead.id,
                user_id=current_user.id,
                team_id=search.team_id,
                kind="deleted",
                payload={"forever": bool(forever)},
            )
        )
        await session.commit()
    return {"ok": True, "forever": bool(forever)}

_enriching_leads: set[str] = set()

@router.post("/api/v1/leads/{lead_id}/re-enrich", response_model=LeadResponse)
async def re_enrich_lead(
    lead_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
) -> LeadResponse:
    """Trigger a fresh AI enrichment pass for a single lead.

    Returns 409 if enrichment is already running for this lead_id.
    """
    from leadgen.collectors import GooglePlacesCollector
    from leadgen.pipeline.enrichment import enrich_leads

    lead_id_str = str(lead_id)
    if lead_id_str in _enriching_leads:
        raise HTTPException(status_code=409, detail="enrichment already in progress")
    _enriching_leads.add(lead_id_str)
    try:
        async with session_factory() as session:
            lead = await session.get(Lead, lead_id)
            if lead is None:
                raise HTTPException(status_code=404, detail="lead not found")
            search = await session.get(SearchQuery, lead.query_id)
            if search is None:
                raise HTTPException(status_code=404, detail="search not found")
            allowed = search.user_id == current_user.id
            if not allowed and search.team_id is not None:
                allowed = (
                    await membership(
                        session, search.team_id, current_user.id
                    )
                ) is not None
            if not allowed:
                raise HTTPException(
                    status_code=404, detail="lead not found"
                )

        collector = GooglePlacesCollector()
        await enrich_leads(
            [lead],
            collector,
            search.niche,
            search.region,
        )

        async with session_factory() as session:
            updated = await session.get(Lead, lead_id)
            if updated is None:
                raise HTTPException(status_code=404, detail="lead not found")
            return LeadResponse.model_validate(updated)
    finally:
        _enriching_leads.discard(lead_id_str)

# ── /api/v1/saved-searches (bookmark + recurring re-run) ───────────

def _saved_to_schema(row: SavedSearch) -> SavedSearchSchema:
    return SavedSearchSchema(
        id=str(row.id),
        name=row.name,
        team_id=str(row.team_id) if row.team_id else None,
        niche=row.niche,
        region=row.region,
        target_languages=row.target_languages,
        scope=row.scope,
        radius_m=row.radius_m,
        max_results=row.max_results,
        schedule=row.schedule,
        next_run_at=row.next_run_at,
        last_run_at=row.last_run_at,
        last_leads_count=row.last_leads_count,
        active=row.active,
        created_at=row.created_at,
        updated_at=row.updated_at,
    )

def _normalize_schedule(raw: str | None) -> str | None:
    """Map ``"off"`` and the empty string onto ``None`` so the
    worker query can ``WHERE schedule IS NOT NULL`` cleanly."""
    if not raw:
        return None
    value = raw.strip().lower()
    if value in ("off", "none", "manual", ""):
        return None
    from leadgen.core.services.saved_searches import VALID_SCHEDULES

    if value not in VALID_SCHEDULES:
        raise HTTPException(
            status_code=400,
            detail=(
                "schedule must be one of off / daily / weekly / "
                "biweekly / monthly"
            ),
        )
    return value

@router.get(
    "/api/v1/saved-searches",
    response_model=SavedSearchListResponse,
)
async def list_saved_searches(
    current_user: User = Depends(get_current_user),
) -> SavedSearchListResponse:
    async with session_factory() as session:
        team_ids = (
            (
                await session.execute(
                    select(TeamMembership.team_id).where(
                        TeamMembership.user_id == current_user.id
                    )
                )
            )
            .scalars()
            .all()
        )
        stmt = (
            select(SavedSearch)
            .where(
                sa.or_(
                    sa.and_(
                        SavedSearch.user_id == current_user.id,
                        SavedSearch.team_id.is_(None),
                    ),
                    SavedSearch.team_id.in_(team_ids)
                    if team_ids
                    else sa.false(),
                )
            )
            .order_by(SavedSearch.created_at.desc())
        )
        rows = (await session.execute(stmt)).scalars().all()
    return SavedSearchListResponse(
        items=[_saved_to_schema(r) for r in rows]
    )

@router.post(
    "/api/v1/saved-searches",
    response_model=SavedSearchSchema,
)
async def create_saved_search(
    body: SavedSearchCreate,
    current_user: User = Depends(get_current_user),
) -> SavedSearchSchema:
    from leadgen.core.services.saved_searches import (
        next_run_after,
    )

    team_uuid: uuid.UUID | None = None
    if body.team_id:
        try:
            team_uuid = uuid.UUID(body.team_id)
        except ValueError as exc:
            raise HTTPException(
                status_code=400, detail="invalid team_id"
            ) from exc

    schedule = _normalize_schedule(body.schedule)
    async with session_factory() as session:
        if team_uuid is not None:
            membership = (
                await session.execute(
                    select(TeamMembership)
                    .where(TeamMembership.user_id == current_user.id)
                    .where(TeamMembership.team_id == team_uuid)
                )
            ).scalar_one_or_none()
            if membership is None:
                raise HTTPException(
                    status_code=403, detail="not a team member"
                )
        row = SavedSearch(
            user_id=current_user.id,
            team_id=team_uuid,
            name=body.name.strip(),
            niche=body.niche.strip(),
            region=body.region.strip(),
            target_languages=body.target_languages,
            scope=body.scope,
            radius_m=body.radius_m,
            max_results=body.max_results,
            schedule=schedule,
            next_run_at=next_run_after(schedule)
            if schedule
            else None,
            active=True,
        )
        session.add(row)
        await session.commit()
        await session.refresh(row)
    return _saved_to_schema(row)

@router.patch(
    "/api/v1/saved-searches/{saved_id}",
    response_model=SavedSearchSchema,
)
async def update_saved_search(
    saved_id: uuid.UUID,
    body: SavedSearchUpdate,
    current_user: User = Depends(get_current_user),
) -> SavedSearchSchema:
    from leadgen.core.services.saved_searches import (
        next_run_after,
    )

    async with session_factory() as session:
        row = await session.get(SavedSearch, saved_id)
        if row is None or row.user_id != current_user.id:
            raise HTTPException(
                status_code=404, detail="saved search not found"
            )
        if body.name is not None:
            row.name = body.name.strip()
        if body.schedule is not None:
            row.schedule = _normalize_schedule(body.schedule)
            row.next_run_at = (
                next_run_after(row.schedule) if row.schedule else None
            )
        if body.active is not None:
            row.active = body.active
        if body.max_results is not None:
            row.max_results = body.max_results
        if body.radius_m is not None:
            row.radius_m = body.radius_m
        row.updated_at = datetime.now(timezone.utc)
        await session.commit()
        await session.refresh(row)
    return _saved_to_schema(row)

@router.delete("/api/v1/saved-searches/{saved_id}")
async def delete_saved_search(
    saved_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
) -> dict[str, bool]:
    async with session_factory() as session:
        row = await session.get(SavedSearch, saved_id)
        if row is None or row.user_id != current_user.id:
            raise HTTPException(
                status_code=404, detail="saved search not found"
            )
        await session.delete(row)
        await session.commit()
    return {"ok": True}

@router.post(
    "/api/v1/saved-searches/{saved_id}/run",
    response_model=SearchCreateResponse,
)
async def run_saved_search_now(
    saved_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
) -> SearchCreateResponse:
    """Manually trigger a saved search outside its schedule.

    Useful for the "Run now" button next to each saved search row.
    Reuses the same enqueue plumbing that ``POST /searches`` does so
    the SSE progress stream and the CRM lead-list are identical.
    """
    from leadgen.core.services.saved_searches import build_search_query

    async with session_factory() as session:
        row = await session.get(SavedSearch, saved_id)
        if row is None or row.user_id != current_user.id:
            raise HTTPException(
                status_code=404, detail="saved search not found"
            )
        user = await session.get(User, current_user.id)
        new_query = build_search_query(row)
        session.add(new_query)
        row.last_run_at = datetime.now(timezone.utc)
        await session.commit()
        query_id = new_query.id

    user_profile = {
        "display_name": user.display_name or user.first_name if user else None,
        "language_code": user.language_code if user else None,
    }
    queued_id = await enqueue_search(
        query_id, chat_id=None, user_profile=user_profile
    )
    queued = bool(queued_id)
    if not queued:
        spawn(
            _run_web_search_inline(query_id, user_profile),
            name=f"convioo-web-search-{query_id}",
        )
    return SearchCreateResponse(id=query_id, queued=queued)

# ── /api/v1/leads/{id}/custom-fields ────────────────────────────────

@router.get(
    "/api/v1/leads/{lead_id}/custom-fields",
    response_model=LeadCustomFieldsResponse,
)
async def list_lead_custom_fields(
    lead_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
) -> LeadCustomFieldsResponse:
    user_id = current_user.id
    async with session_factory() as session:
        stmt = (
            select(LeadCustomField)
            .where(LeadCustomField.lead_id == lead_id)
            .where(LeadCustomField.user_id == user_id)
            .order_by(LeadCustomField.key)
        )
        rows = (await session.execute(stmt)).scalars().all()
        items = [
            LeadCustomFieldSchema.model_validate(r) for r in rows
        ]
    return LeadCustomFieldsResponse(items=items)

@router.put(
    "/api/v1/leads/{lead_id}/custom-fields",
    response_model=LeadCustomFieldSchema,
)
async def upsert_lead_custom_field(
    lead_id: uuid.UUID,
    body: LeadCustomFieldUpsert,
    current_user: User = Depends(get_current_user),
) -> LeadCustomFieldSchema:
    """Create or update one (key, value) pair on this lead.

    Schemaless — the user picks any key from the UI. ``value`` may
    be NULL, which acts as a soft-delete on the row (we still keep
    the row so the timeline can reference the historical key).
    """
    user_id = current_user.id
    key = body.key.strip()
    if not key:
        raise HTTPException(status_code=400, detail="key is required")
    value = body.value if body.value is None else body.value.strip()
    async with session_factory() as session:
        existing = (
            await session.execute(
                select(LeadCustomField)
                .where(LeadCustomField.lead_id == lead_id)
                .where(LeadCustomField.user_id == user_id)
                .where(LeadCustomField.key == key)
                .limit(1)
            )
        ).scalar_one_or_none()
        now = datetime.now(timezone.utc)
        search = (
            await session.execute(
                select(SearchQuery)
                .join(Lead, Lead.query_id == SearchQuery.id)
                .where(Lead.id == lead_id)
                .limit(1)
            )
        ).scalar_one_or_none()
        team_id_for_activity = search.team_id if search else None
        if existing is None:
            existing = LeadCustomField(
                lead_id=lead_id,
                user_id=user_id,
                key=key,
                value=value,
            )
            session.add(existing)
        else:
            existing.value = value
            existing.updated_at = now
        session.add(
            LeadActivity(
                lead_id=lead_id,
                user_id=user_id,
                team_id=team_id_for_activity,
                kind="custom_field",
                payload={"key": key, "value": value},
            )
        )
        await session.commit()
        await session.refresh(existing)
        return LeadCustomFieldSchema.model_validate(existing)

@router.delete("/api/v1/leads/{lead_id}/custom-fields/{key}")
async def delete_lead_custom_field(
    lead_id: uuid.UUID,
    key: str,
    current_user: User = Depends(get_current_user),
) -> dict[str, bool]:
    user_id = current_user.id
    async with session_factory() as session:
        row = (
            await session.execute(
                select(LeadCustomField)
                .where(LeadCustomField.lead_id == lead_id)
                .where(LeadCustomField.user_id == user_id)
                .where(LeadCustomField.key == key)
                .limit(1)
            )
        ).scalar_one_or_none()
        if row is None:
            return {"deleted": False}
        await session.delete(row)
        await session.commit()
    return {"deleted": True}

# ── /api/v1/leads/{id}/activity ─────────────────────────────────────

@router.get(
    "/api/v1/leads/{lead_id}/activity",
    response_model=LeadActivityListResponse,
)
async def list_lead_activity(
    lead_id: uuid.UUID,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
) -> LeadActivityListResponse:
    limit = max(1, min(limit, 200))
    async with session_factory() as session:
        await _authorise_lead_access(session, lead_id, current_user.id)
        stmt = (
            select(LeadActivity)
            .where(LeadActivity.lead_id == lead_id)
            .order_by(LeadActivity.created_at.desc())
            .limit(limit)
        )
        rows = (await session.execute(stmt)).scalars().all()
        items = [LeadActivitySchema.model_validate(r) for r in rows]
    return LeadActivityListResponse(items=items)

# ── /api/v1/leads/{id}/tasks ────────────────────────────────────────

@router.get(
    "/api/v1/leads/{lead_id}/tasks",
    response_model=LeadTaskListResponse,
)
async def list_lead_tasks(
    lead_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
) -> LeadTaskListResponse:
    user_id = current_user.id
    async with session_factory() as session:
        stmt = (
            select(LeadTask)
            .where(LeadTask.lead_id == lead_id)
            .where(LeadTask.user_id == user_id)
            .order_by(
                LeadTask.done_at.is_(None).desc(),
                LeadTask.due_at.asc().nullslast(),
                LeadTask.created_at.desc(),
            )
        )
        rows = (await session.execute(stmt)).scalars().all()
        items = [LeadTaskSchema.model_validate(r) for r in rows]
    return LeadTaskListResponse(items=items)

@router.post(
    "/api/v1/leads/{lead_id}/tasks",
    response_model=LeadTaskSchema,
)
async def create_lead_task(
    lead_id: uuid.UUID,
    body: LeadTaskCreate,
    current_user: User = Depends(get_current_user),
) -> LeadTaskSchema:
    user_id = current_user.id
    async with session_factory() as session:
        row = LeadTask(
            lead_id=lead_id,
            user_id=user_id,
            content=body.content.strip(),
            due_at=body.due_at,
        )
        session.add(row)
        search = (
            await session.execute(
                select(SearchQuery)
                .join(Lead, Lead.query_id == SearchQuery.id)
                .where(Lead.id == lead_id)
                .limit(1)
            )
        ).scalar_one_or_none()
        session.add(
            LeadActivity(
                lead_id=lead_id,
                user_id=user_id,
                team_id=search.team_id if search else None,
                kind="task",
                payload={
                    "content": body.content.strip()[:200],
                    "due_at": body.due_at.isoformat() if body.due_at else None,
                },
            )
        )
        await session.commit()
        await session.refresh(row)
        return LeadTaskSchema.model_validate(row)

@router.patch(
    "/api/v1/tasks/{task_id}",
    response_model=LeadTaskSchema,
)
async def update_lead_task(
    task_id: uuid.UUID,
    body: LeadTaskUpdate,
    current_user: User = Depends(get_current_user),
) -> LeadTaskSchema:
    user_id = current_user.id
    async with session_factory() as session:
        row = await session.get(LeadTask, task_id)
        if row is None or row.user_id != user_id:
            raise HTTPException(status_code=404, detail="task not found")
        data = body.model_dump(exclude_unset=True)
        if "content" in data and data["content"]:
            row.content = data["content"].strip()
        if "due_at" in data:
            row.due_at = data["due_at"]
        if "done" in data and data["done"] is not None:
            row.done_at = (
                datetime.now(timezone.utc) if data["done"] else None
            )
        await session.commit()
        await session.refresh(row)
        return LeadTaskSchema.model_validate(row)

@router.delete("/api/v1/tasks/{task_id}")
async def delete_lead_task(
    task_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
) -> dict[str, bool]:
    user_id = current_user.id
    async with session_factory() as session:
        row = await session.get(LeadTask, task_id)
        if row is None or row.user_id != user_id:
            return {"deleted": False}
        await session.delete(row)
        await session.commit()
    return {"deleted": True}

@router.get(
    "/api/v1/users/me/tasks",
    response_model=LeadTaskListResponse,
)
async def list_my_tasks(
    open_only: bool = True,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
) -> LeadTaskListResponse:
    """Today's-tasks widget feed: open tasks across every lead."""
    user_id = current_user.id
    limit = max(1, min(limit, 500))
    async with session_factory() as session:
        stmt = select(LeadTask).where(LeadTask.user_id == user_id)
        if open_only:
            stmt = stmt.where(LeadTask.done_at.is_(None))
        stmt = stmt.order_by(
            LeadTask.due_at.asc().nullslast(),
            LeadTask.created_at.desc(),
        ).limit(limit)
        rows = (await session.execute(stmt)).scalars().all()
        items = [LeadTaskSchema.model_validate(r) for r in rows]
    return LeadTaskListResponse(items=items)

@router.post(
    "/api/v1/leads/{lead_id}/draft-email",
    response_model=LeadEmailDraftResponse,
)
async def draft_lead_email(
    lead_id: uuid.UUID,
    body: LeadEmailDraftRequest,
    current_user: User = Depends(get_current_user),
) -> LeadEmailDraftResponse:
    """Generate a personalised cold-email draft for one lead.

    The frontend opens the draft inline in the lead modal — the
    salesperson can copy the subject + body (or regenerate with a
    different tone) and paste into Gmail. Real send-via-Gmail
    ships once the OAuth connector lands.
    """
    async with session_factory() as session:
        lead, _search = await _authorise_lead_access(
            session, lead_id, current_user.id
        )
        user = await session.get(User, current_user.id)

    user_profile: dict[str, Any] = {}
    if user is not None:
        user_profile = {
            "display_name": user.display_name or user.first_name,
            "age_range": user.age_range,
            "gender": user.gender,
            "business_size": user.business_size,
            "profession": user.profession,
            "service_description": user.service_description,
            "home_region": user.home_region,
            "niches": list(user.niches or []),
            "language_code": user.language_code,
            "calendly_url": user.calendly_url,
            "icp_profile": user.icp_profile,
        }

    lead_payload = {
        "name": lead.name,
        "category": lead.category,
        "address": lead.address,
        "website": lead.website,
        "rating": lead.rating,
        "reviews_count": lead.reviews_count,
        "score_ai": lead.score_ai,
        "summary": lead.summary,
        "advice": lead.advice,
        "strengths": list(lead.strengths) if lead.strengths else None,
        "weaknesses": list(lead.weaknesses) if lead.weaknesses else None,
        "red_flags": list(lead.red_flags) if lead.red_flags else None,
    }

    analyzer = AIAnalyzer()

    # UI language (for the research headings shown to the user) vs
    # email language (per-draft override → UI language → ru).
    ui_lang = normalize_lang(user.language_code if user else None)
    email_language = normalize_lang(
        body.language or (user.language_code if user else None)
    )

    # Optional: deep research pass — fresh website fetch + Claude
    # extraction of notable facts. Threaded into ``extra_context``
    # so the existing email prompt naturally cites the lead's own
    # site instead of leaning on cached enrichment.
    notable_facts: list[str] = []
    recent_signal: str | None = None
    merged_extra = body.extra_context
    if body.deep_research:
        research = await analyzer.research_lead_for_outreach(
            lead_payload,
            user_profile=user_profile or None,
        )
        notable_facts = list(research.get("notable_facts") or [])
        recent_signal = research.get("recent_signal")
        opener = research.get("suggested_opener")
        research_block_parts: list[str] = []
        if notable_facts:
            research_block_parts.append(
                locale_pick(
                    ui_lang,
                    ru="Свежие факты с сайта (можно цитировать в opener):",
                    uk="Свіжі факти з сайту (можна цитувати в opener):",
                    en="Fresh facts from the site (quotable in the opener):",
                )
            )
            for fact in notable_facts:
                research_block_parts.append(f"- {fact}")
        if recent_signal:
            research_block_parts.append(
                locale_pick(
                    ui_lang,
                    ru=f"Recent signal (что-то новое у них): {recent_signal}",
                    uk=f"Recent signal (щось нове у них): {recent_signal}",
                    en=f"Recent signal (something new on their side): {recent_signal}",
                )
            )
        if opener:
            research_block_parts.append(
                locale_pick(
                    ui_lang,
                    ru=f"Подсказанный opener: {opener}",
                    uk=f"Підказаний opener: {opener}",
                    en=f"Suggested opener: {opener}",
                )
            )
        if research_block_parts:
            research_block = "\n".join(research_block_parts)
            merged_extra = (
                f"{body.extra_context}\n\n{research_block}"
                if body.extra_context
                else research_block
            )

    result = await analyzer.generate_cold_email(
        lead_payload,
        user_profile=user_profile or None,
        tone=body.tone,
        extra_context=merged_extra,
        language=email_language,
    )
    return LeadEmailDraftResponse(
        subject=result["subject"],
        body=result["body"],
        tone=result["tone"],
        notable_facts=notable_facts,
        recent_signal=recent_signal,
    )

@router.post(
    "/api/v1/leads/bulk-draft",
    response_model=BulkDraftEmailResponse,
)
async def bulk_draft_emails(
    body: BulkDraftEmailRequest,
    current_user: User = Depends(get_current_user),
) -> BulkDraftEmailResponse:
    """Generate cold-email drafts for up to 20 leads in one shot.

    The salesperson selects rows on /app/leads, hits "Написать
    всем", and gets back a stitched list ready for review. Per-
    lead errors don't take the whole batch down — failed entries
    come back with ``error`` populated.

    Concurrency is throttled (3 in-flight) so a 20-lead batch
    doesn't stampede Anthropic. Authorisation is per-lead: each
    lead must belong to a search the caller owns or is a member
    of via team.
    """
    async with session_factory() as session:
        user_profile: dict[str, Any] = {
            "display_name": current_user.display_name or current_user.first_name,
            "age_range": current_user.age_range,
            "gender": current_user.gender,
            "business_size": current_user.business_size,
            "profession": current_user.profession,
            "service_description": current_user.service_description,
            "home_region": current_user.home_region,
            "niches": list(current_user.niches or []),
            "language_code": current_user.language_code,
        }
        lead_rows = (
            (
                await session.execute(
                    select(Lead, SearchQuery)
                    .join(SearchQuery, SearchQuery.id == Lead.query_id)
                    .where(Lead.id.in_(list(body.lead_ids)))
                )
            )
            .all()
        )
        authorised: dict[uuid.UUID, Lead] = {}
        for lead, search in lead_rows:
            if search.user_id == current_user.id:
                authorised[lead.id] = lead
                continue
            if search.team_id is not None and (
                await membership(session, search.team_id, current_user.id)
            ):
                authorised[lead.id] = lead

    analyzer = AIAnalyzer()
    sem = asyncio.Semaphore(3)
    tone = (body.tone or "professional").strip().lower()
    # Per-batch email language: explicit override → UI language → ru.
    email_language = normalize_lang(
        body.language or current_user.language_code
    )

    async def _one(lead_id: uuid.UUID) -> BulkDraftEmailItem:
        lead = authorised.get(lead_id)
        if lead is None:
            return BulkDraftEmailItem(
                lead_id=lead_id, error="not authorised"
            )
        payload = {
            "name": lead.name,
            "category": lead.category,
            "address": lead.address,
            "website": lead.website,
            "rating": lead.rating,
            "reviews_count": lead.reviews_count,
            "score_ai": lead.score_ai,
            "summary": lead.summary,
            "advice": lead.advice,
            "strengths": list(lead.strengths) if lead.strengths else None,
            "weaknesses": list(lead.weaknesses) if lead.weaknesses else None,
            "red_flags": list(lead.red_flags) if lead.red_flags else None,
        }
        async with sem:
            try:
                result = await analyzer.generate_cold_email(
                    payload,
                    user_profile=user_profile,
                    tone=tone,
                    extra_context=body.extra_context,
                    language=email_language,
                )
                return BulkDraftEmailItem(
                    lead_id=lead_id,
                    subject=result.get("subject"),
                    body=result.get("body"),
                )
            except Exception as exc:  # noqa: BLE001
                logger.exception(
                    "bulk-draft: failed for lead %s", lead_id
                )
                return BulkDraftEmailItem(
                    lead_id=lead_id, error=str(exc)[:200]
                )

    items = await asyncio.gather(*(_one(lid) for lid in body.lead_ids))
    return BulkDraftEmailResponse(items=list(items))

@router.put("/api/v1/leads/{lead_id}/mark", response_model=LeadResponse)
async def set_lead_mark(
    lead_id: uuid.UUID,
    body: LeadMarkRequest,
    current_user: User = Depends(get_current_user),
) -> LeadResponse:
    """Set or clear the caller's private colour mark on a lead.

    Pass ``color: null`` to remove. The mark is only ever visible
    to the caller; teammates see their own marks (or none).
    """
    async with session_factory() as session:
        lead, _search = await _authorise_lead_access(
            session, lead_id, current_user.id
        )

        existing = (
            await session.execute(
                select(LeadMark)
                .where(LeadMark.user_id == current_user.id)
                .where(LeadMark.lead_id == lead_id)
                .limit(1)
            )
        ).scalar_one_or_none()

        color = (body.color or "").strip() or None
        if color is None:
            if existing is not None:
                await session.delete(existing)
            final_color: str | None = None
        elif existing is None:
            session.add(
                LeadMark(
                    user_id=current_user.id,
                    lead_id=lead_id,
                    color=color,
                )
            )
            final_color = color
        else:
            existing.color = color
            existing.updated_at = datetime.now(timezone.utc)
            final_color = color

        await session.commit()
        await session.refresh(lead)
        return to_lead_response(lead, final_color)

